"""
Script d'analyse de la documentation KBA105 pour extraction des tâches de maintenance préventive
"""

import zipfile
import re
import os
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Import des jours fériés tunisiens du projet18
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from logic.projet18 import JOURS_FERIES_TUNISIE_2026, is_jour_ferie, get_nom_jour_ferie

# Dictionnaires de traduction allemand -> français pour la maintenance
TRANSLATIONS = {
    # Maintenance
    'wartung': 'maintenance',
    'wartungs': 'maintenance',
    'wartungsplan': 'plan de maintenance',
    'wartungsanleitung': 'manuel de maintenance',
    'wartungsvorschrift': 'prescription de maintenance',
    
    # Vérification/Contrôle
    'prüfung': 'vérification',
    'prüf': 'vérification',
    'prüfplan': 'plan de vérification',
    'prüfanleitung': 'manuel de vérification',
    'prüfvorschrift': 'prescription de vérification',
    'kontrolle': 'contrôle',
    'kontroll': 'contrôle',
    
    # Nettoyage
    'reinigung': 'nettoyage',
    'reinigen': 'nettoyer',
    'rein': 'nettoyage',
    
    # Lubrification
    'schmierung': 'lubrification',
    'schmier': 'lubrification',
    'ölen': 'huiler',
    'öl': 'huile',
    
    # Inspection
    'inspektion': 'inspection',
    'inspekt': 'inspection',
    
    # Révision
    'revision': 'révision',
    'revis': 'révision',
    
    # Périodicité
    'täglich': 'quotidien',
    'tägliche': 'quotidienne',
    'tages': 'jour',
    'wöchentlich': 'hebdomadaire',
    'wöchentliche': 'hebdomadaire',
    'woche': 'semaine',
    'monatlich': 'mensuel',
    'monatliche': 'mensuelle',
    'monat': 'mois',
    'vierteljährlich': 'trimestriel',
    'vierteljährliche': 'trimestrielle',
    'quartal': 'trimestre',
    'jährlich': 'annuel',
    'jährliche': 'annuelle',
    'jahr': 'année',
    'stündlich': 'horaire',
    'stunde': 'heure',
}

# Patterns pour identifier la périodicité (améliorés avec plus de variantes)
PERIODICITY_PATTERNS = {
    'Quotidienne': [
        # Allemand
        r'täglich', r'tägliche', r'tages', r'taglich', r'tagliche',
        # Français
        r'quotidien', r'quotidienne', r'quotidiennement', r'jour', r'journalier', r'journalière',
        r'tous les jours', r'chaque jour', r'par jour', r'par jour', r'journellement',
        # Anglais
        r'daily', r'day', r'per day', r'each day', r'every day',
        # Autres
        r'stündlich', r'stunde', r'heure', r'horaire', r'à chaque heure',
        r'hebdomadaire', r'wöchentlich'  # Parfois utilisé pour quotidien dans certains contextes
    ],
    'Hebdomadaire': [
        # Allemand
        r'wöchentlich', r'wöchentliche', r'woche', r'wochen',
        # Français
        r'hebdomadaire', r'hebdomadairement', r'semaine', r'semainier', r'semainaire',
        r'toutes les semaines', r'chaque semaine', r'par semaine', r'une fois par semaine',
        r'hebdo', r'hebdomadaire',
        # Anglais
        r'weekly', r'week', r'per week', r'each week', r'every week', r'once a week',
        # Patterns avec nombres
        r'\d+\s*(semaine|week|woche)'
    ],
    'Mensuelle': [
        # Allemand
        r'monatlich', r'monatliche', r'monat', r'monate',
        # Français
        r'mensuel', r'mensuelle', r'mensuellement', r'mois', r'mensualité',
        r'tous les mois', r'chaque mois', r'par mois', r'une fois par mois',
        # Anglais
        r'monthly', r'month', r'per month', r'each month', r'every month', r'once a month',
        # Patterns avec nombres (1-2 mois = mensuelle)
        r'\b([12]|une|un|one|two)\s*(mois|month|monat)',
        r'tous les\s*\d*\s*mois'
    ],
    'Trimestrielle': [
        # Allemand
        r'vierteljährlich', r'vierteljährliche', r'quartal', r'vierteljahr',
        # Français
        r'trimestriel', r'trimestrielle', r'trimestriellement', r'trimestre',
        r'tous les trimestres', r'chaque trimestre', r'par trimestre', r'une fois par trimestre',
        r'quartal', r'quaterly',
        # Anglais
        r'quarterly', r'quarter', r'per quarter', r'each quarter', r'every quarter',
        # Patterns avec nombres (3-4 mois = trimestrielle)
        r'\b([34]|trois|four|three)\s*(mois|month|monat)',
        r'tous les\s*[34]\s*mois'
    ],
    'Annuelle': [
        # Allemand
        r'jährlich', r'jährliche', r'jahr', r'jahre', r'jahrlich',
        # Français
        r'annuel', r'annuelle', r'annuellement', r'année', r'annuellement',
        r'tous les ans', r'chaque année', r'par an', r'une fois par an', r'une fois par année',
        r'annuel', r'année',
        # Anglais
        r'yearly', r'year', r'per year', r'each year', r'every year', r'once a year', r'annually',
        # Patterns avec nombres (12 mois, 365 jours = annuelle)
        r'\b(12|douze|twelve)\s*(mois|month|monat)',
        r'\b(365|trois cent soixante-cinq|three hundred sixty-five)\s*(jour|day|tag)',
        r'tous les\s*\d*\s*ans'
    ]
}

# Patterns pour identifier les composants
COMPONENT_PATTERNS = [
    r'walze', r'walzen',  # rouleau
    r'druck', r'drucken',  # impression
    r'farbwerk', r'farb',  # unité d'impression
    r'zylinder',  # cylindre
    r'motor', r'antrieb',  # moteur
    r'pumpe',  # pompe
    r'ventil', r'ventile',  # valve
    r'filter', r'filterung',  # filtre
    r'lager', r'lagerung',  # palier/bearing
    r'kette', r'ketten',  # chaîne
    r'riemen',  # courroie
    r'getriebe',  # transmission
    r'kühlung', r'kühler',  # refroidissement
    r'luft', r'luftung',  # air
    r'öl', r'ölsystem',  # huile
    r'wasser', r'wassersystem',  # eau
    r'elektrik', r'elektrisch',  # électrique
    r'hydraulik', r'hydraulisch',  # hydraulique
    r'pneumatik', r'pneumatisch',  # pneumatique
    r'steuerung', r'steuer',  # contrôle/commande
    r'sicherheit', r'sicher',  # sécurité
    r'brems', r'bremse',  # frein
    r'kupplung',  # embrayage
]

# Patterns pour identifier les types de documents
DOCUMENT_TYPES = {
    'Manuel': [r'handbuch', r'manual', r'anleitung', r'manuel'],
    'Plan': [r'plan', r'planung', r'schedule'],
    'Procédure': [r'vorschrift', r'procedure', r'prozess', r'procédure'],
    'Liste': [r'liste', r'list', r'checkliste', r'checklist'],
    'Guide': [r'guide', r'führung', r'anleitung'],
    'Spécification': [r'spezifikation', r'specification', r'spec'],
    'Instruction': [r'instruction', r'anweisung', r'instruction'],
}


def translate_to_french(text):
    """Traduit les termes allemands en français"""
    text_lower = text.lower()
    for de, fr in TRANSLATIONS.items():
        if de in text_lower:
            text = text.replace(de, fr)
            text = text.replace(de.capitalize(), fr.capitalize())
    return text


def extract_periodicity(filename):
    """Extrait la périodicité depuis le nom de fichier"""
    filename_lower = filename.lower()
    
    for periodicity, patterns in PERIODICITY_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, filename_lower, re.IGNORECASE):
                return periodicity
    
    # Par défaut, si aucun pattern trouvé
    return 'Non spécifiée'


def extract_periodicity_from_tables(content):
    """Extrait la périodicité depuis les tableaux HTML en analysant la colonne Intervalle"""
    
    # Chercher tous les tableaux
    tables = re.findall(r'<table[^>]*>(.*?)</table>', content, re.DOTALL | re.IGNORECASE)
    
    periodicity_found = None
    
    for table in tables:
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', table, re.DOTALL | re.IGNORECASE)
        
        if len(rows) < 2:  # Pas assez de lignes pour avoir des données
            continue
        
        # Trouver l'index de la colonne "Intervalle"
        intervalle_col_index = None
        header_row = rows[0]
        header_cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', header_row, re.DOTALL | re.IGNORECASE)
        
        for i, cell in enumerate(header_cells):
            cell_text = re.sub(r'<[^>]+>', ' ', cell).strip().lower()
            if 'intervalle' in cell_text or 'intervall' in cell_text:
                intervalle_col_index = i
                break
        
        # Si pas trouvé dans l'en-tête, chercher dans les premières lignes de données
        if intervalle_col_index is None:
            # Analyser les premières lignes pour trouver où sont les valeurs d'intervalle
            for row_idx in range(1, min(5, len(rows))):
                cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', rows[row_idx], re.DOTALL | re.IGNORECASE)
                # Chercher une cellule qui contient "million" ou des valeurs numériques avec unités
                for i, cell in enumerate(cells):
                    cell_text = re.sub(r'<[^>]+>', ' ', cell).strip().lower()
                    if 'million' in cell_text or re.search(r'\d+\s*(?:h|heure|hour|stunde|mois|month|monat|semaine|week|woche|an|année|year|jahr)', cell_text):
                        intervalle_col_index = i
                        break
                if intervalle_col_index is not None:
                    break
        
        # Si toujours pas trouvé, utiliser la dernière colonne (souvent l'intervalle)
        if intervalle_col_index is None and len(rows) > 1:
            first_data_row = rows[1]
            cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', first_data_row, re.DOTALL | re.IGNORECASE)
            if len(cells) > 0:
                # Vérifier si la dernière colonne contient des valeurs d'intervalle
                last_cell = re.sub(r'<[^>]+>', ' ', cells[-1]).strip().lower()
                if 'million' in last_cell or re.search(r'\d+\s*(?:h|heure|hour|stunde|mois|month|monat|semaine|week|woche|an|année|year|jahr)', last_cell):
                    intervalle_col_index = len(cells) - 1
        
        # Analyser chaque ligne de données pour trouver la périodicité
        for row in rows[1:]:  # Sauter l'en-tête
            cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', row, re.DOTALL | re.IGNORECASE)
            
            # Essayer toutes les colonnes si l'index n'est pas trouvé
            cols_to_check = [intervalle_col_index] if intervalle_col_index is not None else range(len(cells))
            
            for col_idx in cols_to_check:
                if col_idx is None or col_idx >= len(cells):
                    continue
                    
                intervalle_cell = cells[col_idx]
                intervalle_text = re.sub(r'<[^>]+>', ' ', intervalle_cell).strip()
                
                if not intervalle_text or len(intervalle_text) < 2:
                    continue
                    
                intervalle_lower = intervalle_text.lower()
                
                # Analyser la valeur d'intervalle
                # "X millions de feuilles" - PRIORITÉ ABSOLUE
                million_match = re.search(r'(\d+)\s*millions?\s*(?:de\s*)?(?:feuilles?|sheets?|bogen)', intervalle_lower, re.IGNORECASE)
                if million_match:
                    try:
                        value = int(million_match.group(1))
                        # Conversion basée sur la production typique d'une presse offset
                        # Une presse KBA105 peut produire 10-15 millions de feuilles par an
                        # Donc les intervalles en millions de feuilles indiquent la fréquence
                        if value >= 15:
                            return 'Annuelle'  # 15+ millions = maintenance annuelle (après 1 an de production)
                        elif value >= 10:
                            return 'Annuelle'  # 10-14 millions = annuelle
                        elif value >= 5:
                            return 'Annuelle'  # 5-9 millions = annuelle (environ 6 mois à 1 an)
                        elif value >= 2:
                            return 'Trimestrielle'  # 2-4 millions = trimestrielle (environ 2-4 mois)
                        elif value == 1:
                            return 'Mensuelle'  # 1 million = mensuelle (environ 1 mois de production)
                    except:
                        pass
                
                # Patterns textuels directs (plus fiables que les heures)
                if any(word in intervalle_lower for word in ['quotidien', 'daily', 'täglich', 'jour', 'tous les jours']):
                    periodicity_found = 'Quotidienne'
                    continue
                if any(word in intervalle_lower for word in ['hebdomadaire', 'weekly', 'wöchentlich', 'semaine', 'toutes les semaines']):
                    periodicity_found = 'Hebdomadaire'
                    continue
                if any(word in intervalle_lower for word in ['mensuel', 'monthly', 'monatlich', 'mois', 'tous les mois']):
                    periodicity_found = 'Mensuelle'
                    continue
                if any(word in intervalle_lower for word in ['trimestriel', 'quarterly', 'vierteljährlich', 'trimestre']):
                    periodicity_found = 'Trimestrielle'
                    continue
                if any(word in intervalle_lower for word in ['annuel', 'yearly', 'jährlich', 'année', 'jahr', 'tous les ans']):
                    periodicity_found = 'Annuelle'
                    continue
                
                # "X heures" ou "X h" - seulement si c'est dans un contexte d'intervalle de maintenance
                # Ignorer les valeurs de vitesse (ex: "5000 Bg/h") ou autres contextes
                if 'million' not in intervalle_lower and 'feuille' not in intervalle_lower:
                    # Vérifier que ce n'est pas une vitesse de production
                    if not re.search(r'\d+\s*(?:bg|bogen|feuille|sheet)\s*/\s*h', intervalle_lower, re.IGNORECASE):
                        hour_match = re.search(r'\b(\d+)\s*(?:h|heure|hour|stunde)\b', intervalle_lower, re.IGNORECASE)
                        if hour_match:
                            try:
                                hours = int(hour_match.group(1))
                                # Conversion heures -> périodicité (seulement si valeur raisonnable pour maintenance)
                                # Les valeurs très petites (< 1h) ou très grandes sont suspectes
                                if 1 <= hours <= 8:
                                    # Seulement si c'est vraiment un intervalle de maintenance, pas une vitesse
                                    if 'intervalle' in intervalle_lower or 'intervall' in intervalle_lower:
                                        periodicity_found = 'Quotidienne'
                                elif 9 <= hours <= 168:  # 1 semaine = 168h
                                    periodicity_found = 'Hebdomadaire'
                                elif 169 <= hours <= 720:  # 1 mois = ~720h (30 jours * 24h)
                                    periodicity_found = 'Mensuelle'
                                elif 721 <= hours <= 2160:  # 3 mois = ~2160h (90 jours * 24h)
                                    periodicity_found = 'Trimestrielle'
                                elif hours > 2160:
                                    periodicity_found = 'Annuelle'
                            except:
                                pass
    
    return periodicity_found if periodicity_found else 'Non spécifiée'


def extract_periodicity_from_content(content_lower):
    """Extrait la périodicité depuis le contenu HTML avec analyse approfondie"""
    
    # Extraire le texte visible (sans balises HTML) pour une meilleure analyse
    text_content = re.sub(r'<script[^>]*>.*?</script>', '', content_lower, flags=re.DOTALL | re.IGNORECASE)
    text_content = re.sub(r'<style[^>]*>.*?</style>', '', text_content, flags=re.DOTALL | re.IGNORECASE)
    
    # Extraire spécifiquement le contenu des tableaux (où la périodicité est souvent indiquée)
    tables = re.findall(r'<table[^>]*>(.*?)</table>', content_lower, re.DOTALL | re.IGNORECASE)
    table_text = ' '.join(tables) if tables else ''
    
    # Analyser les lignes de tableau pour trouver la colonne "Intervalle"
    intervalle_values = []
    for table in tables:
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', table, re.DOTALL | re.IGNORECASE)
        for row in rows:
            cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', row, re.DOTALL | re.IGNORECASE)
            # Chercher la colonne qui contient "Intervalle" ou la dernière colonne (souvent l'intervalle)
            for i, cell in enumerate(cells):
                cell_clean = re.sub(r'<[^>]+>', ' ', cell).strip()
                # Si la cellule contient "intervalle" ou est dans une colonne d'intervalle
                if 'intervalle' in cell_clean.lower() or 'intervall' in cell_clean.lower():
                    # La valeur d'intervalle est souvent dans la même ligne, colonne suivante ou dernière
                    if i + 1 < len(cells):
                        intervalle_cell = re.sub(r'<[^>]+>', ' ', cells[i + 1]).strip()
                        if intervalle_cell and intervalle_cell.lower() not in ['intervalle', 'intervall', 'h', '(h)']:
                            intervalle_values.append(intervalle_cell)
                    # Ou chercher dans toutes les cellules de la ligne
                    for cell_val in cells:
                        cell_val_clean = re.sub(r'<[^>]+>', ' ', cell_val).strip()
                        if cell_val_clean and 'million' in cell_val_clean.lower():
                            intervalle_values.append(cell_val_clean)
    
    # Extraire le texte des cellules de tableau
    cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', content_lower, re.DOTALL | re.IGNORECASE)
    cell_text = ' '.join(cells)
    
    # Ajouter les valeurs d'intervalle trouvées
    if intervalle_values:
        cell_text = ' '.join(intervalle_values) + ' ' + cell_text
    
    # Combiner tout le texte pour analyse
    all_text = text_content + ' ' + table_text + ' ' + cell_text
    all_text = re.sub(r'<[^>]+>', ' ', all_text)  # Retirer toutes les balises HTML restantes
    all_text = re.sub(r'\s+', ' ', all_text)  # Normaliser les espaces
    
    # Chercher les patterns de périodicité dans le texte combiné
    # Mais être plus sélectif : ignorer les références aux heures dans d'autres contextes
    periodicity_scores = {}
    
    # Filtrer le texte pour exclure les vitesses de production et autres contextes non-maintenance
    filtered_text = all_text
    # Retirer les références aux vitesses (ex: "5000 Bg/h")
    filtered_text = re.sub(r'\d+\s*(?:bg|bogen|feuille|sheet)\s*/\s*h', '', filtered_text, flags=re.IGNORECASE)
    
    for periodicity, patterns in PERIODICITY_PATTERNS.items():
        score = 0
        for pattern in patterns:
            # Ignorer les patterns d'heures génériques qui peuvent être dans d'autres contextes
            if 'h' in pattern or 'heure' in pattern or 'hour' in pattern or 'stunde' in pattern:
                # Seulement chercher dans un contexte de maintenance/intervalle
                if 'intervalle' not in filtered_text.lower() and 'maintenance' not in filtered_text.lower() and 'wartung' not in filtered_text.lower():
                    continue
            
            # Chercher dans le texte filtré
            matches_all = re.findall(pattern, filtered_text, re.IGNORECASE)
            # Chercher spécifiquement dans les cellules de tableau (plus fiable)
            matches_cells = re.findall(pattern, cell_text, re.IGNORECASE)
            
            if matches_all:
                score += len(matches_all)
            if matches_cells:
                score += len(matches_cells) * 3  # Les cellules de tableau ont un poids plus élevé
            
            # Patterns plus spécifiques ont un poids plus élevé
            if any(p in pattern for p in ['tous les', 'chaque', 'every', 'each', 'per']):
                score += 2
        
        if score > 0:
            periodicity_scores[periodicity] = score
    
    # Retourner la périodicité avec le score le plus élevé
    if periodicity_scores:
        # Trier par score décroissant
        sorted_periodicities = sorted(periodicity_scores.items(), key=lambda x: x[1], reverse=True)
        return sorted_periodicities[0][0]
    
    # Si aucune périodicité trouvée, chercher des indices contextuels dans les tableaux
    # Analyser les intervalles en "millions de feuilles" ou heures
    
    # D'abord, analyser les valeurs d'intervalle extraites des tableaux
    if intervalle_values:
        for intervalle_val in intervalle_values:
            intervalle_lower = intervalle_val.lower()
            
            # Analyser les "millions de feuilles"
            million_match = re.search(r'(\d+)\s*millions?\s*(?:de\s*)?(?:feuilles?|sheets?|bogen)', intervalle_lower, re.IGNORECASE)
            if million_match:
                try:
                    value = int(million_match.group(1))
                    # Conversion précise basée sur la production
                    if value >= 15:
                        return 'Annuelle'
                    elif value >= 10:
                        return 'Annuelle'
                    elif value >= 5:
                        return 'Annuelle'
                    elif value >= 2:
                        return 'Trimestrielle'  # 2-4 millions
                    elif value == 1:
                        return 'Mensuelle'  # 1 million
                except:
                    pass
            
            # Si "million" présent sans nombre spécifique
            if 'million' in intervalle_lower and ('feuille' in intervalle_lower or 'sheet' in intervalle_lower or 'bogen' in intervalle_lower):
                return 'Annuelle'
    
    # Chercher dans tout le texte les patterns "millions de feuilles"
    million_patterns = [
        (r'(\d+)\s*millions?\s*(?:de\s*)?(?:feuilles?|sheets?|bogen)', 'Annuelle'),
        (r'(\d+)\s*million\s*(?:de\s*)?(?:feuilles?|sheets?|bogen)', 'Annuelle'),
    ]
    
    for pattern, periodicity in million_patterns:
        matches = re.findall(pattern, cell_text, re.IGNORECASE)
        if matches:
            for match in matches:
                if isinstance(match, tuple):
                    value_str = match[0] if match[0].isdigit() else '0'
                else:
                    value_str = str(match) if str(match).isdigit() else '0'
                
                try:
                    value = int(value_str)
                    # Conversion basée sur la production typique
                    if value >= 15:
                        return 'Annuelle'
                    elif value >= 10:
                        return 'Annuelle'
                    elif value >= 5:
                        return 'Annuelle'
                    elif value >= 2:
                        return 'Trimestrielle'
                    elif value == 1:
                        return 'Mensuelle'
                except:
                    pass
            
            # Si pattern trouvé mais pas de nombre, retourner annuel par défaut
            return 'Annuelle'
    
    # Chercher des nombres avec des unités de temps explicites
    time_patterns = [
        (r'\b(\d+)\s*(jour|day|tag)\b', 'Quotidienne'),
        (r'\b(\d+)\s*(semaine|week|woche)\b', 'Hebdomadaire'),
        (r'\b([1-2]|une|un|one|two)\s*(mois|month|monat)\b', 'Mensuelle'),
        (r'\b([3-4]|trois|four|three)\s*(mois|month|monat)\b', 'Trimestrielle'),
        (r'\b(12|douze|twelve)\s*(mois|month|monat)\b', 'Annuelle'),
        (r'\b(\d+)\s*(an|année|year|jahr)\b', 'Annuelle'),
    ]
    
    # Chercher d'abord dans les cellules de tableau (plus fiable)
    for pattern, periodicity in time_patterns:
        if re.search(pattern, cell_text, re.IGNORECASE):
            return periodicity
    
    # Analyser les intervalles en heures (seulement si pas de "million" trouvé et contexte approprié)
    if 'million' not in cell_text.lower() and 'feuille' not in cell_text.lower():
        # Ignorer les vitesses de production (ex: "5000 Bg/h")
        if not re.search(r'\d+\s*(?:bg|bogen|feuille|sheet)\s*/\s*h', cell_text, re.IGNORECASE):
            hour_patterns = [
                # Seulement si c'est dans un contexte d'intervalle de maintenance
                (r'intervalle.*?\b([1-8])\s*(h|heure|hour|stunde)\b', 'Quotidienne'),
                (r'\b([1-7]\d|8[0-9]|9[0-9]|1[0-6]\d)\s*(h|heure|hour|stunde)\b.*?(?:intervalle|maintenance|wartung)', 'Hebdomadaire'),
                (r'\b([1-7]\d{2}|8[0-9]{2}|9[0-9]{2})\s*(h|heure|hour|stunde)\b.*?(?:intervalle|maintenance|wartung)', 'Mensuelle'),
                (r'\b([1-9]\d{3,})\s*(h|heure|hour|stunde)\b.*?(?:intervalle|maintenance|wartung)', 'Annuelle'),
            ]
            
            for pattern, periodicity in hour_patterns:
                if re.search(pattern, cell_text, re.IGNORECASE):
                    return periodicity
    
    # Puis dans tout le texte (sans les patterns d'heures pour éviter les faux positifs)
    for pattern, periodicity in time_patterns:
        if re.search(pattern, all_text, re.IGNORECASE):
            return periodicity
    
    return 'Non spécifiée'


def extract_component_from_content(content_lower):
    """Extrait le composant depuis le contenu HTML"""
    components_found = []
    
    for pattern in COMPONENT_PATTERNS:
        matches = re.findall(pattern, content_lower, re.IGNORECASE)
        if matches:
            components_found.extend(matches)
    
    if components_found:
        # Retourner le premier composant trouvé, traduit en français
        component = components_found[0]
        return translate_to_french(component).capitalize()
    
    return 'Machine complète'


def extract_component(filename):
    """Extrait le composant concerné depuis le nom de fichier"""
    filename_lower = filename.lower()
    components_found = []
    
    for pattern in COMPONENT_PATTERNS:
        matches = re.findall(pattern, filename_lower, re.IGNORECASE)
        if matches:
            components_found.extend(matches)
    
    if components_found:
        # Retourner le premier composant trouvé, traduit en français
        component = components_found[0]
        return translate_to_french(component).capitalize()
    
    return 'Machine complète'


def extract_document_type(filename):
    """Extrait le type de document depuis le nom de fichier"""
    filename_lower = filename.lower()
    
    for doc_type, patterns in DOCUMENT_TYPES.items():
        for pattern in patterns:
            if re.search(pattern, filename_lower, re.IGNORECASE):
                return doc_type
    
    return 'Documentation'


def is_preventive_maintenance(filename):
    """Détermine si le fichier concerne la maintenance préventive"""
    filename_lower = filename.lower()
    
    # Mots-clés de maintenance préventive
    preventive_keywords = [
        'wartung', 'wartungs', 'maintenance', 'mainten',
        'prüfung', 'prüf', 'prüfplan', 'vérification', 'verif',
        'inspektion', 'inspection', 'insp',
        'revision', 'révision', 'revis',
        'wartungsplan', 'plan de maintenance',
        'wartungsvorschrift', 'prescription',
        'wartungsanleitung', 'manuel de maintenance',
        'wartungskalender', 'calendrier de maintenance',
        'wartungsliste', 'liste de maintenance',
        'wartungsprogramm', 'programme de maintenance',
    ]
    
    # Mots-clés à exclure (maintenance corrective)
    corrective_keywords = [
        'reparatur', 'réparation', 'repair',
        'störung', 'panne', 'defekt', 'défaut', 'fault',
        'notfall', 'urgence', 'emergency',
        'ausfall', 'échec', 'failure',
    ]
    
    # Vérifier les mots-clés correctifs d'abord
    for keyword in corrective_keywords:
        if keyword in filename_lower:
            return False
    
    # Vérifier les mots-clés préventifs
    for keyword in preventive_keywords:
        if keyword in filename_lower:
            return True
    
    return False


def extract_task_label(filename):
    """Extrait le libellé de la tâche depuis le nom de fichier"""
    # Nettoyer le nom de fichier
    name = filename
    
    # Retirer l'extension
    name = os.path.splitext(name)[0]
    
    # Retirer les préfixes communs (ex: kba105_, kba105-)
    name = re.sub(r'^kba105[_-]?', '', name, flags=re.IGNORECASE)
    
    # Remplacer les underscores et tirets par des espaces
    name = re.sub(r'[_-]', ' ', name)
    
    # Traduire en français
    name = translate_to_french(name)
    
    # Capitaliser chaque mot
    words = name.split()
    name = ' '.join(word.capitalize() for word in words)
    
    return name


def analyze_kba105_archive(zip_path):
    """Analyse l'archive kba105.zip ou les fichiers extraits et extrait les tâches de maintenance préventive"""
    
    # Convertir zip_path en Path si c'est une chaîne
    zip_path_obj = Path(zip_path) if isinstance(zip_path, str) else zip_path
    
    # script_dir est le dossier kba105_analysis (parent du dossier kba105 qui contient le zip)
    script_dir = zip_path_obj.parent.parent if "kba105" in str(zip_path_obj.parent) else zip_path_obj.parent
    extracted_dir = script_dir / "kba105"
    txt_file_path = extracted_dir / "kba105.txt"
    
    tasks = []
    example_files = []
    
    # Chercher d'abord le dossier non compressé dans kba105_analysis
    # Essayer plusieurs variantes du nom du dossier
    maintenance_folder_local = None
    possible_names = [
        "3543 fichiers de  maintenance kba 105",
        "3543 fichiers de maintenance kba 105",
        "maintenance kba 105",
        "kba105_files",
        "fichiers maintenance kba105"
    ]
    
    for name in possible_names:
        test_path = script_dir / name
        if test_path.exists():
            maintenance_folder_local = test_path
            print(f"Dossier trouve avec le nom: {name}")
            break
    
    maintenance_folder_onedrive = Path(r"C:\Users\pack2\OneDrive\Bureau\projet 16 GMAO\maintenance kba 105")
    zip_ref = None
    
    # Priorité au dossier local dans kba105_analysis
    maintenance_folder = None
    if maintenance_folder_local and maintenance_folder_local.exists():
        maintenance_folder = maintenance_folder_local
        print(f"Dossier maintenance local trouve: {maintenance_folder}")
    elif maintenance_folder_onedrive.exists():
        maintenance_folder = maintenance_folder_onedrive
        print(f"Dossier maintenance OneDrive trouve: {maintenance_folder}")
    
    if maintenance_folder:
        # Lister tous les fichiers HTML du dossier
        html_files_in_folder = list(maintenance_folder.glob("*.html"))
        print(f"Fichiers HTML trouves dans le dossier: {len(html_files_in_folder)}")
        
        # Créer une liste avec les chemins complets pour lecture directe
        file_list_with_paths = [(f.name, str(f)) for f in html_files_in_folder]
        file_list = [f[0] for f in file_list_with_paths]
        
        print(f"Analyse de {len(file_list)} fichiers HTML depuis le dossier...")
    elif os.path.exists(zip_path):
        print(f"Analyse de l'archive: {zip_path}")
        
        zip_ref = zipfile.ZipFile(zip_path, 'r')
        # Lister TOUS les fichiers HTML dans l'archive directement
        all_files = zip_ref.namelist()
        html_files_in_zip = [f for f in all_files if f.lower().endswith('.html')]
        
        print(f"Fichiers HTML trouves dans l'archive: {len(html_files_in_zip)}")
        
        # Chercher le fichier kba105.txt dans l'archive
        txt_file = None
        for file in all_files:
            if 'kba105.txt' in file.lower():
                txt_file = file
                break
        
        if txt_file:
            print(f"Lecture du fichier liste: {txt_file}")
            with zip_ref.open(txt_file) as f:
                file_list_raw = f.read().decode('utf-8', errors='ignore').splitlines()
            
            # Extraire les noms de fichiers depuis les chemins complets
            file_list_from_txt = []
            file_list_with_paths = []
            for line in file_list_raw:
                line = line.strip()
                if not line:
                    continue
                # Retirer les guillemets si présents
                line = line.strip('"')
                # Extraire le nom de fichier depuis le chemin complet
                filename = os.path.basename(line)
                if filename.lower().endswith('.html'):
                    file_list_from_txt.append(filename)
                    file_list_with_paths.append((filename, line))
            
            print(f"Fichiers HTML listes dans kba105.txt: {len(file_list_from_txt)}")
            # Utiliser la liste complète depuis kba105.txt
            file_list = file_list_from_txt
        else:
            # Si pas de fichier txt, utiliser tous les fichiers HTML de l'archive
            print("Aucun fichier kba105.txt trouve, analyse de tous les fichiers HTML de l'archive")
            file_list = [os.path.basename(f) for f in html_files_in_zip]
            file_list_with_paths = [(os.path.basename(f), f) for f in html_files_in_zip]
    else:
        raise FileNotFoundError(f"Ni le dossier maintenance ni l'archive {zip_path} n'existent.")
    
    # Analyser tous les fichiers HTML pour trouver les tâches de maintenance préventive
    print("Analyse du contenu HTML des fichiers pour identifier les tâches de maintenance...")
    html_files = [f for f in file_list if f.lower().endswith('.html')]
    
    print(f"Analyse de {len(html_files)} fichiers HTML...")
    
    # zip_ref est déjà ouvert plus haut, ne pas le rouvrir
    
    # Analyser chaque fichier HTML
    preventive_keywords_html = [
            'maintenance', 'wartung', 'wartungs',
            'vérification', 'prüfung', 'prüf',
            'inspection', 'inspektion',
            'nettoyage', 'reinigung', 'reinigen',
            'lubrification', 'schmierung', 'schmier',
            'quotidien', 'täglich', 'daily',
            'hebdomadaire', 'wöchentlich', 'weekly',
            'mensuel', 'monatlich', 'monthly',
            'trimestriel', 'vierteljährlich', 'quarterly',
            'annuel', 'jährlich', 'yearly',
    ]
    
    corrective_keywords_html = [
            'réparation', 'reparatur', 'repair',
            'panne', 'störung', 'defekt', 'fault',
            'urgence', 'notfall', 'emergency',
    ]
    
    for i, filename in enumerate(html_files):
        if not filename.strip():
            continue
        
        filename = filename.strip()
        content = None
        
        # Chercher le fichier dans file_list_with_paths
        file_path = None
        if 'file_list_with_paths' in locals():
            for fname, fpath in file_list_with_paths:
                if fname == filename:
                    file_path = fpath
                    break
        
        # Lire le contenu depuis le chemin trouvé
        if file_path and os.path.exists(file_path):
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
            except Exception as e:
                pass
        
        # Si pas trouvé, chercher dans l'archive ZIP
        if not content and zip_ref:
            filename_base = os.path.basename(filename)
            for zip_file in zip_ref.namelist():
                zip_file_base = os.path.basename(zip_file)
                if filename_base == zip_file_base or filename_base.lower() == zip_file_base.lower():
                    try:
                        with zip_ref.open(zip_file) as f:
                            content = f.read().decode('utf-8', errors='ignore')
                        break
                    except Exception as e:
                        pass
        
        # Si toujours pas de contenu, passer au fichier suivant
        if not content:
            continue
        
        try:
            
            content_lower = content.lower()
            
            # Extraire le titre et h2 depuis le HTML
            title_match = re.search(r'<title>(.*?)</title>', content, re.IGNORECASE | re.DOTALL)
            h2_match = re.search(r'<h2>(.*?)</h2>', content, re.IGNORECASE | re.DOTALL)
            title = title_match.group(1).strip() if title_match else ''
            h2 = h2_match.group(1).strip() if h2_match else ''
            
            # Retirer les balises HTML
            title = re.sub(r'<[^>]+>', '', title)
            h2 = re.sub(r'<[^>]+>', '', h2)
            
            # Vérifier les mots-clés correctifs (réparation, panne, etc.)
            # Liste plus restrictive pour éviter d'exclure trop de fichiers
            has_corrective = any(kw in content_lower for kw in [
                'réparation d\'urgence', 'reparatur', 'repair', 
                'panne', 'störung', 'defekt', 'défaut',
                'notfall', 'urgence', 'emergency repair'
            ])
            
            # Puisque tous les fichiers sont dans un dossier "maintenance kba 105",
            # considérer comme maintenance préventive sauf si explicitement correctif
            # Inclure tous les fichiers qui ont du contenu technique
            has_technical_content = any(kw in content_lower for kw in [
                'machine', 'presse', 'groupe', 'système', 'system', 'composant', 'component',
                'kba105', 'rapida', 'walze', 'zylinder', 'motor', 'pumpe', 'ventil', 'filter',
                'maintenance', 'wartung', 'vérification', 'prüfung', 'inspection', 'inspektion'
            ])
            
            # Inclure si : pas correctif ET (a du contenu technique OU est dans dossier maintenance)
            if not has_corrective and (has_technical_content or True):  # Inclure tous sauf correctifs explicites
                # Extraire le libellé depuis le titre ou h2, ou depuis le contenu
                task_label = h2 if h2 else title
                if not task_label.strip():
                    # Essayer d'extraire depuis d'autres balises HTML
                    h1_match = re.search(r'<h1>(.*?)</h1>', content, re.IGNORECASE | re.DOTALL)
                    h3_match = re.search(r'<h3>(.*?)</h3>', content, re.IGNORECASE | re.DOTALL)
                    if h1_match:
                        task_label = re.sub(r'<[^>]+>', '', h1_match.group(1)).strip()
                    elif h3_match:
                        task_label = re.sub(r'<[^>]+>', '', h3_match.group(1)).strip()
                
                if not task_label.strip():
                    task_label = extract_task_label(filename)
                
                # Extraire les informations
                # Analyser d'abord les tableaux pour extraire la périodicité depuis la colonne "Intervalle"
                periodicity = extract_periodicity_from_tables(content)
                if periodicity == 'Non spécifiée':
                    periodicity = extract_periodicity_from_content(content_lower)
                if periodicity == 'Non spécifiée':
                    periodicity = extract_periodicity(filename)
                
                component = extract_component_from_content(content_lower)
                if component == 'Machine complète':
                    component = extract_component(filename)
                
                doc_type = extract_document_type(filename)
                
                tasks.append({
                    'numero': len(tasks) + 1,
                    'libelle': task_label.strip(),
                    'periodicite': periodicity,
                    'type_document': doc_type,
                    'fichier_source': filename,
                    'composant': component
                })
                
                if len(example_files) < 20:
                    example_files.append(filename)
                    
        except Exception as e:
            # Ignorer les erreurs de lecture
            pass
            
            if (i + 1) % 500 == 0:
                print(f"  Traite {i + 1}/{len(html_files)} fichiers HTML...")
    
    # Fermer l'archive
    if zip_ref:
        zip_ref.close()
    
    print(f"\n[OK] {len(tasks)} tâches de maintenance préventive identifiées")
    
    return tasks, example_files


def generate_calendar_with_holidays(year=2026):
    """Génère un calendrier avec les jours fériés tunisiens"""
    calendar = {}
    
    # Créer un dictionnaire des jours fériés
    holidays_dict = {}
    for holiday in JOURS_FERIES_TUNISIE_2026:
        if holiday.year == year:
            holidays_dict[holiday.date()] = get_nom_jour_ferie(holiday)
    
    # Générer toutes les dates de l'année
    start_date = datetime(year, 1, 1)
    end_date = datetime(year, 12, 31)
    
    current_date = start_date
    while current_date <= end_date:
        date_key = current_date.date()
        calendar[date_key] = {
            'is_holiday': is_jour_ferie(current_date),
            'holiday_name': holidays_dict.get(date_key, ''),
            'weekday': current_date.strftime('%A'),
            'week_number': current_date.isocalendar()[1]
        }
        current_date += timedelta(days=1)
    
    return calendar


def calculate_next_dates(task_periodicity, start_date=datetime(2026, 1, 1), num_occurrences=12):
    """Calcule les prochaines dates d'exécution d'une tâche selon sa périodicité"""
    dates = []
    current_date = start_date
    
    if task_periodicity == 'Quotidienne':
        for i in range(num_occurrences):
            dates.append(current_date + timedelta(days=i))
    
    elif task_periodicity == 'Hebdomadaire':
        for i in range(num_occurrences):
            dates.append(current_date + timedelta(weeks=i))
    
    elif task_periodicity == 'Mensuelle':
        for i in range(num_occurrences):
            # Ajouter i mois
            month = start_date.month + i
            year = start_date.year
            while month > 12:
                month -= 12
                year += 1
            # Utiliser le même jour du mois si possible
            try:
                dates.append(datetime(year, month, start_date.day))
            except ValueError:
                # Si le jour n'existe pas (ex: 31 février), utiliser le dernier jour du mois
                from calendar import monthrange
                last_day = monthrange(year, month)[1]
                dates.append(datetime(year, month, last_day))
    
    elif task_periodicity == 'Trimestrielle':
        for i in range(num_occurrences):
            month = start_date.month + (i * 3)
            year = start_date.year
            while month > 12:
                month -= 12
                year += 1
            try:
                dates.append(datetime(year, month, start_date.day))
            except ValueError:
                from calendar import monthrange
                last_day = monthrange(year, month)[1]
                dates.append(datetime(year, month, last_day))
    
    elif task_periodicity == 'Annuelle':
        for i in range(num_occurrences):
            dates.append(datetime(start_date.year + i, start_date.month, start_date.day))
    
    else:
        # Non spécifiée - pas de dates
        pass
    
    return dates


def create_excel_report(tasks, output_path):
    """Crée le fichier Excel avec les tâches et le calendrier de prévision"""
    
    print(f"\nGénération du fichier Excel: {output_path}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tâches Maintenance Préventive KBA105"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # En-têtes
    headers = [
        'N°',
        'Libellé de la tâche',
        'Périodicité',
        'Type de document',
        'Nom du fichier source',
        'Composant concerné'
    ]
    
    # Écrire les en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Écrire les données
    for row_num, task in enumerate(tasks, 2):
        ws.cell(row=row_num, column=1, value=task['numero']).border = border
        ws.cell(row=row_num, column=2, value=task['libelle']).border = border
        ws.cell(row=row_num, column=2).alignment = wrap_alignment
        ws.cell(row=row_num, column=3, value=task['periodicite']).border = border
        ws.cell(row=row_num, column=3).alignment = center_alignment
        ws.cell(row=row_num, column=4, value=task['type_document']).border = border
        ws.cell(row=row_num, column=4).alignment = center_alignment
        ws.cell(row=row_num, column=5, value=task['fichier_source']).border = border
        ws.cell(row=row_num, column=5).alignment = wrap_alignment
        ws.cell(row=row_num, column=6, value=task['composant']).border = border
        ws.cell(row=row_num, column=6).alignment = wrap_alignment
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 30
    
    # Ajouter une feuille pour le calendrier de prévision
    ws_calendar = wb.create_sheet("Calendrier Prévision 2026")
    
    # Générer le calendrier
    calendar = generate_calendar_with_holidays(2026)
    
    # En-têtes du calendrier
    calendar_headers = ['Date', 'Jour', 'Semaine', 'Jour férié', 'Tâches prévues']
    for col_num, header in enumerate(calendar_headers, 1):
        cell = ws_calendar.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Calculer les dates d'exécution pour chaque tâche
    task_dates = defaultdict(list)
    for task in tasks:
        dates = calculate_next_dates(task['periodicite'], datetime(2026, 1, 1), 100)
        for date in dates:
            if date.year == 2026:
                task_dates[date.date()].append(task['libelle'])
    
    # Écrire le calendrier
    row_num = 2
    for date in sorted(calendar.keys()):
        tasks_for_date = task_dates.get(date, [])
        
        # Date
        ws_calendar.cell(row=row_num, column=1, value=date.strftime('%d/%m/%Y')).border = border
        ws_calendar.cell(row=row_num, column=1).alignment = center_alignment
        
        # Jour de la semaine
        ws_calendar.cell(row=row_num, column=2, value=calendar[date]['weekday']).border = border
        ws_calendar.cell(row=row_num, column=2).alignment = center_alignment
        
        # Numéro de semaine
        ws_calendar.cell(row=row_num, column=3, value=calendar[date]['week_number']).border = border
        ws_calendar.cell(row=row_num, column=3).alignment = center_alignment
        
        # Jour férié
        if calendar[date]['is_holiday']:
            cell = ws_calendar.cell(row=row_num, column=4, value=calendar[date]['holiday_name'])
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        else:
            ws_calendar.cell(row=row_num, column=4, value='').border = border
        ws_calendar.cell(row=row_num, column=4).alignment = center_alignment
        ws_calendar.cell(row=row_num, column=4).border = border
        
        # Tâches prévues
        tasks_text = '\n'.join(tasks_for_date) if tasks_for_date else ''
        cell = ws_calendar.cell(row=row_num, column=5, value=tasks_text)
        cell.alignment = wrap_alignment
        cell.border = border
        
        # Mettre en évidence les jours avec des tâches
        if tasks_for_date:
            cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        row_num += 1
    
    # Ajuster la largeur des colonnes du calendrier
    ws_calendar.column_dimensions['A'].width = 12
    ws_calendar.column_dimensions['B'].width = 12
    ws_calendar.column_dimensions['C'].width = 10
    ws_calendar.column_dimensions['D'].width = 25
    ws_calendar.column_dimensions['E'].width = 80
    
    # Geler la première ligne
    ws.freeze_panes = 'A2'
    ws_calendar.freeze_panes = 'A2'
    
    # Sauvegarder
    wb.save(output_path)
    print(f"[OK] Fichier Excel cree: {output_path}")


def main():
    """Fonction principale"""
    print("=" * 70)
    print("ANALYSE DE LA DOCUMENTATION KBA105")
    print("Extraction des tâches de maintenance préventive")
    print("=" * 70)
    
    # Chemins
    script_dir = Path(__file__).parent
    zip_path = script_dir / "kba105.zip"
    # Vérifier aussi dans le sous-dossier kba105
    if not zip_path.exists():
        zip_path = script_dir / "kba105" / "kba105.zip"
    output_path = script_dir / "kba105_maintenance_preventive.xlsx"
    
    try:
        # Analyser l'archive
        tasks, example_files = analyze_kba105_archive(str(zip_path))
        
        if not tasks:
            print("\n[ATTENTION] Aucune tache de maintenance preventive trouvee.")
            print("Vérifiez que le fichier kba105.zip contient bien des fichiers de documentation.")
            return
        
        # Afficher quelques statistiques
        print("\n" + "=" * 70)
        print("STATISTIQUES")
        print("=" * 70)
        
        periodicity_count = defaultdict(int)
        component_count = defaultdict(int)
        doc_type_count = defaultdict(int)
        
        for task in tasks:
            periodicity_count[task['periodicite']] += 1
            component_count[task['composant']] += 1
            doc_type_count[task['type_document']] += 1
        
        print(f"\nRépartition par périodicité:")
        for periodicity, count in sorted(periodicity_count.items()):
            print(f"  - {periodicity}: {count}")
        
        print(f"\nTop 10 composants concernés:")
        for component, count in sorted(component_count.items(), key=lambda x: x[1], reverse=True)[:10]:
            print(f"  - {component}: {count}")
        
        print(f"\nRépartition par type de document:")
        for doc_type, count in sorted(doc_type_count.items()):
            print(f"  - {doc_type}: {count}")
        
        # Générer le fichier Excel
        create_excel_report(tasks, str(output_path))
        
        print("\n" + "=" * 70)
        print("[OK] ANALYSE TERMINEE AVEC SUCCES")
        print("=" * 70)
        print(f"\nFichier Excel généré: {output_path}")
        print(f"Nombre de tâches extraites: {len(tasks)}")
        
    except FileNotFoundError as e:
        print(f"\n[ERREUR] {e}")
        print(f"\nVeuillez placer le fichier kba105.zip dans le dossier:")
        print(f"  {script_dir}")
    except Exception as e:
        print(f"\n[ERREUR] {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

