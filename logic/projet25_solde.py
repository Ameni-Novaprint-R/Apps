# -*- coding: utf-8 -*-
"""Projet 25 – Soldes de congés (fiche type Excel, import unique)."""
import os
import re
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP

from db import get_db_cursor
from logic.projet25 import get_person, list_personel_actifs, _int_mat

IMPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'imports_projet25', 'solde_conge')
DEFAULT_IMPORT_FILE = 'Copie de Congés-Restants-2026.xlsx'
SHEET_NAME = 'Personnels'

MOIS_LABELS = ['JANV.', 'FEV.', 'MARS', 'AVRIL', 'MAI', 'JUIN', 'JUILLET', 'AOUT', 'SEP.', 'OCT.', 'NOV.', 'DEC.']

DROITS_SUPPL = [
    ('ANCIENNETE', 'Ancienneté', 2),
    ('MARIAGE', 'Mariage', 8),
    ('MATERNITE', 'Maternité', 15),
    ('MATERNITE_90', 'Maternité (90 j)', 90),
    ('PATERNITE', 'Paternité', 7),
    ('NAISSANCE', 'Naissance', 3),
    ('CIRCONCISION', 'Circoncision', 2),
    ('MARIAGE_ENFANT', 'Mariage enfant', 2),
    ('DECES_PROCHE', 'Décès proche', 4),
    ('DECES_ETENDU', 'Décès famille élargie', 3),
]

Q2 = Decimal('0.01')
Q217 = Decimal('2.17')
Q15 = Decimal('1.5')
PLAFOND_15 = Decimal('26')
PLAFOND_217_ANNUEL = Decimal('2.17') * Decimal('12')


def _d(val):
    if val is None:
        return Decimal('0')
    if isinstance(val, Decimal):
        return val
    return Decimal(str(val).replace(',', '.'))


def _f2(val):
    return float(_d(val).quantize(Q2, rounding=ROUND_HALF_UP))


def init_solde_fiche_tables():
    sql_blocks = [
        """IF OBJECT_ID('dbo.WEB_CONGE_SOLDE_FICHE', 'U') IS NULL
           CREATE TABLE dbo.WEB_CONGE_SOLDE_FICHE (
               ID INT IDENTITY(1,1) PRIMARY KEY,
               Matricule INT NOT NULL,
               Annee INT NOT NULL,
               TauxMensuel DECIMAL(6,2) NOT NULL,
               DroitFixe DECIMAL(6,2) NULL,
               DroitFixeManuel BIT NOT NULL DEFAULT 0,
               DateProchaineAdditionFixe DATE NULL,
               ReliquatAnneePrecedente DECIMAL(8,2) NOT NULL DEFAULT 0,
               Departement NVARCHAR(120) NULL,
               DateDernierImport DATETIME NULL,
               CONSTRAINT UQ_WEB_CONGE_SOLDE_FICHE UNIQUE (Matricule, Annee))""",
        """IF OBJECT_ID('dbo.WEB_CONGE_SOLDE_MENSUEL', 'U') IS NULL
           CREATE TABLE dbo.WEB_CONGE_SOLDE_MENSUEL (
               ID INT IDENTITY(1,1) PRIMARY KEY,
               Matricule INT NOT NULL,
               Annee INT NOT NULL,
               Mois INT NOT NULL,
               CongeAccorde DECIMAL(8,2) NOT NULL DEFAULT 0,
               SourceImport DECIMAL(8,2) NOT NULL DEFAULT 0,
               SourceP25 DECIMAL(8,2) NOT NULL DEFAULT 0,
               CONSTRAINT UQ_WEB_CONGE_SOLDE_MENSUEL UNIQUE (Matricule, Annee, Mois))""",
        """IF OBJECT_ID('dbo.WEB_CONGE_IMPORT_ECART', 'U') IS NULL
           CREATE TABLE dbo.WEB_CONGE_IMPORT_ECART (
               ID INT IDENTITY(1,1) PRIMARY KEY,
               DateImport DATETIME NOT NULL DEFAULT GETDATE(),
               MatriculeExcel NVARCHAR(20) NULL,
               NomExcel NVARCHAR(200) NULL,
               Raison NVARCHAR(500) NOT NULL,
               FichierSource NVARCHAR(500) NULL)""",
    ]
    with get_db_cursor() as cursor:
        for block in sql_blocks:
            cursor.execute(block)
        cursor.connection.commit()


def mois_complets_ecoules(ref=None):
    """Mois calendaires acquis dans l'année (mois en cours inclus dès le 1er jour)."""
    if ref is None:
        ref = date.today()
    return ref.month


def _parse_embauche_from_cell(val):
    if not val:
        return None
    if hasattr(val, 'date'):
        return val.date() if hasattr(val, 'date') else val
    s = str(val)
    m = re.search(r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})', s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None


def _parse_matricule_cell(val):
    if not val:
        return None
    m = re.search(r'(\d+)', str(val))
    return _int_mat(m.group(1)) if m else None


def _find_block_starts(ws):
    starts = []
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row, 2).value
        if v and isinstance(v, str) and 'MATRICULE' in v.upper():
            starts.append(row)
    return starts


def _parse_block(ws, start_row):
    mat_excel = ws.cell(start_row, 2).value or ''
    mat = _parse_matricule_cell(mat_excel)
    nom_raw = ws.cell(start_row + 1, 2).value or ''
    nom = re.sub(r'^NOM\s*&\s*PRENOM\s*:\s*', '', str(nom_raw), flags=re.I).strip()
    emb_cell = ws.cell(start_row + 1, 16).value or ws.cell(start_row + 1, 15).value
    embauche = _parse_embauche_from_cell(emb_cell)
    dept = ws.cell(start_row, 18).value or ws.cell(start_row, 17).value
    if dept and isinstance(dept, str) and 'DEPART' in dept.upper():
        dept = re.sub(r'^.*DEPARTEMENT\s*:\s*', '', dept, flags=re.I).strip()
    reliquat = ws.cell(start_row + 5, 3).value
    annee_ref = ws.cell(start_row + 5, 1).value

    taux = None
    fixe = None
    next_add = None
    pris = {}
    acquis_import = {}

    for dr in range(start_row + 5, start_row + 18):
        e = ws.cell(dr, 5).value
        if not e:
            continue
        eu = str(e).upper()
        if 'DROIT DE CONG' in eu:
            taux = ws.cell(dr, 6).value
            for m in range(12):
                v = ws.cell(dr, 7 + m).value
                if v is not None:
                    acquis_import[m + 1] = _f2(v)
        if 'CONG' in eu and 'ACCORD' in eu:
            for m in range(12):
                v = ws.cell(dr, 7 + m).value
                if v is not None:
                    pris[m + 1] = _f2(v)
        v21 = ws.cell(dr, 21).value
        if hasattr(v21, 'year'):
            next_add = v21.date() if hasattr(v21, 'date') else v21

    for dr in range(start_row + 7, start_row + 13):
        bv = ws.cell(dr, 2).value
        if isinstance(bv, (int, float)) and bv in (0, 2, 4, 6, 8):
            fixe = _f2(bv)

    solde_excel = ws.cell(start_row + 5, 20).value

    return {
        'matricule_excel': mat_excel.strip(),
        'matricule': mat,
        'nom_excel': nom,
        'embauche': embauche,
        'departement': str(dept).strip() if dept else None,
        'reliquat': _f2(reliquat) if reliquat is not None else 0.0,
        'annee_ref': int(annee_ref) if annee_ref else None,
        'taux_mensuel': _f2(taux) if taux is not None else None,
        'droit_fixe': fixe,
        'date_prochaine_addition': next_add,
        'conges_pris': pris,
        'acquis_import': acquis_import,
        'solde_excel': _f2(solde_excel) if solde_excel is not None else None,
        'start_row': start_row,
    }


def parse_excel_fiches(path=None):
    import openpyxl
    if path is None:
        path = os.path.join(IMPORT_DIR, DEFAULT_IMPORT_FILE)
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Fichier introuvable : {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    blocks = []
    for sr in _find_block_starts(ws):
        blocks.append(_parse_block(ws, sr))
    return blocks, os.path.basename(path)


def _personel_matricules_set():
    return {p['matricule'] for p in list_personel_actifs('')}


def _personel_statut(matricule):
    """actif | archive | absent"""
    m = _int_mat(matricule)
    if m is None:
        return 'absent'
    with get_db_cursor() as cursor:
        cursor.execute("SELECT archive FROM personel WHERE Matricule = ?", (m,))
        row = cursor.fetchone()
        if not row:
            return 'absent'
        if row.archive in (1, True):
            return 'archive'
        return 'actif'


def purge_import_ecarts_obsoletes():
    """Supprime les écarts liés aux anciens collaborateurs ou fiches désormais importables."""
    with get_db_cursor() as cursor:
        cursor.execute("""
            DELETE E FROM WEB_CONGE_IMPORT_ECART E
            WHERE E.Raison LIKE N'Matricule % absent de la table personel%'
        """)
        cursor.connection.commit()


def _calc_fixe_nouvelle_embauche(embauche, ref=None):
    if not embauche:
        return Decimal('0'), None
    if ref is None:
        ref = date.today()
    years = (ref - embauche).days // 365
    palier = (years // 5) * 2
    fixe = min(Decimal('8'), Decimal(str(palier)))
    years_next = ((years // 5) + 1) * 5
    try:
        next_d = embauche.replace(year=embauche.year + years_next)
    except ValueError:
        next_d = embauche.replace(year=embauche.year + years_next, day=28)
    return fixe, next_d


def apply_fixe_anniversary(fiche, ref=None):
    """Applique +2 j / 5 ans si date dépassée (sauf si manuel RH)."""
    if ref is None:
        ref = date.today()
    if fiche.get('droit_fixe_manuel'):
        return fiche
    if _d(fiche.get('taux_mensuel')) == Q217:
        return fiche
    fixe = _d(fiche.get('droit_fixe') or 0)
    next_d = fiche.get('date_prochaine_addition_fixe')
    if isinstance(next_d, str):
        next_d = datetime.strptime(next_d[:10], '%Y-%m-%d').date()
    changed = False
    while next_d and ref >= next_d and fixe < Decimal('8'):
        fixe += Decimal('2')
        try:
            next_d = next_d.replace(year=next_d.year + 5)
        except ValueError:
            next_d = next_d.replace(year=next_d.year + 5, day=28)
        changed = True
    if changed:
        fiche['droit_fixe'] = _f2(fixe)
        fiche['date_prochaine_addition_fixe'] = next_d.isoformat() if next_d else None
    return fiche


def calculer_fiche(fiche, ref=None):
    if ref is None:
        ref = date.today()
    annee = fiche.get('annee') or ref.year
    taux = _d(fiche.get('taux_mensuel') or Q15)
    reliquat = _d(fiche.get('reliquat_annee_precedente') or 0)
    fixe = _d(fiche.get('droit_fixe') or 0) if taux == Q15 else Decimal('0')
    mois_ec = mois_complets_ecoules(ref)
    acquis_mois = taux * Decimal(mois_ec)

    mensuel = fiche.get('mensuel') or {}
    pris_total = Decimal('0')
    mois_data = []
    for m in range(1, 13):
        row = mensuel.get(m) or {}
        pris = _d(row.get('conge_accorde') or 0)
        pris_total += pris
        mois_data.append({
            'mois': m,
            'label': MOIS_LABELS[m - 1],
            'conge_accorde': _f2(pris),
            'acquis': _f2(taux) if m <= mois_ec else 0.0,
            'source_import': _f2(row.get('source_import') or 0),
            'source_p25': _f2(row.get('source_p25') or 0),
        })

    droit_annuel_theorique = fixe + taux * Decimal('12')
    if taux == Q15:
        plafond = PLAFOND_15
        plafond_atteint = droit_annuel_theorique >= PLAFOND_15
    else:
        plafond = PLAFOND_217_ANNUEL
        plafond_atteint = False

    solde = reliquat + fixe + acquis_mois - pris_total

    return {
        'annee': annee,
        'taux_mensuel': _f2(taux),
        'regime': '2.17' if taux == Q217 else '1.5',
        'droit_fixe': _f2(fixe) if taux == Q15 else None,
        'reliquat_annee_precedente': _f2(reliquat),
        'mois_ecoules': mois_ec,
        'acquis_periode': _f2(acquis_mois),
        'conges_accordes_total': _f2(pris_total),
        'solde_restant': _f2(solde),
        'droit_annuel_theorique': _f2(droit_annuel_theorique),
        'plafond_annuel': _f2(plafond),
        'plafond_atteint': plafond_atteint,
        'mois': mois_data,
        'droits_supplementaires': DROITS_SUPPL,
    }


def _load_mensuel(cursor, mat, annee):
    cursor.execute("""
        SELECT Mois, CongeAccorde, SourceImport, SourceP25
        FROM WEB_CONGE_SOLDE_MENSUEL WHERE Matricule = ? AND Annee = ?
    """, (mat, annee))
    out = {}
    for r in cursor.fetchall():
        out[r.Mois] = {
            'conge_accorde': _f2(r.CongeAccorde),
            'source_import': _f2(r.SourceImport),
            'source_p25': _f2(r.SourceP25),
        }
    return out


def get_fiche_solde(matricule, annee=None, apply_p25=True):
    m = _int_mat(matricule)
    if m is None:
        return None
    if annee is None:
        annee = date.today().year
    with get_db_cursor() as cursor:
        cursor.execute("""
            SELECT Matricule, Annee, TauxMensuel, DroitFixe, DroitFixeManuel,
                   DateProchaineAdditionFixe, ReliquatAnneePrecedente, Departement
            FROM WEB_CONGE_SOLDE_FICHE WHERE Matricule = ? AND Annee = ?
        """, (m, annee))
        row = cursor.fetchone()
        if not row:
            return None
        fiche = {
            'matricule': row.Matricule,
            'annee': row.Annee,
            'taux_mensuel': _f2(row.TauxMensuel),
            'droit_fixe': _f2(row.DroitFixe) if row.DroitFixe is not None else None,
            'droit_fixe_manuel': bool(row.DroitFixeManuel),
            'date_prochaine_addition_fixe': row.DateProchaineAdditionFixe.isoformat()
            if row.DateProchaineAdditionFixe else None,
            'reliquat_annee_precedente': _f2(row.ReliquatAnneePrecedente),
            'departement': row.Departement,
            'mensuel': _load_mensuel(cursor, m, annee),
        }
    if apply_p25:
        sync_p25_conges_to_mensuel(m, annee)
        with get_db_cursor() as cursor:
            fiche['mensuel'] = _load_mensuel(cursor, m, annee)
    fiche = apply_fixe_anniversary(fiche)
    p = get_person(m)
    fiche['label'] = p['label'] if p else str(m)
    fiche['embauche'] = p.get('embauche') if p and p.get('embauche') else None
    calc = calculer_fiche(fiche)
    fiche.update(calc)
    return fiche


def get_solde_demande_conge(matricule, annee=None):
    """Solde pour le popup « Demande de congé » (fiche mensuelle importée)."""
    if annee is None:
        annee = date.today().year
    fiche = get_fiche_solde(matricule, annee, apply_p25=True)
    if not fiche:
        return {
            'annee': annee,
            'restant': 0.0,
            'consomme': 0.0,
            'solde': 0.0,
            'source': 'none',
            'message': 'Fiche solde non disponible — contactez les RH.',
        }
    acquis_total = _f2(
        _d(fiche['solde_restant']) + _d(fiche['conges_accordes_total'])
    )
    return {
        'annee': fiche['annee'],
        'restant': fiche['solde_restant'],
        'consomme': fiche['conges_accordes_total'],
        'solde': acquis_total,
        'regime': fiche.get('regime'),
        'source': 'fiche',
    }


def list_fiches_solde(annee=None, q=''):
    if annee is None:
        annee = date.today().year
    q = (q or '').strip().lower()
    with get_db_cursor() as cursor:
        cursor.execute("""
            SELECT F.Matricule, F.TauxMensuel, F.DroitFixe, F.DroitFixeManuel,
                   F.DateProchaineAdditionFixe, F.ReliquatAnneePrecedente, F.Departement,
                   P.Nom, P.Prenom
            FROM WEB_CONGE_SOLDE_FICHE F
            LEFT JOIN personel P ON P.Matricule = F.Matricule
            WHERE F.Annee = ?
            ORDER BY P.Nom, P.Prenom
        """, (annee,))
        rows = cursor.fetchall()
    out = []
    for r in rows:
        label = f"{r.Nom or ''} {r.Prenom or ''}".strip()
        if q and q not in label.lower() and q not in str(r.Matricule):
            continue
        with get_db_cursor() as c2:
            mensuel = _load_mensuel(c2, r.Matricule, annee)
        fiche = {
            'matricule': r.Matricule,
            'annee': annee,
            'label': label,
            'taux_mensuel': _f2(r.TauxMensuel),
            'droit_fixe': _f2(r.DroitFixe) if r.DroitFixe is not None else None,
            'droit_fixe_manuel': bool(r.DroitFixeManuel),
            'date_prochaine_addition_fixe': r.DateProchaineAdditionFixe.isoformat()
            if r.DateProchaineAdditionFixe else None,
            'reliquat_annee_precedente': _f2(r.ReliquatAnneePrecedente),
            'departement': r.Departement,
            'mensuel': mensuel,
        }
        fiche = apply_fixe_anniversary(fiche)
        calc = calculer_fiche(fiche)
        out.append({
            'matricule': r.Matricule,
            'label': label,
            'departement': r.Departement,
            'taux_mensuel': calc['taux_mensuel'],
            'regime': calc['regime'],
            'droit_fixe': calc['droit_fixe'],
            'reliquat': calc['reliquat_annee_precedente'],
            'solde_restant': calc['solde_restant'],
            'plafond_atteint': calc['plafond_atteint'],
            'date_prochaine_addition_fixe': fiche.get('date_prochaine_addition_fixe'),
        })
    return out


def import_excel_fiches(path=None, annee=None):
    if annee is None:
        annee = date.today().year
    blocks, fname = parse_excel_fiches(path)
    imported = 0
    skipped_archive = 0
    skipped_absent = 0
    ecarts = []

    purge_import_ecarts_obsoletes()

    for blk in blocks:
        mat = blk['matricule']
        if mat is None:
            ecarts.append((blk.get('matricule_excel'), blk.get('nom_excel'), 'Matricule illisible'))
            continue
        statut = _personel_statut(mat)
        if statut == 'archive':
            skipped_archive += 1
            continue
        if statut == 'absent':
            skipped_absent += 1
            continue
        taux = blk.get('taux_mensuel')
        if taux is None:
            ecarts.append((blk.get('matricule_excel'), blk.get('nom_excel'), 'Taux mensuel introuvable'))
            continue

        fixe = blk.get('droit_fixe')
        next_add = blk.get('date_prochaine_addition')
        fixe_manuel = False
        if taux == 2.17:
            fixe = None
            next_add = None
        elif fixe is None:
            p = get_person(mat)
            emb = blk.get('embauche')
            if not emb and p:
                emb = p.get('embauche')
            if emb:
                if isinstance(emb, str):
                    emb = _parse_embauche_from_cell(emb)
                fixe_dec, next_add = _calc_fixe_nouvelle_embauche(emb)
                fixe = _f2(fixe_dec)
            else:
                fixe = 0.0

        with get_db_cursor() as cursor:
            cursor.execute("""
                MERGE WEB_CONGE_SOLDE_FICHE AS t
                USING (SELECT ? AS Matricule, ? AS Annee) AS s
                ON t.Matricule = s.Matricule AND t.Annee = s.Annee
                WHEN MATCHED THEN UPDATE SET
                    TauxMensuel=?, DroitFixe=?, DroitFixeManuel=?, DateProchaineAdditionFixe=?,
                    ReliquatAnneePrecedente=?, Departement=?, DateDernierImport=GETDATE()
                WHEN NOT MATCHED THEN INSERT (
                    Matricule, Annee, TauxMensuel, DroitFixe, DroitFixeManuel,
                    DateProchaineAdditionFixe, ReliquatAnneePrecedente, Departement, DateDernierImport
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, GETDATE());
            """, (
                mat, annee,
                taux, fixe, 1 if fixe_manuel else 0, next_add, blk['reliquat'], blk.get('departement'),
                mat, annee, taux, fixe, 1 if fixe_manuel else 0, next_add, blk['reliquat'], blk.get('departement'),
            ))
            for m in range(1, 13):
                pris = blk['conges_pris'].get(m, 0.0)
                cursor.execute("""
                    MERGE WEB_CONGE_SOLDE_MENSUEL AS t
                    USING (SELECT ? AS Matricule, ? AS Annee, ? AS Mois) AS s
                    ON t.Matricule=s.Matricule AND t.Annee=s.Annee AND t.Mois=s.Mois
                    WHEN MATCHED THEN UPDATE SET CongeAccorde=?, SourceImport=?, SourceP25=0
                    WHEN NOT MATCHED THEN INSERT (Matricule, Annee, Mois, CongeAccorde, SourceImport, SourceP25)
                    VALUES (?, ?, ?, ?, ?, 0);
                """, (mat, annee, m, pris, pris, mat, annee, m, pris, pris))
            cursor.connection.commit()
        imported += 1

    with get_db_cursor() as cursor:
        for mex, nom, raison in ecarts:
            cursor.execute("""
                INSERT INTO WEB_CONGE_IMPORT_ECART (MatriculeExcel, NomExcel, Raison, FichierSource)
                VALUES (?, ?, ?, ?)
            """, (mex, nom, raison, fname))
        cursor.connection.commit()

    return {
        'imported': imported,
        'ecarts': len(ecarts),
        'ecarts_detail': [{'matricule_excel': e[0], 'nom': e[1], 'raison': e[2]} for e in ecarts],
        'ignores_archives': skipped_archive,
        'ignores_absents': skipped_absent,
        'fichier': fname,
        'annee': annee,
    }


def list_import_ecarts(limit=200):
    with get_db_cursor() as cursor:
        cursor.execute("""
            SELECT TOP (?) MatriculeExcel, NomExcel, Raison, FichierSource, DateImport
            FROM WEB_CONGE_IMPORT_ECART ORDER BY DateImport DESC
        """, (limit,))
        return [
            {
                'matricule_excel': r.MatriculeExcel,
                'nom': r.NomExcel,
                'raison': r.Raison,
                'fichier': r.FichierSource,
                'date': r.DateImport.strftime('%Y-%m-%d %H:%M') if r.DateImport else '',
            }
            for r in cursor.fetchall()
        ]


def update_fiche_rh(matricule, annee, data):
    m = _int_mat(matricule)
    with get_db_cursor() as cursor:
        cursor.execute("SELECT ID FROM WEB_CONGE_SOLDE_FICHE WHERE Matricule=? AND Annee=?", (m, annee))
        if not cursor.fetchone():
            return False, "Fiche introuvable."
        sets = []
        params = []
        if 'droit_fixe' in data:
            sets.append('DroitFixe=?')
            params.append(data['droit_fixe'])
            sets.append('DroitFixeManuel=1')
        if 'date_prochaine_addition_fixe' in data:
            sets.append('DateProchaineAdditionFixe=?')
            d = data['date_prochaine_addition_fixe']
            params.append(d if d else None)
        if 'reliquat_annee_precedente' in data:
            sets.append('ReliquatAnneePrecedente=?')
            params.append(data['reliquat_annee_precedente'])
        if not sets:
            return True, None
        params.extend([m, annee])
        cursor.execute(
            f"UPDATE WEB_CONGE_SOLDE_FICHE SET {', '.join(sets)} WHERE Matricule=? AND Annee=?",
            params,
        )
        cursor.connection.commit()
    return True, None


def sync_p25_conges_to_mensuel(matricule, annee):
    """Recalcule les congés annuels validés P25 par mois (sans écraser l'import)."""
    m = _int_mat(matricule)
    with get_db_cursor() as cursor:
        cursor.execute("""
            SELECT D.DateDebut, D.NbJoursOuvres
            FROM WEB_CONGE_DEMANDE D
            INNER JOIN WEB_CONGE_TYPE T ON T.ID = D.ID_TypeConge
            WHERE D.TypeDemande='CONGE' AND D.MatriculeDemandeur=? AND D.Statut='VALIDE'
              AND T.Code='ANNUEL' AND D.DateDebut IS NOT NULL
              AND YEAR(D.DateDebut)=?
        """, (m, annee))
        p25_by_month = {i: Decimal('0') for i in range(1, 13)}
        for r in cursor.fetchall():
            nb = _d(r.NbJoursOuvres or 0)
            if nb > 0:
                p25_by_month[r.DateDebut.month] += nb
        for mois in range(1, 13):
            nb = p25_by_month[mois]
            cursor.execute("""
                SELECT SourceImport FROM WEB_CONGE_SOLDE_MENSUEL
                WHERE Matricule=? AND Annee=? AND Mois=?
            """, (m, annee, mois))
            ex = cursor.fetchone()
            src_imp = _d(ex.SourceImport) if ex else Decimal('0')
            total = src_imp + nb
            if ex or nb > 0:
                cursor.execute("""
                    MERGE WEB_CONGE_SOLDE_MENSUEL AS t
                    USING (SELECT ? AS Matricule, ? AS Annee, ? AS Mois) AS s
                    ON t.Matricule=s.Matricule AND t.Annee=s.Annee AND t.Mois=s.Mois
                    WHEN MATCHED THEN UPDATE SET SourceP25=?, CongeAccorde=?
                    WHEN NOT MATCHED THEN INSERT (Matricule, Annee, Mois, CongeAccorde, SourceImport, SourceP25)
                    VALUES (?, ?, ?, ?, 0, ?);
                """, (m, annee, mois, _f2(nb), _f2(total), m, annee, mois, _f2(total), _f2(nb)))
        cursor.connection.commit()


def clore_annee(annee):
    """R1 : solde au 31/12 devient reliquat de l'année suivante."""
    annee_suiv = annee + 1
    with get_db_cursor() as cursor:
        cursor.execute("SELECT Matricule FROM WEB_CONGE_SOLDE_FICHE WHERE Annee=?", (annee,))
        mats = [r.Matricule for r in cursor.fetchall()]
    for mat in mats:
        fiche = get_fiche_solde(mat, annee, apply_p25=True)
        if not fiche:
            continue
        solde_fin = _d(fiche['solde_restant'])
        with get_db_cursor() as cursor:
            cursor.execute("""
                SELECT ID FROM WEB_CONGE_SOLDE_FICHE WHERE Matricule=? AND Annee=?
            """, (mat, annee_suiv))
            if cursor.fetchone():
                cursor.execute("""
                    UPDATE WEB_CONGE_SOLDE_FICHE SET ReliquatAnneePrecedente=? WHERE Matricule=? AND Annee=?
                """, (_f2(solde_fin), mat, annee_suiv))
            else:
                prev = fiche
                cursor.execute("""
                    INSERT INTO WEB_CONGE_SOLDE_FICHE (
                        Matricule, Annee, TauxMensuel, DroitFixe, DroitFixeManuel,
                        DateProchaineAdditionFixe, ReliquatAnneePrecedente, Departement
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    mat, annee_suiv, prev['taux_mensuel'], prev.get('droit_fixe'),
                    1 if prev.get('droit_fixe_manuel') else 0,
                    prev.get('date_prochaine_addition_fixe'),
                    _f2(solde_fin), prev.get('departement'),
                ))
            for m in range(1, 13):
                cursor.execute("""
                    DELETE FROM WEB_CONGE_SOLDE_MENSUEL WHERE Matricule=? AND Annee=? AND Mois=?
                """, (mat, annee_suiv, m))
            cursor.connection.commit()
    return {'annee_source': annee, 'annee_cible': annee_suiv, 'employes': len(mats)}
