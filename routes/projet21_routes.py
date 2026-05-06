"""
Projet 21 - Mise à jour base de données
Synchronisation Novaprint -> novaprint_restored
"""

from flask import Blueprint, render_template, jsonify, send_file
import pyodbc
import threading
from datetime import datetime
from flask import request
from io import BytesIO
import pandas as pd
from openpyxl.styles import PatternFill
import os
import socket
import time

try:
    from local_env import load_project_env

    load_project_env()
except ImportError:
    pass

projet21_bp = Blueprint('projet21', __name__, url_prefix='/projet21')

# Marqueur pour vérifier quelle version du code est chargée par Flask
PROJET21_CODE_VERSION = "projet21_routes.py::orphan-delete-target-map-v4::2026-04-06"

# Configuration des bases de données
SOURCE_CONFIG = {
    'server': 'SageSRV\\Graphisoft',
    'database': 'Novaprint',
    'username': 'sa',
    'password': 'Graphis0ft'
}

XRT_SOURCE_CONFIG = {
    # IMPORTANT: la base XRT est STRICTEMENT en lecture seule côté synchro.
    # Le mot de passe et le nom de base sont fournis via variables d'environnement.
    # NOTE: l'instance nommée nécessite SQL Browser (UDP 1434) pour la résolution.
    # Pour fiabilité, on force TCP+port connu.
    'server': r'tcp:SRV-XRT2,1433',
    'database_env': 'XRT_DATABASE',
    'database_default': 'SXA',
    'username': 'NOVAPRINT1',
    'password_env': 'XRT_PASSWORD',
}

TARGET_CONFIG = {
    'server': '192.168.10.225',
    'database': 'novaprint_restored',
    'trusted_connection': False,  # Utiliser l'authentification SQL Server pour Task Scheduler
    'username': 'sa',
    'password': 'bA8ALvct9QtX'
}

sync_status = {'running': False, 'message': '', 'progress': 0, 'details': [], 'code_version': PROJET21_CODE_VERSION}
xrt_sync_status = {'running': False, 'message': '', 'progress': 0, 'details': [], 'code_version': PROJET21_CODE_VERSION}

# Synchro XRT « miroir » (INSERT/UPDATE) : plafond de lignes pour chargement mémoire complet.
# Au-delà, si écart COUNT ou checksum : recopie table via xrt_copy_table_full (flux fetchmany).
XRT_MIRROR_MAX_ROWS_IN_MEMORY = int(os.environ.get("XRT_MIRROR_MAX_ROWS_IN_MEMORY", "50000"))

def get_xrt_source_runtime_config():
    db = os.environ.get(XRT_SOURCE_CONFIG['database_env'], '').strip() or XRT_SOURCE_CONFIG.get('database_default', '')
    pwd = os.environ.get(XRT_SOURCE_CONFIG['password_env'], '').strip()
    if not db:
        raise Exception(
            f"Configuration manquante: variable d'environnement {XRT_SOURCE_CONFIG['database_env']} (nom de base XRT)."
        )
    if not pwd:
        raise Exception(
            f"Configuration manquante: variable d'environnement {XRT_SOURCE_CONFIG['password_env']} "
            f"(mot de passe SQL de {XRT_SOURCE_CONFIG['username']}). "
            "Définissez-la pour le compte qui lance l’application, ou créez un fichier .env à la racine du projet "
            f"avec une ligne {XRT_SOURCE_CONFIG['password_env']}=… puis redémarrez Flask."
        )
    return {
        'server': XRT_SOURCE_CONFIG['server'],
        'database': db,
        'username': XRT_SOURCE_CONFIG['username'],
        'password': pwd,
    }

def get_connection(config, readonly=False):
    """
    Connexion SQL Server.

    Important: éviter le driver legacy "{SQL Server}" côté cible, car il déclenche
    des erreurs HYC00 sur certains types lors des INSERT/SQLBindParameter.
    On privilégie ODBC Driver 17/18 avec fallback contrôlé.
    """
    if config.get('trusted_connection'):
        driver_candidates = ["ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server", "SQL Server"]
        last_err = None
        for drv in driver_candidates:
            try:
                conn_str = (
                    f"DRIVER={{{drv}}};"
                    f"SERVER={config['server']};"
                    f"DATABASE={config['database']};"
                    f"Trusted_Connection=yes;"
                    f"TrustServerCertificate=yes"
                )
                conn = pyodbc.connect(conn_str)
                break
            except Exception as e:
                last_err = e
                conn = None
        if conn is None:
            raise last_err
    else:
        conn_str = (
            f"DRIVER={{ODBC Driver 17 for SQL Server}};"
            f"SERVER={config['server']};"
            f"DATABASE={config['database']};"
            f"UID={config['username']};"
            f"PWD={config['password']}"
        )
        conn = pyodbc.connect(conn_str)
    if readonly:
        conn.execute("SET TRANSACTION ISOLATION LEVEL READ UNCOMMITTED")
    return conn

def xrt_connection_diagnostics():
    """
    Diagnostic détaillé (lecture seule) de la connectivité XRT.
    - Ne modifie aucune donnée.
    - Tente plusieurs drivers/options et formats serveur.
    """
    started = datetime.now()
    report = {
        'started_at': started.isoformat(),
        'source': {
            'server_configured': XRT_SOURCE_CONFIG.get('server'),
            'database': None,
            'username': XRT_SOURCE_CONFIG.get('username'),
        },
        'odbc': {
            'available_drivers': [],
        },
        'dns': {},
        'tcp': {},
        'odbc_attempts': [],
    }

    # Config runtime (base + password)
    try:
        cfg = get_xrt_source_runtime_config()
        report['source']['database'] = cfg.get('database')
    except Exception as e:
        report['config_error'] = str(e)
        report['ended_at'] = datetime.now().isoformat()
        return report

    # DNS resolution (best-effort)
    raw_server = (cfg.get('server') or '').strip()
    host = raw_server.split('\\')[0].strip()
    if host.lower().startswith('tcp:'):
        host = host[4:]
    # retirer port éventuel "host,1433"
    if ',' in host:
        host = host.split(',', 1)[0].strip()
    report['dns']['host'] = host
    try:
        infos = socket.getaddrinfo(host, None)
        addrs = sorted({i[4][0] for i in infos if i and i[4]})
        report['dns']['resolved'] = addrs
    except Exception as e:
        report['dns']['error'] = str(e)

    # TCP reachability test (1433) - best effort
    def test_tcp(port: int, timeout_s: float = 2.0):
        t0 = time.time()
        try:
            with socket.create_connection((host, port), timeout=timeout_s):
                return {'port': port, 'ok': True, 'ms': int((time.time() - t0) * 1000)}
        except Exception as e:
            return {'port': port, 'ok': False, 'ms': int((time.time() - t0) * 1000), 'error': str(e)}

    if host:
        report['tcp']['1433'] = test_tcp(1433)
        # 1434 (SQL Browser) - TCP test only (UDP non testé ici)
        report['tcp']['1434'] = test_tcp(1434)

    # ODBC attempts
    try:
        report['odbc']['available_drivers'] = list(pyodbc.drivers())
    except Exception as e:
        report['odbc']['available_drivers_error'] = str(e)

    available = set(report.get('odbc', {}).get('available_drivers') or [])
    driver_candidates = [d for d in ["ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server"] if d in available]
    if not driver_candidates:
        report['odbc']['error'] = "Aucun driver ODBC SQL Server compatible trouvé (ODBC Driver 17/18)."
        report['ended_at'] = datetime.now().isoformat()
        return report
    server_candidates = [
        cfg['server'],               # ex: tcp:SRV-XRT2,1433
        f"tcp:{host},1433",          # fallback: force TCP+port
        f"tcp:{host}",               # fallback: force TCP sans port
        host,                        # fallback: default instance
    ]
    # Deduplicate while keeping order
    seen = set()
    server_candidates = [s for s in server_candidates if s and not (s in seen or seen.add(s))]

    option_sets = [
        {'Encrypt': 'no', 'TrustServerCertificate': 'yes'},
        {'Encrypt': 'yes', 'TrustServerCertificate': 'yes'},
        {},  # no options
    ]

    for drv in driver_candidates:
        for srv in server_candidates:
            for opts in option_sets:
                attempt = {
                    'driver': drv,
                    'server': srv,
                    'database': cfg['database'],
                    'options': opts,
                    'ok': False,
                }
                try:
                    opt_str = ''.join([f"{k}={v};" for k, v in opts.items()])
                    conn_str = (
                        f"DRIVER={{{drv}}};"
                        f"SERVER={srv};"
                        f"DATABASE={cfg['database']};"
                        f"UID={cfg['username']};"
                        f"PWD={cfg['password']};"
                        f"{opt_str}"
                        f"Connection Timeout=3;"
                    )
                    t0 = time.time()
                    conn = pyodbc.connect(conn_str)
                    cur = conn.cursor()
                    cur.execute("SELECT 1")
                    _ = cur.fetchone()
                    conn.close()
                    attempt['ok'] = True
                    attempt['ms'] = int((time.time() - t0) * 1000)
                    report['odbc_attempts'].append(attempt)
                    report['ended_at'] = datetime.now().isoformat()
                    return report
                except Exception as e:
                    attempt['error'] = str(e)
                    report['odbc_attempts'].append(attempt)

    report['ended_at'] = datetime.now().isoformat()
    return report

def get_primary_keys(cursor, table_name):
    # Utiliser sys.objects et sys.index_columns pour une meilleure compatibilité
    cursor.execute("""
        SELECT c.name
        FROM sys.indexes i
        INNER JOIN sys.index_columns ic ON i.object_id = ic.object_id AND i.index_id = ic.index_id
        INNER JOIN sys.columns c ON ic.object_id = c.object_id AND ic.column_id = c.column_id
        INNER JOIN sys.tables t ON i.object_id = t.object_id
        WHERE i.is_primary_key = 1
        AND t.name = ?
        ORDER BY ic.key_ordinal
    """, (table_name,))
    result = [row[0] for row in cursor.fetchall()]
    if not result:
        # Fallback: chercher colonne ID
        cursor.execute("""
            SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = ? AND COLUMN_NAME = 'ID'
        """, (table_name,))
        row = cursor.fetchone()
        if row:
            return [row[0]]
    return result

def get_table_columns(cursor, table_name):
    cursor.execute("""
        SELECT COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, IS_NULLABLE, COLUMN_DEFAULT
        FROM INFORMATION_SCHEMA.COLUMNS 
        WHERE TABLE_NAME = ?
        ORDER BY ORDINAL_POSITION
    """, (table_name,))
    return cursor.fetchall()

def get_table_columns_full(cursor, schema_name, table_name):
    cursor.execute(
        """
        SELECT
            COLUMN_NAME,
            DATA_TYPE,
            CHARACTER_MAXIMUM_LENGTH,
            NUMERIC_PRECISION,
            NUMERIC_SCALE,
            DATETIME_PRECISION,
            IS_NULLABLE,
            COLUMN_DEFAULT
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
        ORDER BY ORDINAL_POSITION
        """,
        (schema_name, table_name),
    )
    return cursor.fetchall()

def get_identity_columns(cursor, schema_name, table_name):
    cursor.execute(
        """
        SELECT c.name
        FROM sys.columns c
        INNER JOIN sys.tables t ON c.object_id = t.object_id
        INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
        WHERE s.name = ? AND t.name = ? AND c.is_identity = 1
        """,
        (schema_name, table_name),
    )
    return {r[0] for r in cursor.fetchall()}


def has_identity_column_scoped(cursor, schema_name, table_name):
    """Colonne IDENTITY pour une table qualifiée schéma."""
    cursor.execute(
        """
        SELECT COUNT(*)
        FROM sys.columns c
        INNER JOIN sys.tables t ON c.object_id = t.object_id
        INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
        WHERE s.name = ? AND t.name = ? AND c.is_identity = 1
        """,
        (schema_name, table_name),
    )
    return cursor.fetchone()[0] > 0


def get_primary_keys_scoped(cursor, schema_name, table_name):
    """Clé primaire pour une table source qualifiée schéma (évite l’ambiguïté multi-schémas)."""
    cursor.execute(
        """
        SELECT c.name
        FROM sys.indexes i
        INNER JOIN sys.index_columns ic ON i.object_id = ic.object_id AND i.index_id = ic.index_id
        INNER JOIN sys.columns c ON ic.object_id = c.object_id AND ic.column_id = c.column_id
        INNER JOIN sys.tables t ON i.object_id = t.object_id
        INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
        WHERE i.is_primary_key = 1 AND s.name = ? AND t.name = ?
        ORDER BY ic.key_ordinal
        """,
        (schema_name, table_name),
    )
    result = [row[0] for row in cursor.fetchall()]
    if not result:
        cursor.execute(
            """
            SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ? AND COLUMN_NAME = 'ID'
            """,
            (schema_name, table_name),
        )
        row = cursor.fetchone()
        if row:
            return [row[0]]
    return result


def xrt_count_rows(cursor, schema_name, table_name):
    cursor.execute(f"SELECT COUNT(*) FROM [{schema_name}].[{table_name}]")
    return int(cursor.fetchone()[0] or 0)


def xrt_table_checksum_agg(cursor, schema_name, table_name):
    """
    Agrégat CHECKSUM par table (détection de divergence avec même COUNT).
    Retourne None si la requête échoue (types non supportés, etc.).
    """
    try:
        cursor.execute(
            f"SELECT CHECKSUM_AGG(CAST(BINARY_CHECKSUM(*) AS BIGINT)) FROM [{schema_name}].[{table_name}]"
        )
        row = cursor.fetchone()
        return None if row is None else row[0]
    except Exception:
        return None


def ensure_target_schema(cursor, schema_name):
    cursor.execute(
        f"""
        IF SCHEMA_ID(N'{schema_name}') IS NULL
            EXEC(N'CREATE SCHEMA [{schema_name}] AUTHORIZATION [dbo];');
        """
    )

def build_create_table_sql_xrt(source_cursor, source_schema, table_name, target_schema='XRT'):
    tgt_name = xrt_target_table_name(source_schema, table_name)
    cols = get_table_columns_full(source_cursor, source_schema, table_name)
    identity_cols = get_identity_columns(source_cursor, source_schema, table_name)

    col_defs = []
    for (
        col_name,
        data_type,
        char_len,
        num_prec,
        num_scale,
        dt_prec,
        is_nullable,
        col_default,
    ) in cols:
        dt = (data_type or '').lower()
        type_sql = data_type
        if dt in ('varchar', 'nvarchar', 'char', 'nchar', 'binary', 'varbinary'):
            if char_len is None:
                type_sql = data_type
            elif int(char_len) < 0:
                type_sql = f"{data_type}(MAX)"
            else:
                type_sql = f"{data_type}({int(char_len)})"
        elif dt in ('decimal', 'numeric'):
            if num_prec is not None and num_scale is not None:
                type_sql = f"{data_type}({int(num_prec)},{int(num_scale)})"
        elif dt in ('datetime2', 'time', 'datetimeoffset'):
            if dt_prec is not None:
                type_sql = f"{data_type}({int(dt_prec)})"

        null_sql = "NULL" if (is_nullable or '').upper() == 'YES' else "NOT NULL"
        identity_sql = " IDENTITY(1,1)" if col_name in identity_cols else ""
        default_sql = f" DEFAULT {col_default}" if col_default else ""
        col_defs.append(f"[{col_name}] {type_sql}{identity_sql}{default_sql} {null_sql}")

    create_sql = (
        f"CREATE TABLE [{target_schema}].[{tgt_name}] (\n  " + ",\n  ".join(col_defs) + "\n)"
    )
    return create_sql

def xrt_list_source_tables(source_cursor):
    source_cursor.execute(
        """
        SELECT s.name AS schema_name, t.name AS table_name
        FROM sys.tables t
        INNER JOIN sys.schemas s ON s.schema_id = t.schema_id
        WHERE t.is_ms_shipped = 0
        ORDER BY s.name, t.name
        """
    )
    return [(r[0], r[1]) for r in source_cursor.fetchall()]

def xrt_target_table_name(source_schema: str, table_name: str) -> str:
    """
    Nom de table dans le schéma XRT (cible) pour éviter les collisions entre schémas sources.
    - dbo.* : on conserve le nom
    - autre schéma : prefixe SCHEMA__TABLE
    """
    if (source_schema or "").lower() == "dbo":
        return table_name
    safe_schema = (source_schema or "").replace(".", "_").replace(" ", "_")
    return f"{safe_schema}__{table_name}"


def xrt_target_table_exists(target_cursor, target_schema: str, tgt_name: str) -> bool:
    """True si la table cible existe déjà dans le schéma (pour mode complétion sans DROP global)."""
    target_cursor.execute(
        """
        SELECT 1
        FROM sys.tables t
        INNER JOIN sys.schemas s ON t.schema_id = s.schema_id
        WHERE LOWER(s.name) = LOWER(?) AND LOWER(t.name) = LOWER(?)
        """,
        (target_schema, tgt_name),
    )
    return target_cursor.fetchone() is not None


def xrt_drop_all_user_tables_in_schema(cursor, schema_name: str) -> int:
    """
    Supprime toutes les tables utilisateur d'un schéma (cible), en respectant un ordre compatible FK intra-schéma.
    Retourne le nombre de DROP tentés.
    """
    cursor.execute(
        """
        SELECT t.name AS table_name
        FROM sys.tables t
        INNER JOIN sys.schemas s ON s.schema_id = t.schema_id
        WHERE s.name = ? AND t.is_ms_shipped = 0;
        """,
        (schema_name,),
    )
    all_tbl = [r[0] for r in cursor.fetchall()]
    if not all_tbl:
        return 0
    all_set = set(all_tbl)

    cursor.execute(
        """
        SELECT tp.name AS parent_table, tr.name AS child_table
        FROM sys.foreign_keys fk
        INNER JOIN sys.tables tp ON fk.referenced_object_id = tp.object_id
        INNER JOIN sys.schemas sch_p ON tp.schema_id = sch_p.schema_id
        INNER JOIN sys.tables tr ON fk.parent_object_id = tr.object_id
        INNER JOIN sys.schemas sch_c ON tr.schema_id = sch_c.schema_id
        WHERE sch_p.name = ? AND sch_c.name = ?;
        """,
        (schema_name, schema_name),
    )

    parent_to_children = {t: set() for t in all_tbl}
    indeg = {t: 0 for t in all_tbl}
    for parent_name, child_name in cursor.fetchall():
        if parent_name in all_set and child_name in all_set:
            parent_to_children[parent_name].add(child_name)
            indeg[child_name] += 1

    queue = [t for t in all_tbl if indeg[t] == 0]
    topo = []
    while queue:
        n = queue.pop(0)
        topo.append(n)
        for ch in parent_to_children.get(n, set()):
            indeg[ch] -= 1
            if indeg[ch] == 0:
                queue.append(ch)

    if len(topo) < len(all_tbl):
        # cycles ou FK complexes: compléter par nom (dernier recours)
        remaining = [t for t in sorted(all_tbl) if t not in topo]
        topo.extend(remaining)

    dropped = 0
    for tname in reversed(topo):
        cursor.execute(
            f"IF OBJECT_ID(N'[{schema_name}].[{tname}]', N'U') IS NOT NULL DROP TABLE [{schema_name}].[{tname}];"
        )
        dropped += 1
    return dropped


def xrt_sort_tables_for_copy_cursor(source_cursor, tables):
    if not tables:
        return tables

    table_set = set(tables)

    source_cursor.execute(
        """
        SELECT sch_p.name AS parent_schema, tp.name AS parent_table,
               sch_c.name AS child_schema, tr.name AS child_table
        FROM sys.foreign_keys fk
        INNER JOIN sys.tables tr ON fk.parent_object_id = tr.object_id
        INNER JOIN sys.schemas sch_c ON tr.schema_id = sch_c.schema_id
        INNER JOIN sys.tables tp ON fk.referenced_object_id = tp.object_id
        INNER JOIN sys.schemas sch_p ON tp.schema_id = sch_p.schema_id
        WHERE tr.is_ms_shipped = 0 AND tp.is_ms_shipped = 0;
        """
    )

    # Graphe: enfant -> parents
    parents_of = {t: set() for t in table_set}
    children_of = {t: set() for t in table_set}

    for ps, pt, cs, ct in source_cursor.fetchall():
        parent = (ps, pt)
        child = (cs, ct)
        if parent in table_set and child in table_set and parent != child:
            parents_of[child].add(parent)
            children_of[parent].add(child)

    # Kahn: privilégier les tables sans parent dans l'ensemble
    indeg = {t: len(parents_of[t]) for t in table_set}
    queue = [t for t in table_set if indeg[t] == 0]
    queue.sort(key=lambda x: (x[0].lower(), x[1].lower()))
    ordered = []
    while queue:
        n = queue.pop(0)
        ordered.append(n)
        for ch in sorted(children_of.get(n, set()), key=lambda x: (x[0].lower(), x[1].lower())):
            indeg[ch] -= 1
            if indeg[ch] == 0:
                queue.append(ch)
        queue.sort(key=lambda x: (x[0].lower(), x[1].lower()))

    if len(ordered) < len(table_set):
        remaining = [t for t in sorted(table_set, key=lambda x: (x[0].lower(), x[1].lower())) if t not in ordered]
        ordered.extend(remaining)

    return ordered


def xrt_copy_table_full(source_cursor, target_cursor, source_schema, table_name, status, target_schema='XRT', batch_size=2000):
    # Drop + recreate to guarantee full mirror
    tgt_name = xrt_target_table_name(source_schema, table_name)
    identity_cols = get_identity_columns(source_cursor, source_schema, table_name)
    identity_on = False

    target_cursor.execute(
        f"IF OBJECT_ID(N'[{target_schema}].[{tgt_name}]', N'U') IS NOT NULL DROP TABLE [{target_schema}].[{tgt_name}];"
    )
    create_sql = build_create_table_sql_xrt(source_cursor, source_schema, table_name, target_schema=target_schema)
    target_cursor.execute(create_sql)

    # Read-only source
    source_cursor.execute(f"SELECT * FROM [{source_schema}].[{table_name}]")
    columns = [d[0] for d in source_cursor.description]
    if not columns:
        return 0

    col_list = ", ".join([f"[{c}]" for c in columns])
    placeholders = ", ".join(["?"] * len(columns))
    insert_sql = f"INSERT INTO [{target_schema}].[{tgt_name}] ({col_list}) VALUES ({placeholders})"

    inserted = 0
    target_cursor.fast_executemany = True
    try:
        if identity_cols:
            target_cursor.execute(
                f"SET IDENTITY_INSERT [{target_schema}].[{tgt_name}] ON"
            )
            identity_on = True
        while True:
            rows = source_cursor.fetchmany(batch_size)
            if not rows:
                break
            target_cursor.executemany(insert_sql, rows)
            inserted += len(rows)
            if inserted % (batch_size * 5) == 0:
                status['details'].append(
                    f"   … {source_schema}.{table_name} -> {target_schema}.{tgt_name}: {inserted} lignes copiées"
                )
    finally:
        if identity_on:
            target_cursor.execute(
                f"SET IDENTITY_INSERT [{target_schema}].[{tgt_name}] OFF"
            )

    return inserted

def xrt_sync_databases():
    global xrt_sync_status
    xrt_sync_status = {'running': True, 'message': 'Démarrage XRT...', 'progress': 0, 'details': [], 'code_version': PROJET21_CODE_VERSION}
    xrt_sync_status['details'].append(
        f"🧩 Version code Projet21: {PROJET21_CODE_VERSION} ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})"
    )
    xrt_sync_status['details'].append("🔒 Source XRT: lecture seule (aucune écriture).")
    started_at = datetime.now()

    try:
        xrt_cfg = get_xrt_source_runtime_config()
        source_conn = get_connection(xrt_cfg, readonly=True)
        source_cursor = source_conn.cursor()

        target_conn = get_connection(TARGET_CONFIG)
        target_cursor = target_conn.cursor()
        ensure_target_schema(target_cursor, 'XRT')
        target_conn.commit()

        dropped = xrt_drop_all_user_tables_in_schema(target_cursor, 'XRT')
        target_conn.commit()
        xrt_sync_status['details'].append(
            f"🧹 Schéma XRT nettoyé en cible: {dropped} table(s) supprimée(s) (si présentes)"
        )

        tables = xrt_list_source_tables(source_cursor)
        tables = xrt_sort_tables_for_copy_cursor(source_cursor, tables)
        total = len(tables)
        xrt_sync_status['details'].append(f"📦 Tables XRT détectées: {total}")

        total_inserted = 0
        errors = 0

        for i, (schema_name, table_name) in enumerate(tables, start=1):
            xrt_sync_status['message'] = f"XRT: copie {schema_name}.{table_name}"
            xrt_sync_status['progress'] = int((i / max(1, total)) * 100)
            try:
                inserted = xrt_copy_table_full(source_cursor, target_cursor, schema_name, table_name, xrt_sync_status)
                target_conn.commit()
                total_inserted += inserted
                xrt_sync_status['details'].append(f"✓ {schema_name}.{table_name}: {inserted} ligne(s) copiée(s)")
            except Exception as e_tbl:
                target_conn.rollback()
                errors += 1
                xrt_sync_status['details'].append(f"✗ {schema_name}.{table_name}: {str(e_tbl)[:220]}")

        ended_at = datetime.now()
        xrt_sync_status['details'].append("")
        xrt_sync_status['details'].append(f"🕒 Début: {started_at.strftime('%Y-%m-%d %H:%M:%S')}")
        xrt_sync_status['details'].append(f"🕒 Fin:   {ended_at.strftime('%Y-%m-%d %H:%M:%S')}")
        xrt_sync_status['details'].append(f"📌 Total lignes copiées: {total_inserted}")
        xrt_sync_status['details'].append(f"⚠ Erreurs: {errors}")

        xrt_sync_status['message'] = 'Synchronisation XRT terminée'
        xrt_sync_status['progress'] = 100

        try:
            source_conn.close()
        except:
            pass
        try:
            target_conn.close()
        except:
            pass
    except Exception as e:
        xrt_sync_status['message'] = f'Erreur XRT: {str(e)}'
        xrt_sync_status['details'].append(f"Erreur globale XRT: {str(e)}")
    finally:
        xrt_sync_status['running'] = False


def xrt_sync_databases_missing_only():
    """
    Copie uniquement les tables source dont la table XRT correspondante n'existe pas encore en cible.
    Ne supprime pas le schéma XRT ni les tables déjà synchronisées (reprise après synchro interrompue).
    """
    global xrt_sync_status
    xrt_sync_status = {
        'running': True,
        'message': 'Analyse des tables XRT manquantes...',
        'progress': 0,
        'details': [],
        'code_version': PROJET21_CODE_VERSION,
    }
    xrt_sync_status['details'].append(
        f"🧩 Version code Projet21: {PROJET21_CODE_VERSION} ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})"
    )
    xrt_sync_status['details'].append("🔒 Source XRT: lecture seule (aucune écriture).")
    xrt_sync_status['details'].append("➕ Mode: compléter uniquement les tables absentes en XRT (pas de nettoyage global).")
    started_at = datetime.now()

    try:
        xrt_cfg = get_xrt_source_runtime_config()
        source_conn = get_connection(xrt_cfg, readonly=True)
        source_cursor = source_conn.cursor()

        target_conn = get_connection(TARGET_CONFIG)
        target_cursor = target_conn.cursor()
        ensure_target_schema(target_cursor, 'XRT')
        target_conn.commit()

        tables = xrt_list_source_tables(source_cursor)
        tables = xrt_sort_tables_for_copy_cursor(source_cursor, tables)

        missing_list = []
        existing_n = 0
        for schema_name, table_name in tables:
            tgt_name = xrt_target_table_name(schema_name, table_name)
            if xrt_target_table_exists(target_cursor, 'XRT', tgt_name):
                existing_n += 1
            else:
                missing_list.append((schema_name, table_name))

        xrt_sync_status['details'].append(
            f"📊 Tables source: {len(tables)} | Déjà présentes en XRT: {existing_n} | À copier: {len(missing_list)}"
        )

        total = len(missing_list)
        total_inserted = 0
        errors = 0

        if total == 0:
            xrt_sync_status['message'] = 'Aucune table manquante'
            xrt_sync_status['progress'] = 100
        else:
            for i, (schema_name, table_name) in enumerate(missing_list, start=1):
                xrt_sync_status['message'] = f"XRT (complément): {schema_name}.{table_name}"
                xrt_sync_status['progress'] = int((i / max(1, total)) * 100)
                try:
                    inserted = xrt_copy_table_full(
                        source_cursor, target_cursor, schema_name, table_name, xrt_sync_status
                    )
                    target_conn.commit()
                    total_inserted += inserted
                    xrt_sync_status['details'].append(
                        f"✓ {schema_name}.{table_name}: {inserted} ligne(s) copiée(s)"
                    )
                except Exception as e_tbl:
                    target_conn.rollback()
                    errors += 1
                    xrt_sync_status['details'].append(
                        f"✗ {schema_name}.{table_name}: {str(e_tbl)[:220]}"
                    )

            xrt_sync_status['message'] = 'Complément XRT terminé'
            xrt_sync_status['progress'] = 100

        ended_at = datetime.now()
        xrt_sync_status['details'].append("")
        xrt_sync_status['details'].append(f"🕒 Début: {started_at.strftime('%Y-%m-%d %H:%M:%S')}")
        xrt_sync_status['details'].append(f"🕒 Fin:   {ended_at.strftime('%Y-%m-%d %H:%M:%S')}")
        xrt_sync_status['details'].append(f"📌 Lignes copiées (ce run): {total_inserted}")
        xrt_sync_status['details'].append(f"⚠ Erreurs: {errors}")

        try:
            source_conn.close()
        except Exception:
            pass
        try:
            target_conn.close()
        except Exception:
            pass
    except Exception as e:
        xrt_sync_status['message'] = f'Erreur XRT (complément): {str(e)}'
        xrt_sync_status['details'].append(f"Erreur globale XRT: {str(e)}")
    finally:
        xrt_sync_status['running'] = False


def xrt_sync_databases_count_mismatch_only():
    """
    Recopie (DROP + CREATE + données) uniquement les tables déjà présentes en XRT
    dont COUNT(source) diffère de COUNT(cible). Utile après une vérif « KO comptes différents »
    sans tout resynchroniser.
    """
    global xrt_sync_status
    xrt_sync_status = {
        'running': True,
        'message': 'Analyse des écarts de comptage XRT…',
        'progress': 0,
        'details': [],
        'code_version': PROJET21_CODE_VERSION,
    }
    xrt_sync_status['details'].append(
        f"🧩 Version code Projet21: {PROJET21_CODE_VERSION} ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})"
    )
    xrt_sync_status['details'].append("🔒 Source XRT: lecture seule (aucune écriture).")
    xrt_sync_status['details'].append(
        "🔧 Mode: recopier uniquement les tables avec COUNT(source) ≠ COUNT(cible) dans le schéma XRT."
    )
    started_at = datetime.now()

    try:
        xrt_cfg = get_xrt_source_runtime_config()
        source_conn = get_connection(xrt_cfg, readonly=True)
        source_cursor = source_conn.cursor()

        target_conn = get_connection(TARGET_CONFIG)
        target_cursor = target_conn.cursor()
        ensure_target_schema(target_cursor, 'XRT')
        target_conn.commit()

        tables = xrt_list_source_tables(source_cursor)

        mismatch_list = []
        for schema_name, table_name in tables:
            tgt_name = xrt_target_table_name(schema_name, table_name)
            try:
                source_cursor.execute(f"SELECT COUNT(*) FROM [{schema_name}].[{table_name}]")
                src_count = int(source_cursor.fetchone()[0] or 0)
                if not xrt_target_table_exists(target_cursor, 'XRT', tgt_name):
                    continue
                target_cursor.execute(f"SELECT COUNT(*) FROM [XRT].[{tgt_name}]")
                tgt_count = int(target_cursor.fetchone()[0] or 0)
                if src_count != tgt_count:
                    mismatch_list.append((schema_name, table_name, src_count, tgt_count))
            except Exception:
                continue

        mismatch_sorted = [
            (s, t) for s, t, _sc, _tc in mismatch_list
        ]
        mismatch_sorted = xrt_sort_tables_for_copy_cursor(source_cursor, mismatch_sorted)

        xrt_sync_status['details'].append(
            f"📊 Tables avec écart de comptage: {len(mismatch_list)} "
            f"(sur {len(tables)} tables source)"
        )
        for schema_name, table_name, sc, tc in mismatch_list:
            tgt_name = xrt_target_table_name(schema_name, table_name)
            xrt_sync_status['details'].append(
                f"   • {schema_name}.{table_name} -> XRT.{tgt_name}: source={sc} cible={tc}"
            )

        total = len(mismatch_sorted)
        total_inserted = 0
        errors = 0

        if total == 0:
            xrt_sync_status['message'] = 'Aucun écart de comptage'
            xrt_sync_status['progress'] = 100
        else:
            for i, (schema_name, table_name) in enumerate(mismatch_sorted, start=1):
                xrt_sync_status['message'] = f"XRT (réalignement comptage): {schema_name}.{table_name}"
                xrt_sync_status['progress'] = int((i / max(1, total)) * 100)
                try:
                    inserted = xrt_copy_table_full(
                        source_cursor, target_cursor, schema_name, table_name, xrt_sync_status
                    )
                    target_conn.commit()
                    total_inserted += inserted
                    xrt_sync_status['details'].append(
                        f"✓ {schema_name}.{table_name}: {inserted} ligne(s) recopiée(s)"
                    )
                except Exception as e_tbl:
                    target_conn.rollback()
                    errors += 1
                    xrt_sync_status['details'].append(
                        f"✗ {schema_name}.{table_name}: {str(e_tbl)[:220]}"
                    )

            xrt_sync_status['message'] = 'Réalignement comptages XRT terminé'
            xrt_sync_status['progress'] = 100

        ended_at = datetime.now()
        xrt_sync_status['details'].append("")
        xrt_sync_status['details'].append(f"🕒 Début: {started_at.strftime('%Y-%m-%d %H:%M:%S')}")
        xrt_sync_status['details'].append(f"🕒 Fin:   {ended_at.strftime('%Y-%m-%d %H:%M:%S')}")
        xrt_sync_status['details'].append(f"📌 Lignes copiées (ce run): {total_inserted}")
        xrt_sync_status['details'].append(f"⚠ Erreurs: {errors}")

        try:
            source_conn.close()
        except Exception:
            pass
        try:
            target_conn.close()
        except Exception:
            pass
    except Exception as e:
        xrt_sync_status['message'] = f'Erreur XRT (réalignement): {str(e)}'
        xrt_sync_status['details'].append(f"Erreur globale XRT: {str(e)}")
    finally:
        xrt_sync_status['running'] = False


def xrt_sync_databases_incremental():
    """
    Synchronisation XRT type « miroir » sans vider le schéma :
    - table absente en XRT : création + copie complète (flux fetchmany) ;
    - table présente : détection d’écart COUNT puis CHECKSUM_AGG(BINARY_CHECKSUM(*)) ;
      si écart et volume ≤ XRT_MIRROR_MAX_ROWS_IN_MEMORY : INSERT / UPDATE en mémoire (PK requise) ;
      sinon : recopie complète de la table (xrt_copy_table_full) ;
    - phase finale : suppression (ordre inverse FK) des lignes cible dont la PK n’existe plus en source.

    Source strictement en lecture seule.
    """
    global xrt_sync_status
    xrt_sync_status = {
        'running': True,
        'message': 'Analyse XRT (détection + miroir)…',
        'progress': 0,
        'details': [],
        'code_version': PROJET21_CODE_VERSION,
    }
    xrt_sync_status['details'].append(
        f"🧩 Version code Projet21: {PROJET21_CODE_VERSION} ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})"
    )
    xrt_sync_status['details'].append("🔒 Source XRT: lecture seule (aucune écriture).")
    xrt_sync_status['details'].append(
        f"📌 Mode: miroir cible — détection COUNT + checksum ; plafond mémoire "
        f"{XRT_MIRROR_MAX_ROWS_IN_MEMORY} lignes/table (env XRT_MIRROR_MAX_ROWS_IN_MEMORY)."
    )
    started_at = datetime.now()

    total_inserted = 0
    total_updated = 0
    full_copies = 0
    tables_aligned = 0
    orphans_deleted_total = 0
    errors = 0

    try:
        xrt_cfg = get_xrt_source_runtime_config()
        source_conn = get_connection(xrt_cfg, readonly=True)
        source_cursor = source_conn.cursor()

        target_conn = get_connection(TARGET_CONFIG)
        target_cursor = target_conn.cursor()
        ensure_target_schema(target_cursor, 'XRT')
        target_conn.commit()

        tables = xrt_list_source_tables(source_cursor)
        sorted_tables = xrt_sort_tables_for_copy_cursor(source_cursor, tables)
        n_tables = len(sorted_tables)

        for i, (schema_name, table_name) in enumerate(sorted_tables, start=1):
            xrt_sync_status['message'] = f"XRT: {schema_name}.{table_name}"
            xrt_sync_status['progress'] = int((i / max(1, n_tables)) * 88)
            tgt_name = xrt_target_table_name(schema_name, table_name)

            try:
                if not xrt_target_table_exists(target_cursor, 'XRT', tgt_name):
                    ins = xrt_copy_table_full(
                        source_cursor, target_cursor, schema_name, table_name, xrt_sync_status
                    )
                    target_conn.commit()
                    total_inserted += ins
                    full_copies += 1
                    xrt_sync_status['details'].append(
                        f"✓ {schema_name}.{table_name}: nouvelle table → {ins} ligne(s) copiée(s)"
                    )
                    continue

                src_n = xrt_count_rows(source_cursor, schema_name, table_name)
                tgt_n = xrt_count_rows(target_cursor, 'XRT', tgt_name)

                if src_n == 0 and tgt_n == 0:
                    tables_aligned += 1
                    xrt_sync_status['details'].append(
                        f"○ {schema_name}.{table_name}: vide source/cible — inchangé"
                    )
                    continue

                if src_n == 0 and tgt_n > 0:
                    xrt_sync_status['details'].append(
                        f"   … {schema_name}.{table_name}: source vide, {tgt_n} ligne(s) en cible "
                        f"— suppression orphelins en fin de synchro"
                    )
                    continue

                need_sync = False
                reason = ""
                if src_n != tgt_n:
                    need_sync = True
                    reason = f"comptage src={src_n} ≠ cible={tgt_n}"
                else:
                    cs_s = xrt_table_checksum_agg(source_cursor, schema_name, table_name)
                    cs_t = xrt_table_checksum_agg(target_cursor, 'XRT', tgt_name)
                    if cs_s is None or cs_t is None:
                        need_sync = True
                        reason = "checksum indisponible — synchro forcée"
                    elif cs_s != cs_t:
                        need_sync = True
                        reason = "checksum différent (même comptage)"

                if not need_sync:
                    tables_aligned += 1
                    xrt_sync_status['details'].append(
                        f"○ {schema_name}.{table_name}: COUNT et checksum alignés — inchangé"
                    )
                    continue

                pk_cols = get_primary_keys_scoped(source_cursor, schema_name, table_name)

                use_full_copy = (not pk_cols) or (src_n > XRT_MIRROR_MAX_ROWS_IN_MEMORY)

                if use_full_copy:
                    why_bits = []
                    if not pk_cols:
                        why_bits.append("sans PK")
                    if pk_cols and src_n > XRT_MIRROR_MAX_ROWS_IN_MEMORY:
                        why_bits.append(f">{XRT_MIRROR_MAX_ROWS_IN_MEMORY} lignes")
                    xrt_sync_status['details'].append(
                        f"   … {schema_name}.{table_name}: écart ({reason}) → recopie complète "
                        f"({', '.join(why_bits)})"
                    )
                    try:
                        ins = xrt_copy_table_full(
                            source_cursor, target_cursor, schema_name, table_name, xrt_sync_status
                        )
                        target_conn.commit()
                        total_inserted += ins
                        full_copies += 1
                        xrt_sync_status['details'].append(
                            f"✓ {schema_name}.{table_name}: recopie complète {ins} ligne(s)"
                        )
                    except Exception as e_full:
                        target_conn.rollback()
                        errors += 1
                        xrt_sync_status['details'].append(
                            f"✗ {schema_name}.{table_name} (recopie): {str(e_full)[:220]}"
                        )
                    continue

                xrt_sync_status['details'].append(
                    f"   … {schema_name}.{table_name}: écart ({reason}) → miroir mémoire "
                    f"(≤{XRT_MIRROR_MAX_ROWS_IN_MEMORY} lignes)"
                )
                try:
                    ins, upd = xrt_sync_table_mirror_memory(
                        source_cursor,
                        target_cursor,
                        schema_name,
                        table_name,
                        'XRT',
                        tgt_name,
                    )
                    target_conn.commit()
                    total_inserted += ins
                    total_updated += upd
                    xrt_sync_status['details'].append(
                        f"✓ {schema_name}.{table_name}: miroir {ins} inséré(s), {upd} mis(e)(s) à jour"
                    )
                except Exception as e_m:
                    target_conn.rollback()
                    xrt_sync_status['details'].append(
                        f"   ⚠ {schema_name}.{table_name}: miroir échoué ({str(e_m)[:120]}) "
                        f"— tentative recopie complète"
                    )
                    try:
                        ins = xrt_copy_table_full(
                            source_cursor, target_cursor, schema_name, table_name, xrt_sync_status
                        )
                        target_conn.commit()
                        total_inserted += ins
                        full_copies += 1
                        xrt_sync_status['details'].append(
                            f"✓ {schema_name}.{table_name}: recopie de secours {ins} ligne(s)"
                        )
                    except Exception as e2:
                        target_conn.rollback()
                        errors += 1
                        xrt_sync_status['details'].append(
                            f"✗ {schema_name}.{table_name}: {str(e2)[:220]}"
                        )

            except Exception as e_tbl:
                try:
                    target_conn.rollback()
                except Exception:
                    pass
                errors += 1
                xrt_sync_status['details'].append(
                    f"✗ {schema_name}.{table_name}: {str(e_tbl)[:220]}"
                )

        xrt_sync_status['message'] = "XRT: suppression des orphelins cible…"
        xrt_sync_status['progress'] = 92
        xrt_sync_status['details'].append(
            "\n🧹 Lignes présentes en cible mais absentes de la source (ordre inverse FK)…"
        )

        for j, (schema_name, table_name) in enumerate(reversed(sorted_tables), start=1):
            xrt_sync_status['progress'] = 92 + int((j / max(1, n_tables)) * 7)
            tgt_name = xrt_target_table_name(schema_name, table_name)
            if not xrt_target_table_exists(target_cursor, 'XRT', tgt_name):
                continue
            pk_cols = get_primary_keys_scoped(source_cursor, schema_name, table_name)
            if not pk_cols:
                continue
            try:
                nd = xrt_delete_orphans_xrt(
                    source_cursor,
                    target_cursor,
                    target_conn,
                    schema_name,
                    table_name,
                    'XRT',
                    tgt_name,
                    pk_cols,
                    xrt_sync_status,
                )
                orphans_deleted_total += nd
                if nd > 0:
                    xrt_sync_status['details'].append(
                        f"🧹 {schema_name}.{table_name}: {nd} ligne(s) orpheline(s) supprimée(s)"
                    )
            except Exception:
                pass

        if orphans_deleted_total == 0:
            xrt_sync_status['details'].append("○ Aucune ligne orpheline supprimée")

        ended_at = datetime.now()
        xrt_sync_status['details'].append("")
        xrt_sync_status['details'].append(f"🕒 Début: {started_at.strftime('%Y-%m-%d %H:%M:%S')}")
        xrt_sync_status['details'].append(f"🕒 Fin:   {ended_at.strftime('%Y-%m-%d %H:%M:%S')}")
        xrt_sync_status['details'].append(
            f"📌 Résumé: tables inchangées={tables_aligned}, recopies complètes={full_copies}, "
            f"lignes insérées={total_inserted}, lignes mises à jour={total_updated}, "
            f"orphelins supprimés={orphans_deleted_total}"
        )
        xrt_sync_status['details'].append(f"⚠ Erreurs: {errors}")

        xrt_sync_status['message'] = 'Synchronisation XRT terminée'
        xrt_sync_status['progress'] = 100

        try:
            source_conn.close()
        except Exception:
            pass
        try:
            target_conn.close()
        except Exception:
            pass
    except Exception as e:
        xrt_sync_status['message'] = f'Erreur XRT (synchro miroir): {str(e)}'
        xrt_sync_status['details'].append(f"Erreur globale XRT: {str(e)}")
    finally:
        xrt_sync_status['running'] = False


def has_identity_column(cursor, table_name):
    """Vérifie si une table a une colonne IDENTITY"""
    cursor.execute("""
        SELECT COUNT(*) 
        FROM sys.columns c
        INNER JOIN sys.tables t ON c.object_id = t.object_id
        WHERE t.name = ? AND c.is_identity = 1
    """, (table_name,))
    return cursor.fetchone()[0] > 0

def extract_fk_info(error_msg):
    """Extrait les informations de FK depuis un message d'erreur SQL Server"""
    import re
    # Pattern: table "dbo.TABLE_NAME", column 'COLUMN_NAME'
    pattern = r'table\s+"[^"]*\.([^"]+)"[^,]*,\s*column\s+\'([^\']+)\''
    match = re.search(pattern, error_msg, re.IGNORECASE)
    if match:
        return match.group(1), match.group(2)
    return None, None

def check_fk_exists(cursor, ref_table, ref_column, value):
    """
    Vérifie si une valeur existe dans la table référencée.
    Gère spécialement la valeur 0 pour garantir qu'elle est toujours vérifiée correctement.
    """
    try:
        # Pour la valeur 0, vérifier explicitement (car 0 peut être valide même si non présent dans un DISTINCT)
        if value == 0:
            cursor.execute(f"SELECT COUNT(*) FROM [{ref_table}] WHERE [{ref_column}] = 0")
            return cursor.fetchone()[0] > 0
        else:
            cursor.execute(f"SELECT COUNT(*) FROM [{ref_table}] WHERE [{ref_column}] = ?", (value,))
            return cursor.fetchone()[0] > 0
    except:
        return False

def get_foreign_keys(cursor, table_name):
    """Récupère les contraintes FK d'une table"""
    cursor.execute("""
        SELECT 
            fk.name AS FK_Name,
            tp.name AS Parent_Table,
            cp.name AS Parent_Column,
            tr.name AS Referenced_Table,
            cr.name AS Referenced_Column
        FROM sys.foreign_keys AS fk
        INNER JOIN sys.foreign_key_columns AS fkc ON fk.object_id = fkc.constraint_object_id
        INNER JOIN sys.tables AS tp ON fkc.parent_object_id = tp.object_id
        INNER JOIN sys.columns AS cp ON fkc.parent_object_id = cp.object_id AND fkc.parent_column_id = cp.column_id
        INNER JOIN sys.tables AS tr ON fkc.referenced_object_id = tr.object_id
        INNER JOIN sys.columns AS cr ON fkc.referenced_object_id = cr.object_id AND fkc.referenced_column_id = cr.column_id
        WHERE tp.name = ?
    """, (table_name,))
    return cursor.fetchall()

def get_table_definition(cursor, table_name):
    columns = get_table_columns(cursor, table_name)
    pk_columns = get_primary_keys(cursor, table_name)
    
    col_defs = []
    for col in columns:
        col_name, data_type, max_len, nullable, default = col
        col_def = f"[{col_name}] {data_type}"
        if max_len and data_type in ('varchar', 'nvarchar', 'char', 'nchar'):
            col_def += f"({max_len if max_len != -1 else 'MAX'})"
        if nullable == 'NO':
            col_def += " NOT NULL"
        col_defs.append(col_def)
    
    create_sql = f"CREATE TABLE [{table_name}] (\n  " + ",\n  ".join(col_defs)
    if pk_columns:
        create_sql += f",\n  PRIMARY KEY ([" + "], [".join(pk_columns) + "])"
    create_sql += "\n)"
    return create_sql

def build_dependency_graph(source_cursor, target_cursor, tables):
    """
    Construit un graphe de dépendances FK : {table: [tables_dont_elle_depend]}
    Utilise la SOURCE pour détecter les FK (plus fiable que la cible qui peut être incomplète)
    Retourne aussi un set de toutes les tables référencées (pour détecter les tables externes)
    """
    graph = {table: [] for table in tables}
    all_referenced_tables = set()
    
    for table_name in tables:
        try:
            # Utiliser la SOURCE pour détecter les FK (plus fiable)
            fks = get_foreign_keys(source_cursor, table_name)
            for fk_info in fks:
                fk_name, parent_table, parent_col, ref_table, ref_col = fk_info
                # Ne considérer que les FK vers d'autres tables de notre liste
                if ref_table in tables and ref_table != table_name:
                    if ref_table not in graph[table_name]:
                        graph[table_name].append(ref_table)
                    all_referenced_tables.add(ref_table)
        except Exception:
            # Si erreur, considérer la table comme sans dépendances
            pass
    
    return graph, all_referenced_tables

def topological_sort(tables, dependency_graph):
    """
    Tri topologique pour déterminer l'ordre de synchronisation.
    Les tables sans dépendances sont traitées en premier.
    Gère les cycles en les plaçant à la fin.
    """
    # Copie du graphe pour ne pas le modifier
    graph = {table: list(deps) for table, deps in dependency_graph.items()}
    
    # Calculer le nombre de dépendances SORTANTES (dépendances de cette table)
    # Pour le tri topologique, on veut traiter d'abord les tables avec le moins de dépendances
    out_degree = {table: len(graph.get(table, [])) for table in tables}
    
    # Calculer aussi les dépendances ENTRANTES (combien de tables dépendent de cette table)
    in_degree = {table: 0 for table in tables}
    for table in tables:
        for dep in graph.get(table, []):
            if dep in in_degree:
                in_degree[dep] += 1
    
    # Tables sans dépendances SORTANTES (ne dépendent de rien) = prioritaires
    queue = [table for table in tables if out_degree[table] == 0]
    result = []
    processed = set()
    
    # Traiter d'abord les tables sans dépendances
    while queue:
        queue.sort()  # Ordre déterministe
        current = queue.pop(0)
        if current in processed:
            continue
        result.append(current)
        processed.add(current)
        
        # Réduire le degré des tables qui dépendent de current
        for table in tables:
            if current in graph.get(table, []):
                out_degree[table] -= 1
                if out_degree[table] == 0 and table not in processed:
                    queue.append(table)
    
    # Ajouter les tables restantes (cycles ou dépendances complexes)
    remaining = [table for table in tables if table not in processed]
    if remaining:
        # Trier par nombre de dépendances (moins de dépendances = prioritaire)
        remaining.sort(key=lambda t: (out_degree[t], t))
        result.extend(remaining)
    
    return result

def find_row_by_unique_index(target_cursor, table_name, unique_index_cols, values):
    """
    Trouve une ligne existante dans la cible via un index unique (hors PK).
    Retourne la ligne complète si trouvée, None sinon.
    Gère correctement les NULL avec IS NULL.
    """
    if not unique_index_cols or len(unique_index_cols) != len(values):
        return None
    
    # Gérer correctement les NULL: "=" ne matche pas NULL -> utiliser IS NULL
    where_parts = []
    params = []
    for col, val in zip(unique_index_cols, values):
        if val is None:
            where_parts.append(f"[{col}] IS NULL")
        else:
            # Pour les chaînes, normaliser les espaces et la casse si nécessaire
            where_parts.append(f"[{col}] = ?")
            params.append(val)
    
    if not where_parts:
        return None
    
    where_clause = " AND ".join(where_parts)
    try:
        target_cursor.execute(f"SELECT * FROM [{table_name}] WHERE {where_clause}", tuple(params))
        row = target_cursor.fetchone()
        if row:
            # Récupérer les noms de colonnes (une seule fois pour optimiser)
            try:
                # Utiliser la description de la requête précédente si disponible
                columns = [desc[0] for desc in target_cursor.description]
            except:
                # Fallback : requête vide pour obtenir les colonnes
                target_cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
                columns = [desc[0] for desc in target_cursor.description]
            return {columns[i]: row[i] for i in range(len(columns))}
    except Exception as e:
        # En cas d'erreur, retourner None (la ligne n'existe peut-être pas)
        pass
    return None

def compare_rows_for_update(source_row_dict, target_row_dict, exclude_cols=None):
    """
    Compare deux lignes et retourne les colonnes qui diffèrent (hors colonnes exclues).
    Retourne (has_diff, diff_cols, update_values) où update_values est un dict {col: source_value}
    """
    if exclude_cols is None:
        exclude_cols = set()
    
    diff_cols = []
    update_values = {}
    
    for col in source_row_dict:
        if col in exclude_cols:
            continue
        source_val = source_row_dict.get(col)
        target_val = target_row_dict.get(col)
        
        # Comparaison robuste (gérer None, types différents)
        if source_val != target_val:
            # Vérifier si c'est vraiment différent (pas juste type)
            if source_val is None and target_val is None:
                continue
            if source_val is None or target_val is None:
                diff_cols.append(col)
                update_values[col] = source_val
            else:
                # Comparaison de valeurs
                try:
                    # Pour les nombres, comparer numériquement
                    if isinstance(source_val, (int, float)) and isinstance(target_val, (int, float)):
                        if abs(float(source_val) - float(target_val)) > 1e-10:
                            diff_cols.append(col)
                            update_values[col] = source_val
                    elif str(source_val) != str(target_val):
                        diff_cols.append(col)
                        update_values[col] = source_val
                except Exception:
                    # En cas de doute, considérer comme différent
                    diff_cols.append(col)
                    update_values[col] = source_val
    
    return len(diff_cols) > 0, diff_cols, update_values


def xrt_delete_orphans_xrt(
    source_cursor,
    target_cursor,
    target_conn,
    source_schema,
    source_table,
    target_schema,
    target_table,
    pk_columns,
    sync_status,
):
    """
    Supprime en schéma XRT les lignes dont la PK n’existe plus dans la source (lecture seule source).
    """
    deleted = 0
    pk_select = ", ".join([f"[{pk}]" for pk in pk_columns])
    try:
        source_cursor.execute(f"SELECT {pk_select} FROM [{source_schema}].[{source_table}]")
        source_pks = set()
        for row in source_cursor.fetchall():
            if len(pk_columns) == 1:
                source_pks.add(row[0])
            else:
                source_pks.add(tuple(row))

        target_cursor.execute(f"SELECT {pk_select} FROM [{target_schema}].[{target_table}]")
        target_pk_list = []
        for row in target_cursor.fetchall():
            if len(pk_columns) == 1:
                target_pk_list.append(row[0])
            else:
                target_pk_list.append(tuple(row))

        orphans = [pk for pk in target_pk_list if pk not in source_pks]
        if not orphans:
            return 0

        batch_size = 500
        if len(pk_columns) == 1:
            col = pk_columns[0]
            for i in range(0, len(orphans), batch_size):
                batch = orphans[i : i + batch_size]
                placeholders = ", ".join(["?"] * len(batch))
                target_cursor.execute(
                    f"DELETE FROM [{target_schema}].[{target_table}] WHERE [{col}] IN ({placeholders})",
                    tuple(batch),
                )
                rc = target_cursor.rowcount
                deleted += len(batch) if rc == -1 else rc
        else:
            where_parts = [f"[{pk}] = ?" for pk in pk_columns]
            where_sql = " AND ".join(where_parts)
            for pk in orphans:
                target_cursor.execute(
                    f"DELETE FROM [{target_schema}].[{target_table}] WHERE {where_sql}",
                    tuple(pk),
                )
                rc = target_cursor.rowcount
                deleted += 1 if rc == -1 else max(rc, 0)

        target_conn.commit()
        return deleted
    except Exception as e:
        err_str = str(e)
        try:
            target_conn.rollback()
        except Exception:
            pass
        sync_status["details"].append(
            f"  ⚠ Orphelins {source_schema}.{source_table}: {err_str[:200]}"
        )
        return 0


def xrt_sync_table_mirror_memory(
    source_cursor,
    target_cursor,
    source_schema,
    source_table,
    target_schema,
    tgt_name,
):
    """
    Alignement ligne à ligne (INSERT lignes nouvelles, UPDATE colonnes divergentes).
    Ne supprime pas les orphelins (phase séparée). Pas de commit ici.
    Retourne (inserted, updated).
    """
    pk_columns = get_primary_keys_scoped(source_cursor, source_schema, source_table)
    if not pk_columns:
        raise RuntimeError("PK introuvable pour synchro miroir mémoire")

    source_cursor.execute(f"SELECT * FROM [{source_schema}].[{source_table}] WHERE 1=0")
    columns = [d[0] for d in source_cursor.description]
    if not columns:
        return 0, 0

    col_list = ", ".join([f"[{c}]" for c in columns])
    placeholders = ", ".join(["?"] * len(columns))

    source_cursor.execute(f"SELECT {col_list} FROM [{source_schema}].[{source_table}]")
    source_rows = source_cursor.fetchall()

    target_cursor.execute(f"SELECT * FROM [{target_schema}].[{tgt_name}] WHERE 1=0")
    tgt_desc = target_cursor.description or []
    tgt_columns = [d[0] for d in tgt_desc]
    target_col_names = set(tgt_columns)

    target_cursor.execute(f"SELECT * FROM [{target_schema}].[{tgt_name}]")
    existing_rows_by_pk = {}
    for row in target_cursor.fetchall():
        row_dict = {tgt_columns[i]: row[i] for i in range(len(tgt_columns))}
        if len(pk_columns) == 1:
            pk_key = row_dict[pk_columns[0]]
        else:
            pk_key = tuple(row_dict[pk] for pk in pk_columns)
        existing_rows_by_pk[pk_key] = row_dict

    pk_indices = [columns.index(pk) for pk in pk_columns]
    exclude_cols = set(pk_columns)

    identity_on = has_identity_column_scoped(target_cursor, target_schema, tgt_name)
    inserted = 0
    updated = 0
    batch = []

    target_cursor.fast_executemany = True

    try:
        if identity_on:
            target_cursor.execute(
                f"SET IDENTITY_INSERT [{target_schema}].[{tgt_name}] ON"
            )

        for row in source_rows:
            row_dict = {columns[i]: row[i] for i in range(len(columns))}
            if len(pk_columns) == 1:
                pk_one = row[pk_indices[0]]
                pk_exists = pk_one in existing_rows_by_pk
                pk_for_where = pk_one
            else:
                pk_tuple = tuple(row[i] for i in pk_indices)
                pk_exists = pk_tuple in existing_rows_by_pk
                pk_for_where = pk_tuple

            if pk_exists:
                target_row_dict = existing_rows_by_pk[pk_for_where]
                has_diff, diff_cols, update_values = compare_rows_for_update(
                    row_dict, target_row_dict, exclude_cols
                )
                if has_diff:
                    diff_cols = [c for c in diff_cols if c in target_col_names]
                    update_values = {
                        c: update_values[c] for c in diff_cols if c in update_values
                    }
                    if diff_cols:
                        if len(pk_columns) == 1:
                            where_pk = f"[{pk_columns[0]}] = ?"
                            set_clauses = [f"[{col}] = ?" for col in diff_cols]
                            set_values = [update_values[col] for col in diff_cols]
                            update_sql = (
                                f"UPDATE [{target_schema}].[{tgt_name}] SET {', '.join(set_clauses)} "
                                f"WHERE {where_pk}"
                            )
                            target_cursor.execute(
                                update_sql, tuple(set_values) + (pk_for_where,)
                            )
                        else:
                            where_parts = [f"[{pk}] = ?" for pk in pk_columns]
                            where_pk = " AND ".join(where_parts)
                            set_clauses = [f"[{col}] = ?" for col in diff_cols]
                            set_values = [update_values[col] for col in diff_cols]
                            update_sql = (
                                f"UPDATE [{target_schema}].[{tgt_name}] SET {', '.join(set_clauses)} "
                                f"WHERE {where_pk}"
                            )
                            target_cursor.execute(
                                update_sql, tuple(set_values) + pk_for_where
                            )
                        updated += 1
            else:
                batch.append(list(row))
                if len(batch) >= 100:
                    target_cursor.executemany(
                        f"INSERT INTO [{target_schema}].[{tgt_name}] ({col_list}) VALUES ({placeholders})",
                        batch,
                    )
                    inserted += len(batch)
                    batch = []

        if batch:
            target_cursor.executemany(
                f"INSERT INTO [{target_schema}].[{tgt_name}] ({col_list}) VALUES ({placeholders})",
                batch,
            )
            inserted += len(batch)

    finally:
        if identity_on:
            target_cursor.execute(
                f"SET IDENTITY_INSERT [{target_schema}].[{tgt_name}] OFF"
            )

    return inserted, updated


def sync_delete_orphan_rows_from_table(source_cursor, target_cursor, target_conn, table_name, pk_columns, sync_status):
    """
    Supprime en base cible les lignes dont la PK n'existe plus dans la source.
    Lecture seule sur la source. Écriture uniquement sur la cible.
    À appeler dans l'ordre inverse du tri topologique (enfants avant parents, FK).
    """
    deleted = 0
    pk_select = ", ".join([f"[{pk}]" for pk in pk_columns])
    try:
        source_cursor.execute(f"SELECT {pk_select} FROM [{table_name}]")
        source_pks = set()
        for row in source_cursor.fetchall():
            if len(pk_columns) == 1:
                source_pks.add(row[0])
            else:
                source_pks.add(tuple(row))

        target_cursor.execute(f"SELECT {pk_select} FROM [{table_name}]")
        target_pk_list = []
        for row in target_cursor.fetchall():
            if len(pk_columns) == 1:
                target_pk_list.append(row[0])
            else:
                target_pk_list.append(tuple(row))

        orphans = [pk for pk in target_pk_list if pk not in source_pks]
        if not orphans:
            return 0

        batch_size = 500
        if len(pk_columns) == 1:
            col = pk_columns[0]
            for i in range(0, len(orphans), batch_size):
                batch = orphans[i : i + batch_size]
                placeholders = ", ".join(["?"] * len(batch))
                target_cursor.execute(
                    f"DELETE FROM [{table_name}] WHERE [{col}] IN ({placeholders})",
                    tuple(batch),
                )
                rc = target_cursor.rowcount
                deleted += len(batch) if rc == -1 else rc
        else:
            where_parts = [f"[{pk}] = ?" for pk in pk_columns]
            where_sql = " AND ".join(where_parts)
            for pk in orphans:
                target_cursor.execute(
                    f"DELETE FROM [{table_name}] WHERE {where_sql}",
                    tuple(pk),
                )
                rc = target_cursor.rowcount
                deleted += 1 if rc == -1 else max(rc, 0)

        target_conn.commit()
        return deleted
    except Exception as e:
        err_str = str(e)
        try:
            target_conn.rollback()
        except Exception:
            pass
        if "F_COMPTET" in err_str or "regNOVA" in err_str:
            return 0
        sync_status["details"].append(f"  ⚠ Orphelins {table_name}: {err_str[:200]}")
        return 0


def auto_realign_table_ids(source_cursor, target_cursor, target_conn, table_name, sync_status):
    """
    Réaligne automatiquement les IDs d'une table si nécessaire.
    
    Pour PAPIERS_ARTICLES et PAPIERS_IMPRIMEURS uniquement.
    Compare les données par clé alternative (toutes colonnes sauf ID) et réaligne les IDs
    si des décalages sont détectés.
    
    Retourne True si un réalignement a été effectué, False sinon.
    """
    
    # Tables supportées pour le réalignement automatique
    SUPPORTED_TABLES = {
        'PAPIERS_ARTICLES': {
            'fk_tables': ['PAPIERS_TARIF_FMT'],  # Tables avec FK vers cette table
            'fk_columns': {'PAPIERS_TARIF_FMT': 'ID_ARTICLE'}  # Colonne FK dans chaque table enfant
        },
        'PAPIERS_IMPRIMEURS': {
            'fk_tables': ['PAPIERS_TARIF_FMT', 'PAPIERS_TARIF_GRAM'],
            'fk_columns': {
                'PAPIERS_TARIF_FMT': 'ID_PAPIMPRIM',
                'PAPIERS_TARIF_GRAM': 'ID_PAPIMPRIM'
            }
        }
    }
    
    if table_name not in SUPPORTED_TABLES:
        return False  # Table non supportée pour le réalignement automatique
    
    table_config = SUPPORTED_TABLES[table_name]
    
    try:
        sync_status['details'].append(f"🔍 Vérification du réalignement automatique pour {table_name}...")
        
        # 1. Vérifier que les deux tables existent
        source_cursor.execute("""
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_NAME = ? AND TABLE_TYPE = 'BASE TABLE'
        """, (table_name,))
        source_exists = source_cursor.fetchone()[0] > 0
        
        target_cursor.execute("""
            SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_NAME = ? AND TABLE_TYPE = 'BASE TABLE'
        """, (table_name,))
        target_exists = target_cursor.fetchone()[0] > 0
        
        if not source_exists or not target_exists:
            sync_status['details'].append(f"  ⚠ {table_name}: Table absente, réalignement ignoré")
            return False
        
        # 2. Récupérer toutes les colonnes (sauf ID) pour la correspondance
        source_cursor.execute("""
            SELECT COLUMN_NAME
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = ?
            AND COLUMN_NAME != 'ID'
            ORDER BY ORDINAL_POSITION
        """, (table_name,))
        source_columns = [row[0] for row in source_cursor.fetchall()]
        
        target_cursor.execute("""
            SELECT COLUMN_NAME
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = ?
            AND COLUMN_NAME != 'ID'
            ORDER BY ORDINAL_POSITION
        """, (table_name,))
        target_columns = [row[0] for row in target_cursor.fetchall()]
        
        # Utiliser uniquement les colonnes communes
        common_columns = [col for col in source_columns if col in target_columns]
        
        if not common_columns:
            sync_status['details'].append(f"  ⚠ {table_name}: Aucune colonne commune pour correspondance")
            return False
        
        # 3. Lire les données source et cible avec toutes les colonnes (sauf ID)
        columns_str = ', '.join([f'[{col}]' for col in common_columns])
        
        source_cursor.execute(f"SELECT ID, {columns_str} FROM [{table_name}]")
        source_rows = source_cursor.fetchall()
        source_data = {}  # {tuple(colonnes): ID_source}
        for row in source_rows:
            key = tuple(row[1:])  # Toutes les colonnes sauf ID
            source_data[key] = row[0]  # ID source
        
        target_cursor.execute(f"SELECT ID, {columns_str} FROM [{table_name}]")
        target_rows = target_cursor.fetchall()
        
        # 4. Créer le mapping des IDs décalés
        id_mapping = []  # [(ancien_ID_cible, nouveau_ID_source)]
        target_ids_by_key = {}  # {tuple(colonnes): ID_cible}
        
        for row in target_rows:
            target_id = row[0]
            key = tuple(row[1:])
            target_ids_by_key[key] = target_id
            
            if key in source_data:
                source_id = source_data[key]
                if target_id != source_id:
                    id_mapping.append((target_id, source_id))
        
        if not id_mapping:
            sync_status['details'].append(f"  ✓ {table_name}: Aucun décalage d'ID détecté")
            return False
        
        sync_status['details'].append(f"  🔄 {table_name}: {len(id_mapping)} IDs à réaligner détectés")
        
        # 5. Vérifier les conflits (nouveaux IDs qui existent déjà hors mapping)
        conflict_count = 0
        new_ids = {new_id for _, new_id in id_mapping}
        old_ids = {old_id for old_id, _ in id_mapping}
        
        target_cursor.execute(f"SELECT ID FROM [{table_name}]")
        existing_ids = {row[0] for row in target_cursor.fetchall()}
        
        conflicts = new_ids & (existing_ids - old_ids)
        if conflicts:
            sync_status['details'].append(f"  ⚠ {table_name}: {len(conflicts)} conflits d'IDs détectés - réalignement partiel uniquement")
            # Filtrer le mapping pour exclure les conflits
            id_mapping = [(old_id, new_id) for old_id, new_id in id_mapping if new_id not in conflicts]
            if not id_mapping:
                sync_status['details'].append(f"  ✗ {table_name}: Impossible de réaligner (tous les IDs sont en conflit)")
                return False
        
        # 6. Démarrer la transaction
        target_conn.autocommit = False
        
        try:
            # 6a. Désactiver temporairement les FK entrantes
            fk_disabled = []
            for fk_table in table_config['fk_tables']:
                fk_col = table_config['fk_columns'][fk_table]
                
                # Trouver le nom de la contrainte FK
                target_cursor.execute("""
                    SELECT fk.name
                    FROM sys.foreign_keys fk
                    INNER JOIN sys.foreign_key_columns fkc ON fk.object_id = fkc.constraint_object_id
                    INNER JOIN sys.tables tp ON fkc.parent_object_id = tp.object_id
                    INNER JOIN sys.columns cp ON fkc.parent_object_id = cp.object_id AND fkc.parent_column_id = cp.column_id
                    INNER JOIN sys.tables tr ON fkc.referenced_object_id = tr.object_id
                    WHERE tp.name = ? AND cp.name = ? AND tr.name = ?
                """, (fk_table, fk_col, table_name))
                
                fk_row = target_cursor.fetchone()
                if fk_row:
                    fk_name = fk_row[0]
                    target_cursor.execute(f"ALTER TABLE [{fk_table}] NOCHECK CONSTRAINT [{fk_name}]")
                    fk_disabled.append((fk_table, fk_name))
                    sync_status['details'].append(f"    ✓ FK désactivée: {fk_table}.{fk_col}")
            
            # 6b. Mettre à jour les FK dans les tables enfants
            # IMPORTANT : Gérer les conflits de PK qui peuvent survenir lors de la mise à jour
            fk_updated_total = 0
            for fk_table in table_config['fk_tables']:
                fk_col = table_config['fk_columns'][fk_table]
                
                # Utiliser une copie du mapping pour chaque table (pour éviter de modifier le mapping original)
                table_id_mapping = list(id_mapping)
                
                # Construire le mapping pour la mise à jour avec gestion des conflits
                update_count = 0
                conflicts_count = 0
                for old_id, new_id in table_id_mapping:
                    try:
                        target_cursor.execute(f"""
                            UPDATE [{fk_table}]
                            SET [{fk_col}] = ?
                            WHERE [{fk_col}] = ?
                        """, (new_id, old_id))
                        update_count += target_cursor.rowcount
                    except Exception as update_err:
                        err_str = str(update_err)
                        if 'duplicate key' in err_str.lower() or 'primary key' in err_str.lower() or '23000' in err_str:
                            # Conflit détecté lors de la mise à jour (plusieurs lignes avec le même nouveau ID créent un doublon de PK)
                            conflicts_count += 1
                            if conflicts_count <= 3:  # Limiter les messages
                                sync_status['details'].append(f"    ⚠ Conflit PK dans {fk_table}: old_id={old_id}->new_id={new_id} ({err_str[:100]})")
                            continue
                        else:
                            raise  # Répercuter les autres erreurs
                
                fk_updated_total += update_count
                if update_count > 0:
                    sync_status['details'].append(f"    ✓ {update_count} références FK mises à jour dans {fk_table}")
                if conflicts_count > 0:
                    sync_status['details'].append(f"    ⚠ {conflicts_count} références FK non mises à jour dans {fk_table} à cause de conflits de PK")
            
            # 6c. Modifier les IDs dans la table principale
            # Vérifier et activer IDENTITY_INSERT si nécessaire
            identity_enabled = False
            target_cursor.execute("""
                SELECT COUNT(*) 
                FROM sys.columns c
                INNER JOIN sys.tables t ON c.object_id = t.object_id
                WHERE t.name = ? AND c.is_identity = 1
            """, (table_name,))
            if target_cursor.fetchone()[0] > 0:
                target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] ON")
                identity_enabled = True
            
            # Récupérer toutes les colonnes de la table
            target_cursor.execute("""
                SELECT COLUMN_NAME
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_NAME = ?
                ORDER BY ORDINAL_POSITION
            """, (table_name,))
            all_columns = [row[0] for row in target_cursor.fetchall()]
            columns_insert = ', '.join([f'[{col}]' for col in all_columns])
            
            # Créer une table temporaire pour stocker le mapping
            target_cursor.execute(f"""
                CREATE TABLE #TEMP_MAPPING_{table_name} (
                    ancien_ID INT PRIMARY KEY,
                    nouveau_ID INT NOT NULL
                )
            """)
            
            # Insérer le mapping
            for old_id, new_id in id_mapping:
                target_cursor.execute(f"""
                    INSERT INTO #TEMP_MAPPING_{table_name} (ancien_ID, nouveau_ID)
                    VALUES (?, ?)
                """, (old_id, new_id))
            
            # Créer une table temporaire avec les enregistrements à modifier (avec nouveaux IDs)
            target_cursor.execute(f"""
                SELECT pa.*, m.nouveau_ID AS NEW_ID
                INTO #TEMP_REALIGN_{table_name}
                FROM [{table_name}] pa
                INNER JOIN #TEMP_MAPPING_{table_name} m ON pa.ID = m.ancien_ID
            """)
            
            # Mettre à jour les IDs dans la table temporaire
            target_cursor.execute(f"""
                UPDATE #TEMP_REALIGN_{table_name}
                SET ID = NEW_ID
            """)
            
            # Supprimer les anciens enregistrements de la table principale
            old_ids_list = [old_id for old_id, _ in id_mapping]
            placeholders = ','.join(['?' for _ in old_ids_list])
            target_cursor.execute(f"""
                DELETE FROM [{table_name}]
                WHERE ID IN ({placeholders})
            """, old_ids_list)
            
            # Insérer les enregistrements avec les nouveaux IDs
            target_cursor.execute(f"""
                INSERT INTO [{table_name}] ({columns_insert})
                SELECT {columns_insert}
                FROM #TEMP_REALIGN_{table_name}
            """)
            
            # Nettoyer les tables temporaires
            target_cursor.execute(f"DROP TABLE #TEMP_REALIGN_{table_name}")
            target_cursor.execute(f"DROP TABLE #TEMP_MAPPING_{table_name}")
            
            # Réactiver IDENTITY si nécessaire
            if identity_enabled:
                target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] OFF")
                # Réinitialiser IDENTITY
                target_cursor.execute(f"SELECT MAX(ID) FROM [{table_name}]")
                max_id = target_cursor.fetchone()[0] or 0
                target_cursor.execute(f"DBCC CHECKIDENT ('[{table_name}]', RESEED, {max_id})")
            
            # 6d. Réactiver les FK entrantes
            for fk_table, fk_name in fk_disabled:
                target_cursor.execute(f"ALTER TABLE [{fk_table}] CHECK CONSTRAINT [{fk_name}]")
            
            # Valider la transaction
            target_conn.commit()
            target_conn.autocommit = True
            
            sync_status['details'].append(f"  ✓ {table_name}: {len(id_mapping)} IDs réalisés, {fk_updated_total} FK mises à jour")
            return True
            
        except Exception as e:
            target_conn.rollback()
            target_conn.autocommit = True
            sync_status['details'].append(f"  ✗ {table_name}: Erreur lors du réalignement: {str(e)[:200]}")
            import traceback
            traceback.print_exc()
            return False
            
    except Exception as e:
        sync_status['details'].append(f"  ✗ {table_name}: Erreur lors de la vérification: {str(e)[:200]}")
        return False

def sync_databases():
    global sync_status
    sync_status = {'running': True, 'message': 'Démarrage...', 'progress': 0, 'details': [], 'code_version': PROJET21_CODE_VERSION}
    sync_status['details'].append(
        f"🧩 Version code Projet21: {PROJET21_CODE_VERSION} ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})"
    )
    
    try:
        source_conn = get_connection(SOURCE_CONFIG, readonly=True)
        source_cursor = source_conn.cursor()
        target_conn = get_connection(TARGET_CONFIG)
        target_cursor = target_conn.cursor()
        
        source_cursor.execute("""
            SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_TYPE = 'BASE TABLE'
        """)
        source_tables = [row.TABLE_NAME for row in source_cursor.fetchall()]
        
        target_cursor.execute("""
            SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_TYPE = 'BASE TABLE'
        """)
        target_tables = [row.TABLE_NAME for row in target_cursor.fetchall()]
        
        # Construire le graphe de dépendances et trier topologiquement
        sync_status['details'].append("🔍 Analyse des dépendances FK...")
        dependency_graph, referenced_tables = build_dependency_graph(source_cursor, target_cursor, source_tables)
        sorted_tables = topological_sort(source_tables, dependency_graph)
        
        # Log des dépendances pour diagnostic
        deps_count = sum(len(deps) for deps in dependency_graph.values())
        sync_status['details'].append(f"✓ Ordre de synchronisation déterminé ({len(sorted_tables)} tables, {deps_count} dépendances FK)")
        
        # Afficher quelques exemples d'ordre pour vérification
        if len(sorted_tables) > 0:
            sample_size = min(10, len(sorted_tables))
            sample_order = sorted_tables[:sample_size]
            sync_status['details'].append(f"  Exemples d'ordre: {', '.join(sample_order)}...")
        
        total_tables = len(sorted_tables)
        failed_tables = []
        
        # ÉTAPE PRÉLIMINAIRE : Réalignement automatique des IDs pour PAPIERS_ARTICLES et PAPIERS_IMPRIMEURS
        # Cette étape garantit que les IDs sont alignés avant la synchronisation normale
        sync_status['details'].append("\n🔄 Étape de réalignement automatique des IDs...")
        
        tables_to_realign = ['PAPIERS_ARTICLES', 'PAPIERS_IMPRIMEURS']
        for table_name in tables_to_realign:
            if table_name in sorted_tables:
                auto_realign_table_ids(source_cursor, target_cursor, target_conn, table_name, sync_status)
        
        # Première passe : synchroniser toutes les tables dans l'ordre topologique
        for i, table_name in enumerate(sorted_tables):
            sync_status['message'] = f"Traitement (Passe 1): {table_name}"
            sync_status['progress'] = int((i / total_tables) * 50)  # 50% pour la première passe
            
            try:
                if table_name not in target_tables:
                    try:
                        create_sql = get_table_definition(source_cursor, table_name)
                        target_cursor.execute(create_sql)
                        target_conn.commit()
                        sync_status['details'].append(f"✓ Table créée: {table_name}")
                    except Exception as create_err:
                        err_str = str(create_err)
                        if 'F_COMPTET' in err_str or 'regNOVA' in err_str:
                            sync_status['details'].append(f"✗ {table_name}: Impossible de créer la table (dépendance vers regNOVA.dbo.F_COMPTET: {err_str[:150]})")
                            continue
                        else:
                            raise
                
                pk_columns = get_primary_keys(source_cursor, table_name)
                
                if not pk_columns:
                    cols = get_table_columns(source_cursor, table_name)
                    pk_columns = [c[0] for c in cols]  # Utiliser toutes les colonnes comme clé
                    sync_status['details'].append(f"⚠ {table_name}: Pas de PK, utilisation de toutes les colonnes")
                
                # Récupérer les colonnes de la table source
                try:
                    source_cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
                    columns = [desc[0] for desc in source_cursor.description]
                    col_list = ", ".join([f"[{c}]" for c in columns])
                except Exception as col_err:
                    err_str = str(col_err)
                    if 'F_COMPTET' in err_str or 'regNOVA' in err_str:
                        # Table FACTURES a une dépendance vers une table externe inexistante (regNOVA.dbo.F_COMPTET)
                        sync_status['details'].append(f"✗ {table_name}: Table ignorée (dépendance vers regNOVA.dbo.F_COMPTET inexistante)")
                        continue
                    else:
                        raise
                
                # Récupérer les données source
                try:
                    source_cursor.execute(f"SELECT {col_list} FROM [{table_name}]")
                    source_rows = source_cursor.fetchall()
                    source_count = len(source_rows)
                except Exception as select_err:
                    err_str = str(select_err)
                    if 'F_COMPTET' in err_str or 'regNOVA' in err_str:
                        sync_status['details'].append(f"✗ {table_name}: Impossible de lire les données (dépendance vers regNOVA.dbo.F_COMPTET: {err_str[:150]})")
                        continue
                    else:
                        raise
                
                # Récupérer les PKs existants dans la cible
                target_cursor.execute(f"SELECT {', '.join([f'[{pk}]' for pk in pk_columns])} FROM [{table_name}]")
                existing_pks = set()
                for row in target_cursor.fetchall():
                    if len(pk_columns) == 1:
                        existing_pks.add(row[0])
                    else:
                        existing_pks.add(tuple(row))
                
                # Colonnes réellement présentes en cible (pour UPDATE sans toucher aux colonnes purement cibles)
                target_col_names = {c[0] for c in get_table_columns(target_cursor, table_name)}

                # OPTIMISATION : Charger toutes les lignes existantes en mémoire (pour UPDATE rapide)
                # Utiliser les noms de colonnes de la CIBLE (ordre et colonnes supplémentaires ≠ source)
                existing_rows_by_pk = {}
                if existing_pks:
                    target_cursor.execute(f"SELECT * FROM [{table_name}]")
                    tgt_desc = target_cursor.description or []
                    tgt_columns = [d[0] for d in tgt_desc]
                    for row in target_cursor.fetchall():
                        row_dict = {tgt_columns[i]: row[i] for i in range(len(tgt_columns))}
                        if len(pk_columns) == 1:
                            existing_rows_by_pk[row_dict[pk_columns[0]]] = row_dict
                        else:
                            pk_tuple = tuple(row_dict[pk] for pk in pk_columns)
                            existing_rows_by_pk[pk_tuple] = row_dict

                # Trouver les index des colonnes PK dans la liste des colonnes
                pk_indices = [columns.index(pk) for pk in pk_columns]
                
                # Vérifier si la table a une colonne IDENTITY
                identity_enabled = False
                if has_identity_column(target_cursor, table_name):
                    target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] ON")
                    identity_enabled = True
                
                inserted = 0
                updated = 0
                batch = []
                update_batch = []  # Batch pour les UPDATEs
                
                # Récupérer les index uniques (hors PK) pour détecter les conflits
                unique_indexes = _get_unique_indexes(target_cursor, table_name)
                
                # OPTIMISATION : Précharger toutes les combinaisons d'index unique existantes
                existing_unique_indexes = {}
                for idx_name, idx_cols in unique_indexes:
                    if idx_cols:
                        idx_col_list = ", ".join([f"[{c}]" for c in idx_cols])
                        target_cursor.execute(f"SELECT {idx_col_list} FROM [{table_name}] WHERE {' AND '.join([f'[{c}] IS NOT NULL' for c in idx_cols])}")
                        existing_unique_indexes[idx_name] = set()
                        for row in target_cursor.fetchall():
                            if len(idx_cols) == 1:
                                existing_unique_indexes[idx_name].add(row[0])
                            else:
                                existing_unique_indexes[idx_name].add(tuple(row))
                
                # Récupérer les contraintes FK pour vérifier avant insertion
                fk_constraints = get_foreign_keys(target_cursor, table_name)
                
                # OPTIMISATION : Précharger toutes les valeurs FK des tables référencées
                fk_values_cache = {}
                for fk_info in fk_constraints:
                    fk_name, parent_table, parent_col, ref_table, ref_col = fk_info
                    if ref_table not in fk_values_cache:
                        try:
                            # Charger toutes les valeurs FK (y compris 0)
                            # Utiliser IS NOT NULL pour exclure uniquement les NULL réels
                            target_cursor.execute(f"SELECT DISTINCT [{ref_col}] FROM [{ref_table}] WHERE [{ref_col}] IS NOT NULL")
                            fk_values_cache[ref_table] = {row[0] for row in target_cursor.fetchall()}
                            
                            # CORRECTION CRITIQUE : Vérifier explicitement si 0 existe dans la table référencée
                            # et l'ajouter au cache si nécessaire (car 0 IS NOT NULL est vrai mais peut être exclu)
                            # Cette vérification garantit que ID_IMPRIMEUR = 0 est toujours traité comme valide
                            # si ID_SOCIETE = 0 existe dans IMPRIMEURS
                            try:
                                target_cursor.execute(f"SELECT COUNT(*) FROM [{ref_table}] WHERE [{ref_col}] = 0")
                                if target_cursor.fetchone()[0] > 0:
                                    fk_values_cache[ref_table].add(0)
                                    sync_status['details'].append(f"    ✓ Valeur FK 0 détectée et ajoutée au cache pour {ref_table}.{ref_col}")
                            except Exception as e:
                                # Si la vérification échoue, continuer avec le cache existant
                                sync_status['details'].append(f"    ⚠ Impossible de vérifier FK 0 pour {ref_table}.{ref_col}: {str(e)[:100]}")
                                pass
                        except Exception:
                            fk_values_cache[ref_table] = set()
                
                # Colonnes à exclure de la comparaison/UPDATE (ID et colonnes système)
                exclude_cols = set(pk_columns)  # Exclure la PK (ID)
                
                try:
                    for row in source_rows:
                        row_dict = {columns[i]: row[i] for i in range(len(columns))}
                        
                        # Extraire la valeur de la clé primaire
                        if len(pk_columns) == 1:
                            pk_value = row[pk_indices[0]]
                            pk_exists = pk_value in existing_pks
                        else:
                            pk_tuple = tuple(row[idx] for idx in pk_indices)
                            pk_exists = pk_tuple in existing_pks
                        
                        # Si PK existe déjà, vérifier si UPDATE nécessaire
                        if pk_exists:
                            # OPTIMISATION : Utiliser le cache en mémoire au lieu de requête SQL
                            pk_key = pk_value if len(pk_columns) == 1 else pk_tuple
                            target_row_dict = existing_rows_by_pk.get(pk_key)
                            
                            if target_row_dict:
                                has_diff, diff_cols, update_values = compare_rows_for_update(row_dict, target_row_dict, exclude_cols)
                                if has_diff:
                                    diff_cols = [c for c in diff_cols if c in target_col_names]
                                    update_values = {c: update_values[c] for c in diff_cols if c in update_values}
                                    if diff_cols:
                                        # Ajouter au batch UPDATE au lieu d'exécuter immédiatement
                                        if len(pk_columns) == 1:
                                            update_batch.append((diff_cols, update_values, pk_value, None))
                                        else:
                                            update_batch.append((diff_cols, update_values, None, pk_tuple))
                            continue  # PK existe, pas besoin d'INSERT
                        
                        # Vérifier les conflits d'index unique AVANT insertion (OPTIMISÉ avec cache)
                        unique_conflict = False
                        for idx_name, idx_cols in unique_indexes:
                            idx_values = []
                            has_all_cols = True
                            has_null = False
                            for col in idx_cols:
                                if col not in row_dict:
                                    has_all_cols = False
                                    break
                                val = row_dict.get(col)
                                if val is None:
                                    has_null = True
                                idx_values.append(val)
                            
                            if not has_all_cols or has_null:
                                continue  # Index unique avec NULL autorise plusieurs NULL
                            
                            # OPTIMISATION : Vérifier dans le cache préchargé au lieu de requête SQL
                            idx_key = idx_values[0] if len(idx_cols) == 1 else tuple(idx_values)
                            if idx_key in existing_unique_indexes.get(idx_name, set()):
                                # Conflit détecté, récupérer la ligne existante (une seule requête)
                                existing_row = find_row_by_unique_index(target_cursor, table_name, idx_cols, idx_values)
                                if existing_row:
                                    unique_conflict = True
                                    
                                    # Comparer les colonnes (hors PK et colonnes de l'index unique)
                                    exclude_for_update = exclude_cols | set(idx_cols)
                                    has_diff, diff_cols, update_values = compare_rows_for_update(row_dict, existing_row, exclude_for_update)
                                    if has_diff:
                                        diff_cols = [c for c in diff_cols if c in target_col_names]
                                        update_values = {c: update_values[c] for c in diff_cols if c in update_values}
                                        if diff_cols:
                                            # Ajouter au batch UPDATE
                                            update_batch.append((diff_cols, update_values, None, None, idx_cols, idx_values))
                                    break
                        
                        if unique_conflict:
                            continue  # Déjà traité par UPDATE
                        
                        # Vérifier les FK avant insertion (OPTIMISÉ avec cache préchargé)
                        skip_row = False
                        for fk_info in fk_constraints:
                            fk_name, parent_table, parent_col, ref_table, ref_col = fk_info
                            try:
                                fk_col_idx = columns.index(parent_col)
                                fk_value = row[fk_col_idx]
                                if fk_value is not None:  # Ignorer les FK NULL
                                    # OPTIMISATION : Utiliser le cache préchargé au lieu de requête SQL
                                    if ref_table in fk_values_cache:
                                        if fk_value not in fk_values_cache[ref_table]:
                                            # CORRECTION : Pour la valeur 0, vérifier directement dans la table
                                            # car 0 peut être valide même s'il n'est pas dans le cache
                                            if fk_value == 0:
                                                try:
                                                    target_cursor.execute(f"SELECT COUNT(*) FROM [{ref_table}] WHERE [{ref_col}] = 0")
                                                    if target_cursor.fetchone()[0] > 0:
                                                        # 0 existe dans la table, l'ajouter au cache et continuer
                                                        fk_values_cache[ref_table].add(0)
                                                        continue  # FK valide, passer à la suivante
                                                except:
                                                    pass  # Si la vérification échoue, traiter comme FK manquante
                                            
                                            # FK n'existe pas encore, ignorer cette ligne (sera réessayée en passe suivante)
                                            skip_row = True
                                            break
                            except (ValueError, IndexError):
                                # Colonne FK non trouvée, continuer
                                pass
                        
                        if skip_row:
                            continue  # FK manquante, sera réessayée en passe suivante
                        
                        # Pas de conflit et FK OK : ajouter au batch pour INSERT
                        if len(pk_columns) == 1:
                            batch.append(list(row))
                            existing_pks.add(pk_value)
                        else:
                            batch.append(list(row))
                            existing_pks.add(pk_tuple)
                        
                        # Insérer par batch de 100
                        if len(batch) >= 100:
                            placeholders = ", ".join(["?" for _ in columns])
                            target_cursor.executemany(
                                f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                batch
                            )
                            inserted += len(batch)
                            batch = []
                    
                    # Insérer le reste
                    if batch:
                        placeholders = ", ".join(["?" for _ in columns])
                        target_cursor.executemany(
                            f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                            batch
                        )
                        inserted += len(batch)
                    
                    # OPTIMISATION : Exécuter les UPDATEs en batch
                    if update_batch:
                        for update_item in update_batch:
                            # Format: (diff_cols, update_values, pk_value, pk_tuple) ou (diff_cols, update_values, None, None, idx_cols, idx_values)
                            if len(update_item) == 4:
                                diff_cols, update_values, pk_value, pk_tuple = update_item
                                idx_cols, idx_values = None, None
                            else:
                                diff_cols, update_values, _, _, idx_cols, idx_values = update_item
                                pk_value, pk_tuple = None, None
                            
                            if pk_value is not None or pk_tuple is not None:
                                # UPDATE par PK
                                if pk_value is not None:
                                    where_pk = f"[{pk_columns[0]}] = ?"
                                    set_clauses = [f"[{col}] = ?" for col in diff_cols]
                                    set_values = [update_values[col] for col in diff_cols]
                                    update_sql = f"UPDATE [{table_name}] SET {', '.join(set_clauses)} WHERE {where_pk}"
                                    target_cursor.execute(update_sql, tuple(set_values) + (pk_value,))
                                else:
                                    where_parts = [f"[{pk}] = ?" for pk in pk_columns]
                                    where_pk = " AND ".join(where_parts)
                                    set_clauses = [f"[{col}] = ?" for col in diff_cols]
                                    set_values = [update_values[col] for col in diff_cols]
                                    update_sql = f"UPDATE [{table_name}] SET {', '.join(set_clauses)} WHERE {where_pk}"
                                    target_cursor.execute(update_sql, tuple(set_values) + pk_tuple)
                            elif idx_cols and idx_values:
                                # UPDATE par index unique
                                where_unique = " AND ".join([f"[{col}] = ?" for col in idx_cols])
                                set_clauses = [f"[{col}] = ?" for col in diff_cols]
                                set_values = [update_values[col] for col in diff_cols]
                                update_sql = f"UPDATE [{table_name}] SET {', '.join(set_clauses)} WHERE {where_unique}"
                                target_cursor.execute(update_sql, tuple(set_values) + tuple(idx_values))
                            updated += 1
                    
                    if inserted > 0 or updated > 0:
                        target_conn.commit()
                        msg_parts = []
                        if inserted > 0:
                            msg_parts.append(f"{inserted} ajoutés")
                        if updated > 0:
                            msg_parts.append(f"{updated} mis à jour")
                        sync_status['details'].append(f"✓ {table_name}: {', '.join(msg_parts)}")
                    else:
                        sync_status['details'].append(f"○ {table_name}: à jour")
                finally:
                    # Désactiver IDENTITY_INSERT si on l'a activé
                    if identity_enabled:
                        target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] OFF")
                    
            except Exception as e:
                target_conn.rollback()
                error_msg = str(e)
                # Ignorer les erreurs de doublons (données déjà présentes)
                if 'duplicate key' in error_msg.lower() or 'unique index' in error_msg.lower():
                    sync_status['details'].append(f"○ {table_name}: doublons ignorés (données déjà présentes)")
                # Pour les erreurs de types non supportés, on réessaiera en passe finale ligne par ligne
                elif 'HYC00' in error_msg or 'Fonctionnalité optionnelle non implémentée' in error_msg:
                    sync_status['details'].append(f"⚠ {table_name}: types de données non supportés (tentative en passe finale)")
                    failed_tables.append(table_name)
                else:
                    sync_status['details'].append(f"✗ {table_name}: {error_msg}")
                    # Si l'erreur est une FK constraint, on réessaiera en deuxième passe
                    if 'FOREIGN KEY constraint' in error_msg or 'FOREIGN KEY SAME TABLE' in error_msg:
                        failed_tables.append(table_name)
        
        # Passes supplémentaires : réessayer les tables qui ont échoué à cause de FK
        pass_num = 2
        max_passes = 10  # Augmenter le nombre de passes pour résoudre les dépendances complexes
        while failed_tables and pass_num <= max_passes:
            sync_status['details'].append(f"\n🔄 Passe {pass_num} pour {len(failed_tables)} tables avec dépendances...")
            still_failed = []
            
            for i, table_name in enumerate(failed_tables):
                progress_base = 50 + ((pass_num - 2) * 10)
                sync_status['message'] = f"Traitement (Passe {pass_num}): {table_name}"
                sync_status['progress'] = progress_base + int((i / len(failed_tables)) * 10)
                
                try:
                    pk_columns = get_primary_keys(source_cursor, table_name)
                    
                    if not pk_columns:
                        cols = get_table_columns(source_cursor, table_name)
                        pk_columns = [c[0] for c in cols]
                    
                    source_cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
                    columns = [desc[0] for desc in source_cursor.description]
                    col_list = ", ".join([f"[{c}]" for c in columns])
                    
                    source_cursor.execute(f"SELECT {col_list} FROM [{table_name}]")
                    source_rows = source_cursor.fetchall()
                    
                    target_cursor.execute(f"SELECT {', '.join([f'[{pk}]' for pk in pk_columns])} FROM [{table_name}]")
                    existing_pks = set()
                    for row in target_cursor.fetchall():
                        if len(pk_columns) == 1:
                            existing_pks.add(row[0])
                        else:
                            existing_pks.add(tuple(row))
                    
                    pk_indices = [columns.index(pk) for pk in pk_columns]
                    
                    identity_enabled = False
                    if has_identity_column(target_cursor, table_name):
                        target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] ON")
                        identity_enabled = True
                    
                    inserted = 0
                    batch = []
                    try:
                        for row in source_rows:
                            if len(pk_columns) == 1:
                                pk_value = row[pk_indices[0]]
                                if pk_value not in existing_pks:
                                    batch.append(list(row))
                                    existing_pks.add(pk_value)
                            else:
                                pk_tuple = tuple(row[idx] for idx in pk_indices)
                                if pk_tuple not in existing_pks:
                                    batch.append(list(row))
                                    existing_pks.add(pk_tuple)
                            
                            if len(batch) >= 100:
                                placeholders = ", ".join(["?" for _ in columns])
                                target_cursor.executemany(
                                    f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                    batch
                                )
                                inserted += len(batch)
                                batch = []
                        
                        if batch:
                            placeholders = ", ".join(["?" for _ in columns])
                            target_cursor.executemany(
                                f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                batch
                            )
                            inserted += len(batch)
                        
                        if inserted > 0:
                            target_conn.commit()
                            sync_status['details'].append(f"✓ {table_name} (Passe {pass_num}): {inserted} enregistrements ajoutés")
                        else:
                            sync_status['details'].append(f"○ {table_name} (Passe {pass_num}): à jour")
                    finally:
                        if identity_enabled:
                            target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] OFF")
                            
                except Exception as e:
                    target_conn.rollback()
                    error_msg = str(e)
                    # Ignorer les erreurs de doublons (données déjà présentes)
                    if 'duplicate key' in error_msg.lower() or 'unique index' in error_msg.lower():
                        sync_status['details'].append(f"○ {table_name} (Passe {pass_num}): doublons ignorés")
                    # Ignorer les erreurs de types non supportés
                    elif 'HYC00' in error_msg or 'Fonctionnalité optionnelle non implémentée' in error_msg:
                        sync_status['details'].append(f"⚠ {table_name} (Passe {pass_num}): types non supportés")
                    else:
                        sync_status['details'].append(f"✗ {table_name} (Passe {pass_num}): {error_msg}")
                        # Si l'erreur est toujours une FK constraint, on réessaiera au prochain tour
                        if 'FOREIGN KEY constraint' in error_msg or 'FOREIGN KEY SAME TABLE' in error_msg:
                            still_failed.append(table_name)
            
            failed_tables = still_failed
            pass_num += 1
        
        # Passe spéciale pour les tables avec types non supportés : utiliser INSERT ... SELECT
        tables_with_unsupported_types = []
        for table_name in failed_tables[:]:  # Copie de la liste pour pouvoir la modifier
            # Vérifier si cette table a eu des erreurs HYC00
            for detail in sync_status['details']:
                if table_name in detail and ('HYC00' in detail or 'types non supportés' in detail or 'Fonctionnalité optionnelle non implémentée' in detail):
                    tables_with_unsupported_types.append(table_name)
                    failed_tables.remove(table_name)
                    break
        
        if tables_with_unsupported_types:
            sync_status['details'].append(f"\n🔄 Passe spéciale (INSERT ... SELECT) pour {len(tables_with_unsupported_types)} tables avec types non supportés...")
            for i, table_name in enumerate(tables_with_unsupported_types):
                sync_status['message'] = f"Traitement spécial (SQL direct): {table_name}"
                sync_status['progress'] = 85 + int((i / len(tables_with_unsupported_types)) * 5)
                
                try:
                    # Récupérer les colonnes et PK
                    pk_columns = get_primary_keys(source_cursor, table_name)
                    if not pk_columns:
                        cols = get_table_columns(source_cursor, table_name)
                        pk_columns = [c[0] for c in cols]
                    
                    # Récupérer les colonnes de la table
                    cols = get_table_columns(source_cursor, table_name)
                    col_names = [c[0] for c in cols]
                    col_list = ", ".join([f"[{c}]" for c in col_names])
                    
                    # Vérifier les PKs existants dans la cible
                    target_cursor.execute(f"SELECT {', '.join([f'[{pk}]' for pk in pk_columns])} FROM [{table_name}]")
                    existing_pks = set()
                    for row in target_cursor.fetchall():
                        if len(pk_columns) == 1:
                            existing_pks.add(row[0])
                        else:
                            existing_pks.add(tuple(row))
                    
                    # Construire la clause WHERE pour exclure les PKs existants
                    if existing_pks:
                        if len(pk_columns) == 1:
                            pk_list = ','.join([f"'{str(pk)}'" if isinstance(pk, str) else str(pk) for pk in existing_pks if pk is not None])
                            where_clause = f"WHERE [{pk_columns[0]}] NOT IN ({pk_list})" if pk_list else ""
                        else:
                            # Pour les PKs composites, utiliser NOT EXISTS
                            pk_conditions = []
                            for pk_tuple in existing_pks:
                                if None not in pk_tuple:
                                    conditions = []
                                    for idx, pk_col in enumerate(pk_columns):
                                        val = pk_tuple[idx]
                                        if isinstance(val, str):
                                            conditions.append(f"[{pk_col}] = '{val.replace("'", "''")}'")
                                        else:
                                            conditions.append(f"[{pk_col}] = {val}")
                                    pk_conditions.append("(" + " AND ".join(conditions) + ")")
                            if pk_conditions:
                                where_clause = "WHERE NOT (" + " OR ".join(pk_conditions) + ")"
                            else:
                                where_clause = ""
                    else:
                        where_clause = ""
                    
                    # Utiliser INSERT ... SELECT avec conversion de types via OPENROWSET
                    identity_enabled = False
                    if has_identity_column(target_cursor, table_name):
                        target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] ON")
                        identity_enabled = True
                    
                    try:
                        # Construire les colonnes avec conversion de types pour éviter HYC00
                        # Convertir les types problématiques en VARCHAR(MAX) puis reconvertir
                        converted_cols = []
                        for col_name, data_type, max_len, nullable, default in cols:
                            # Pour les types problématiques, utiliser CAST/CONVERT
                            if data_type in ('timestamp', 'rowversion', 'sql_variant', 'geography', 'geometry', 'hierarchyid'):
                                converted_cols.append(f"CAST([{col_name}] AS VARBINARY(MAX)) AS [{col_name}]")
                            elif data_type in ('xml', 'text', 'ntext'):
                                converted_cols.append(f"CAST([{col_name}] AS NVARCHAR(MAX)) AS [{col_name}]")
                            elif data_type in ('image'):
                                converted_cols.append(f"CAST([{col_name}] AS VARBINARY(MAX)) AS [{col_name}]")
                            else:
                                converted_cols.append(f"[{col_name}]")
                        
                        converted_col_list = ", ".join(converted_cols)
                        
                        # Utiliser une approche avec conversion SQL explicite via la connexion source
                        # Construire une requête SELECT avec conversion de types pour chaque colonne problématique
                        select_cols = []
                        for col_name, data_type, max_len, nullable, default in cols:
                            # Pour les types problématiques, utiliser CONVERT explicite
                            if data_type in ('timestamp', 'rowversion'):
                                select_cols.append(f"CONVERT(VARBINARY(8), [{col_name}]) AS [{col_name}]")
                            elif data_type in ('sql_variant'):
                                select_cols.append(f"CONVERT(NVARCHAR(MAX), [{col_name}]) AS [{col_name}]")
                            elif data_type in ('geography', 'geometry'):
                                select_cols.append(f"CONVERT(VARBINARY(MAX), [{col_name}]) AS [{col_name}]")
                            elif data_type in ('hierarchyid'):
                                select_cols.append(f"CONVERT(NVARCHAR(4000), [{col_name}]) AS [{col_name}]")
                            elif data_type in ('xml'):
                                select_cols.append(f"CONVERT(NVARCHAR(MAX), CAST([{col_name}] AS NVARCHAR(MAX))) AS [{col_name}]")
                            elif data_type in ('text', 'ntext'):
                                select_cols.append(f"CONVERT(NVARCHAR(MAX), [{col_name}]) AS [{col_name}]")
                            elif data_type in ('image'):
                                select_cols.append(f"CONVERT(VARBINARY(MAX), [{col_name}]) AS [{col_name}]")
                            else:
                                select_cols.append(f"[{col_name}]")
                        
                        select_col_list = ", ".join(select_cols)
                        
                        # Lire depuis source avec conversion SQL côté serveur
                        try:
                            # Essayer d'abord avec conversion SQL
                            try:
                                source_cursor.execute(f"SELECT {select_col_list} FROM [{table_name}] {where_clause}")
                                rows = source_cursor.fetchall()
                            except Exception as convert_err:
                                # Si la conversion SQL échoue, essayer sans conversion mais avec gestion d'erreurs
                                source_cursor.execute(f"SELECT {col_list} FROM [{table_name}] {where_clause}")
                                rows = source_cursor.fetchall()
                            
                            inserted = 0
                            placeholders = ", ".join(["?" for _ in col_names])
                            
                            for row in rows:
                                try:
                                    # Convertir les valeurs None et les types problématiques
                                    row_values = []
                                    for val in row:
                                        if val is None:
                                            row_values.append(None)
                                        elif isinstance(val, bytes):
                                            row_values.append(val)
                                        elif isinstance(val, str):
                                            row_values.append(val)
                                        else:
                                            row_values.append(val)
                                    
                                    target_cursor.execute(
                                        f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                        row_values
                                    )
                                    inserted += 1
                                except Exception as row_err:
                                    # Ignorer les erreurs de ligne individuelle (doublons, etc.)
                                    continue
                            
                            if inserted > 0:
                                target_conn.commit()
                                sync_status['details'].append(f"✓ {table_name} (SQL direct): {inserted} enregistrements ajoutés")
                            else:
                                sync_status['details'].append(f"○ {table_name} (SQL direct): à jour")
                        except Exception as select_err:
                            # Si la conversion SQL ne fonctionne pas, essayer sans conversion
                            try:
                                source_cursor.execute(f"SELECT {col_list} FROM [{table_name}] {where_clause}")
                                rows = source_cursor.fetchall()
                                
                                inserted = 0
                                placeholders = ", ".join(["?" for _ in col_names])
                                
                                for row in rows:
                                    try:
                                        target_cursor.execute(
                                            f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                            list(row)
                                        )
                                        inserted += 1
                                    except Exception:
                                        continue
                                
                                if inserted > 0:
                                    target_conn.commit()
                                    sync_status['details'].append(f"✓ {table_name} (SQL direct): {inserted} enregistrements ajoutés")
                                else:
                                    sync_status['details'].append(f"⚠ {table_name} (SQL direct): Aucun enregistrement inséré")
                            except Exception as fallback_err:
                                sync_status['details'].append(f"✗ {table_name} (SQL direct): {str(fallback_err)}")
                    finally:
                        if identity_enabled:
                            target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] OFF")
                            
                except Exception as e:
                    sync_status['details'].append(f"✗ {table_name} (SQL direct): {str(e)}")
        
        # Passe finale : insertion ligne par ligne avec itération jusqu'à convergence
        if failed_tables:
            sync_status['details'].append(f"\n🔄 Passe finale (ligne par ligne) pour {len(failed_tables)} tables problématiques...")
            for i, table_name in enumerate(failed_tables):
                sync_status['message'] = f"Traitement final (ligne par ligne): {table_name}"
                sync_status['progress'] = 90 + int((i / len(failed_tables)) * 10)
                
                try:
                    pk_columns = get_primary_keys(source_cursor, table_name)
                    if not pk_columns:
                        cols = get_table_columns(source_cursor, table_name)
                        pk_columns = [c[0] for c in cols]
                    
                    source_cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
                    columns = [desc[0] for desc in source_cursor.description]
                    col_list = ", ".join([f"[{c}]" for c in columns])
                    
                    source_cursor.execute(f"SELECT {col_list} FROM [{table_name}]")
                    source_rows = source_cursor.fetchall()
                    
                    # Récupérer les contraintes FK pour cette table
                    fk_constraints = get_foreign_keys(target_cursor, table_name)
                    fk_cache = {}  # Cache pour les valeurs FK vérifiées
                    
                    identity_enabled = False
                    if has_identity_column(target_cursor, table_name):
                        target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] ON")
                        identity_enabled = True
                    
                    placeholders = ", ".join(["?" for _ in columns])
                    total_inserted = 0
                    iteration = 0
                    max_iterations = 20  # Augmenté pour gérer les dépendances complexes
                    consecutive_no_insert = 0
                    max_consecutive_no_insert = 3  # Arrêter après 3 itérations sans insertion
                    
                    try:
                        while iteration < max_iterations:
                            # Récupérer les PKs existants à chaque itération
                            target_cursor.execute(f"SELECT {', '.join([f'[{pk}]' for pk in pk_columns])} FROM [{table_name}]")
                            existing_pks = set()
                            for row in target_cursor.fetchall():
                                if len(pk_columns) == 1:
                                    existing_pks.add(row[0])
                                else:
                                    existing_pks.add(tuple(row))
                            
                            # Vider le cache FK pour refléter les nouvelles insertions dans les tables référencées
                            fk_cache.clear()
                            
                            pk_indices = [columns.index(pk) for pk in pk_columns]
                            inserted_this_iteration = 0
                            failed_rows = 0
                            fk_errors = 0
                            other_errors = 0
                            
                            for row in source_rows:
                                # Vérifier si déjà présent
                                if len(pk_columns) == 1:
                                    pk_value = row[pk_indices[0]]
                                    if pk_value in existing_pks:
                                        continue
                                else:
                                    pk_tuple = tuple(row[idx] for idx in pk_indices)
                                    if pk_tuple in existing_pks:
                                        continue
                                
                                # Vérifier les FK avant insertion
                                skip_row = False
                                for fk_info in fk_constraints:
                                    fk_name, parent_table, parent_col, ref_table, ref_col = fk_info
                                    try:
                                        fk_col_idx = columns.index(parent_col)
                                        fk_value = row[fk_col_idx]
                                        if fk_value is not None:  # Ignorer les FK NULL
                                            # Utiliser le cache si disponible
                                            cache_key = (ref_table, ref_col, fk_value)
                                            if cache_key not in fk_cache:
                                                fk_cache[cache_key] = check_fk_exists(target_cursor, ref_table, ref_col, fk_value)
                                            if not fk_cache[cache_key]:
                                                # CORRECTION : Pour la valeur 0, vérifier directement dans la table
                                                # car 0 peut être valide même s'il n'est pas dans le cache
                                                if fk_value == 0:
                                                    if check_fk_exists(target_cursor, ref_table, ref_col, 0):
                                                        fk_cache[cache_key] = True
                                                    else:
                                                        # FK n'existe pas, ignorer cet enregistrement
                                                        skip_row = True
                                                        fk_errors += 1
                                                        break
                                                else:
                                                    # FK n'existe pas, ignorer cet enregistrement
                                                    skip_row = True
                                                    fk_errors += 1
                                                    break
                                    except (ValueError, IndexError):
                                        # Colonne FK non trouvée, continuer
                                        pass
                                
                                if skip_row:
                                    failed_rows += 1
                                    continue
                                
                                # Essayer d'insérer
                                try:
                                    # Convertir les valeurs None en None (pas de conversion)
                                    row_values = []
                                    for val in row:
                                        if val is None:
                                            row_values.append(None)
                                        else:
                                            # Essayer de convertir les types problématiques
                                            try:
                                                if isinstance(val, bytes):
                                                    # Pour les types binaires, essayer de les convertir
                                                    row_values.append(val)
                                                elif isinstance(val, str) and len(val) > 0:
                                                    row_values.append(val)
                                                else:
                                                    row_values.append(val)
                                            except:
                                                row_values.append(val)
                                    
                                    target_cursor.execute(
                                        f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                        row_values
                                    )
                                    target_conn.commit()
                                    inserted_this_iteration += 1
                                    if len(pk_columns) == 1:
                                        existing_pks.add(row[pk_indices[0]])
                                        # Mettre à jour le cache FK pour les auto-références
                                        for fk_info in fk_constraints:
                                            fk_name, parent_table, parent_col, ref_table, ref_col = fk_info
                                            if ref_table == table_name and ref_col in pk_columns:
                                                cache_key = (ref_table, ref_col, row[pk_indices[0]])
                                                fk_cache[cache_key] = True
                                    else:
                                        existing_pks.add(tuple(row[idx] for idx in pk_indices))
                                        # Mettre à jour le cache FK pour les auto-références
                                        for fk_info in fk_constraints:
                                            fk_name, parent_table, parent_col, ref_table, ref_col = fk_info
                                            if ref_table == table_name and ref_col in pk_columns:
                                                pk_val = tuple(row[columns.index(pk)] for pk in pk_columns if pk == ref_col)
                                                if len(pk_val) == 1:
                                                    cache_key = (ref_table, ref_col, pk_val[0])
                                                    fk_cache[cache_key] = True
                                except Exception as row_err:
                                    failed_rows += 1
                                    error_str = str(row_err)
                                    if 'FOREIGN KEY' in error_str or '547' in error_str:
                                        fk_errors += 1
                                        # Extraire les infos FK pour diagnostic
                                        ref_table, ref_column = extract_fk_info(error_str)
                                        if ref_table and ref_column:
                                            # Trouver la colonne FK dans la ligne actuelle
                                            try:
                                                fk_col_idx = columns.index(ref_column)
                                                fk_value = row[fk_col_idx]
                                                # Vérifier si la valeur existe dans la table référencée
                                                if not check_fk_exists(target_cursor, ref_table, ref_column, fk_value):
                                                    # La valeur FK n'existe pas - enregistrement orphelin, on peut l'ignorer
                                                    # Ne pas compter comme erreur si la FK n'existe vraiment pas
                                                    fk_errors -= 1  # Corriger le compteur
                                                    failed_rows -= 1  # Ne pas compter comme échec
                                            except:
                                                pass
                                    elif 'HYC00' in error_str or 'Fonctionnalité optionnelle non implémentée' in error_str:
                                        # Pour les types non supportés, on ne peut pas les insérer
                                        other_errors += 1
                                    else:
                                        other_errors += 1
                                    continue
                            
                            total_inserted += inserted_this_iteration
                            
                            # Si aucune insertion cette itération
                            if inserted_this_iteration == 0:
                                consecutive_no_insert += 1
                                if consecutive_no_insert >= max_consecutive_no_insert:
                                    # Arrêter si plusieurs itérations sans insertion
                                    break
                            else:
                                consecutive_no_insert = 0  # Réinitialiser le compteur
                            
                            iteration += 1
                        
                        # Rapport détaillé
                        remaining = len(source_rows) - total_inserted
                        if total_inserted > 0:
                            sync_status['details'].append(f"✓ {table_name} (Final): {total_inserted} enregistrements ajoutés en {iteration} itérations")
                        if remaining > 0:
                            sync_status['details'].append(f"⚠ {table_name} (Final): {remaining} enregistrements restants (FK: {fk_errors}, Autres: {other_errors})")
                    finally:
                        if identity_enabled:
                            target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] OFF")
                            
                except Exception as e:
                    sync_status['details'].append(f"✗ {table_name} (Final): {str(e)}")
        
        # Passe ultime : désactiver temporairement les FK pour insérer tous les enregistrements manquants
        # Vérifier TOUTES les tables communes, pas seulement celles qui ont échoué
        sync_status['details'].append(f"\n🔍 Vérification finale de toutes les tables...")
        
        # Récupérer toutes les tables communes
        source_cursor.execute("""
            SELECT TABLE_NAME 
            FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_TYPE = 'BASE TABLE' 
            AND TABLE_NAME NOT LIKE '#%'
            ORDER BY TABLE_NAME
        """)
        all_source_tables = [row[0] for row in source_cursor.fetchall()]
        
        target_cursor.execute("""
            SELECT TABLE_NAME 
            FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_TYPE = 'BASE TABLE' 
            AND TABLE_NAME NOT LIKE '#%'
            ORDER BY TABLE_NAME
        """)
        all_target_tables = [row[0] for row in target_cursor.fetchall()]
        
        # Tables communes
        common_tables = set(all_source_tables) & set(all_target_tables)
        
        # Identifier les tables qui ont encore des enregistrements manquants
        tables_still_missing = []
        for table_name in sorted(common_tables):
            try:
                pk_columns = get_primary_keys(source_cursor, table_name)
                if not pk_columns:
                    cols = get_table_columns(source_cursor, table_name)
                    pk_columns = [c[0] for c in cols]
                
                # Compter les PKs manquants (plus précis que COUNT)
                source_cursor.execute(f"SELECT {', '.join([f'[{pk}]' for pk in pk_columns])} FROM [{table_name}]")
                source_pks = set()
                for row in source_cursor.fetchall():
                    if len(pk_columns) == 1:
                        source_pks.add(row[0])
                    else:
                        source_pks.add(tuple(row))
                
                target_cursor.execute(f"SELECT {', '.join([f'[{pk}]' for pk in pk_columns])} FROM [{table_name}]")
                target_pks = set()
                for row in target_cursor.fetchall():
                    if len(pk_columns) == 1:
                        target_pks.add(row[0])
                    else:
                        target_pks.add(tuple(row))
                
                missing_pks = source_pks - target_pks
                if missing_pks:
                    tables_still_missing.append((table_name, len(missing_pks)))
            except Exception as e:
                # Ignorer les erreurs de vérification
                pass
        
        # Trier par nombre d'enregistrements manquants (du plus grand au plus petit)
        tables_still_missing.sort(key=lambda x: x[1], reverse=True)
        
        if tables_still_missing:
            total_missing = sum(count for _, count in tables_still_missing)
            sync_status['details'].append(f"\n🔄 Passe ultime (sans contraintes FK) pour {len(tables_still_missing)} tables avec {total_missing} enregistrements manquants...")
            for i, (table_name, missing_count) in enumerate(tables_still_missing):
                sync_status['message'] = f"Traitement ultime (sans FK): {table_name} ({missing_count} manquants)"
                sync_status['progress'] = 95 + int((i / len(tables_still_missing)) * 5)
                
                try:
                    # Récupérer les contraintes FK de cette table
                    fk_constraints = get_foreign_keys(target_cursor, table_name)
                    fk_names = [fk[0] for fk in fk_constraints]  # fk[0] est le nom de la contrainte
                    
                    # Désactiver les contraintes FK
                    disabled_fks = []
                    for fk_name in fk_names:
                        try:
                            target_cursor.execute(f"ALTER TABLE [{table_name}] NOCHECK CONSTRAINT [{fk_name}]")
                            disabled_fks.append(fk_name)
                        except:
                            pass
                    
                    try:
                        # Récupérer les PKs et colonnes
                        pk_columns = get_primary_keys(source_cursor, table_name)
                        if not pk_columns:
                            cols = get_table_columns(source_cursor, table_name)
                            pk_columns = [c[0] for c in cols]
                        
                        source_cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
                        columns = [desc[0] for desc in source_cursor.description]
                        col_list = ", ".join([f"[{c}]" for c in columns])
                        
                        # Récupérer les PKs existants dans la cible
                        target_cursor.execute(f"SELECT {', '.join([f'[{pk}]' for pk in pk_columns])} FROM [{table_name}]")
                        existing_pks = set()
                        for row in target_cursor.fetchall():
                            if len(pk_columns) == 1:
                                existing_pks.add(row[0])
                            else:
                                existing_pks.add(tuple(row))
                        
                        # Récupérer les PKs de la source pour identifier les manquants
                        source_pks = set()
                        missing_pks_list = []
                        try:
                            source_cursor.execute(f"SELECT {', '.join([f'[{pk}]' for pk in pk_columns])} FROM [{table_name}]")
                            for row in source_cursor.fetchall():
                                if len(pk_columns) == 1:
                                    pk_val = row[0]
                                    source_pks.add(pk_val)
                                    if pk_val not in existing_pks:
                                        missing_pks_list.append(pk_val)
                                else:
                                    pk_tuple = tuple(row)
                                    source_pks.add(pk_tuple)
                                    if pk_tuple not in existing_pks:
                                        missing_pks_list.append(pk_tuple)
                        except Exception as pk_err:
                            # Si la lecture des PKs échoue, réessayer avec une approche différente
                            sync_status['details'].append(f"  ⚠ Lecture PKs source échouée, utilisation d'une approche alternative...")
                            # On utilisera OPENROWSET directement avec WHERE basé sur les PKs manquantes identifiées précédemment
                            pass
                        
                        # Utiliser l'approche ligne par ligne directement (évite OPENROWSET qui peut être bloqué)
                        identity_enabled = False
                        if has_identity_column(target_cursor, table_name):
                            target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] ON")
                            identity_enabled = True
                        
                        inserted = 0
                        updated_ultime = 0  # Compteur pour les UPDATEs dans la passe ultime
                        # Diagnostics: conserver quelques erreurs représentatives
                        ultime_diag = {
                            "openrowset": None,
                            "read": None,
                            "convert": None,
                            "insert_samples": [],
                            "openrowset_pk_samples": [],
                        }
                        
                        # Vérifier si la table a des types non supportés
                        cols_info = get_table_columns(source_cursor, table_name)
                        has_unsupported_types = any(
                            dt in ('timestamp', 'rowversion', 'sql_variant', 'xml', 'text', 'ntext', 
                                   'geography', 'geometry', 'hierarchyid', 'image')
                            for _, dt, _, _, _ in cols_info
                        )
                        
                        # Utiliser directement l'approche ligne par ligne (plus fiable, évite OPENROWSET)
                        try:
                            # Index uniques disponibles pour résoudre les conflits (INSERT -> UPDATE)
                            unique_indexes = _get_unique_indexes(target_cursor, table_name)
                            exclude_cols = set(pk_columns)
                            updated_ultime = 0
                            
                            # Si on a la liste des PKs manquantes, construire une clause WHERE
                            if missing_pks_list and len(missing_pks_list) > 0:
                                # Construire WHERE pour filtrer uniquement les PKs manquantes
                                if len(pk_columns) == 1:
                                    pk_values = [pk for pk in missing_pks_list if pk is not None]
                                    if pk_values:
                                        # Limiter à 1000 pour éviter les requêtes trop longues
                                        for batch in [pk_values[i:i+1000] for i in range(0, len(pk_values), 1000)]:
                                            pk_strs = []
                                            for pk_val in batch:
                                                if isinstance(pk_val, str):
                                                    pk_strs.append(f"'{pk_val.replace("'", "''")}'")
                                                else:
                                                    pk_strs.append(str(pk_val))
                                            where_clause = f"WHERE [{pk_columns[0]}] IN ({','.join(pk_strs)})"
                                            source_cursor.execute(f"SELECT {col_list} FROM [{table_name}] {where_clause}")
                                            source_rows = source_cursor.fetchall()
                                            
                                            pk_indices = [columns.index(pk) for pk in pk_columns]
                                            placeholders = ", ".join(["?" for _ in columns])
                                            
                                            for row in source_rows:
                                                row_dict = {columns[i]: row[i] for i in range(len(columns))}
                                                
                                                try:
                                                    target_cursor.execute(
                                                        f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                                        list(row)
                                                    )
                                                    inserted += 1
                                                    if inserted % 100 == 0:
                                                        target_conn.commit()
                                                except Exception as ins_err:
                                                        err_str = str(ins_err)
                                                        
                                                        # Si erreur de conflit d'index unique, essayer UPDATE (Option A)
                                                        if 'duplicate key' in err_str.lower() or 'unique index' in err_str.lower() or '2601' in err_str:
                                                            updated_via_unique = False
                                                            
                                                            for idx_name, idx_cols in unique_indexes:
                                                                if not idx_cols:
                                                                    continue
                                                                
                                                                # Construire les valeurs de l'index (NULL inclus)
                                                                idx_values = []
                                                                has_all = True
                                                                for col in idx_cols:
                                                                    if col not in row_dict:
                                                                        has_all = False
                                                                        break
                                                                    idx_values.append(row_dict.get(col))
                                                                if not has_all:
                                                                    continue
                                                                
                                                                # Utiliser find_row_by_unique_index pour gérer les NULL correctement
                                                                existing_row = find_row_by_unique_index(
                                                                    target_cursor, table_name, idx_cols, idx_values
                                                                )
                                                                if not existing_row:
                                                                    # Fallback : essayer de trouver la ligne avec une requête directe
                                                                    # (parfois find_row_by_unique_index échoue à cause de problèmes de types)
                                                                    try:
                                                                        where_parts_fallback = []
                                                                        where_params_fallback = []
                                                                        for col, val in zip(idx_cols, idx_values):
                                                                            if val is None:
                                                                                where_parts_fallback.append(f"[{col}] IS NULL")
                                                                            else:
                                                                                where_parts_fallback.append(f"[{col}] = ?")
                                                                                where_params_fallback.append(val)
                                                                        where_clause_fallback = " AND ".join(where_parts_fallback)
                                                                        target_cursor.execute(f"SELECT * FROM [{table_name}] WHERE {where_clause_fallback}", tuple(where_params_fallback))
                                                                        row_fallback = target_cursor.fetchone()
                                                                        if row_fallback:
                                                                            # Récupérer les noms de colonnes
                                                                            target_cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
                                                                            cols_fallback = [desc[0] for desc in target_cursor.description]
                                                                            existing_row = {cols_fallback[i]: row_fallback[i] for i in range(len(cols_fallback))}
                                                                    except Exception:
                                                                        pass
                                                                
                                                                if not existing_row:
                                                                    # Essayer une dernière fois avec une comparaison plus souple (types différents)
                                                                    # Parfois les types ne correspondent pas exactement (INT vs BIGINT, etc.)
                                                                    try:
                                                                        # Construire WHERE avec CONVERT pour normaliser les types
                                                                        where_parts_convert = []
                                                                        where_params_convert = []
                                                                        for col, val in zip(idx_cols, idx_values):
                                                                            if val is None:
                                                                                where_parts_convert.append(f"[{col}] IS NULL")
                                                                            else:
                                                                                # Essayer avec conversion de type pour les nombres
                                                                                if isinstance(val, (int, float)):
                                                                                    where_parts_convert.append(f"CAST([{col}] AS FLOAT) = ?")
                                                                                    where_params_convert.append(float(val))
                                                                                elif isinstance(val, str):
                                                                                    where_parts_convert.append(f"LTRIM(RTRIM([{col}])) = ?")
                                                                                    where_params_convert.append(val.strip())
                                                                                else:
                                                                                    where_parts_convert.append(f"[{col}] = ?")
                                                                                    where_params_convert.append(val)
                                                                        where_clause_convert = " AND ".join(where_parts_convert)
                                                                        target_cursor.execute(f"SELECT * FROM [{table_name}] WHERE {where_clause_convert}", tuple(where_params_convert))
                                                                        row_convert = target_cursor.fetchone()
                                                                        if row_convert:
                                                                            target_cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
                                                                            cols_convert = [desc[0] for desc in target_cursor.description]
                                                                            existing_row = {cols_convert[i]: row_convert[i] for i in range(len(cols_convert))}
                                                                    except Exception:
                                                                        pass
                                                                
                                                                if not existing_row:
                                                                    continue
                                                                
                                                                # Conflit détecté : UPDATE uniquement si différence
                                                                exclude_for_update = exclude_cols | set(idx_cols)
                                                                has_diff, diff_cols, update_values = compare_rows_for_update(
                                                                    row_dict, existing_row, exclude_for_update
                                                                )
                                                                if not has_diff:
                                                                    updated_via_unique = True
                                                                    break
                                                                
                                                                # WHERE unique avec gestion des NULL
                                                                where_parts = []
                                                                where_params = []
                                                                for col, val in zip(idx_cols, idx_values):
                                                                    if val is None:
                                                                        where_parts.append(f"[{col}] IS NULL")
                                                                    else:
                                                                        where_parts.append(f"[{col}] = ?")
                                                                        where_params.append(val)
                                                                where_unique = " AND ".join(where_parts)
                                                                
                                                                set_clauses = [f"[{col}] = ?" for col in diff_cols]
                                                                set_values = [update_values[col] for col in diff_cols]
                                                                update_sql = f"UPDATE [{table_name}] SET {', '.join(set_clauses)} WHERE {where_unique}"
                                                                target_cursor.execute(update_sql, tuple(set_values) + tuple(where_params))
                                                                
                                                                updated_ultime += 1
                                                                if updated_ultime % 100 == 0:
                                                                    target_conn.commit()
                                                                updated_via_unique = True
                                                                break
                                                            
                                                            if not updated_via_unique:
                                                                # Logging de diagnostic : pourquoi l'UPDATE n'a pas fonctionné
                                                                if len(ultime_diag["insert_samples"]) < 3:
                                                                    # Ajouter plus de détails sur pourquoi l'UPDATE a échoué
                                                                    diag_msg = f"Duplicate key mais UPDATE impossible: {str(ins_err)[:100]}"
                                                                    if unique_indexes:
                                                                        diag_msg += f" (index uniques testés: {len(unique_indexes)})"
                                                                    ultime_diag["insert_samples"].append(diag_msg)
                                                            else:
                                                                if len(ultime_diag["insert_samples"]) < 3:
                                                                    ultime_diag["insert_samples"].append(str(ins_err))
                                                            continue
                                            
                                            if inserted > 0 or updated_ultime > 0:
                                                target_conn.commit()
                                else:
                                    # PK composite - lire tous et filtrer
                                    source_cursor.execute(f"SELECT {col_list} FROM [{table_name}]")
                                    source_rows = source_cursor.fetchall()
                                    
                                    pk_indices = [columns.index(pk) for pk in pk_columns]
                                    placeholders = ", ".join(["?" for _ in columns])
                                    
                                    for row in source_rows:
                                        pk_tuple = tuple(row[idx] for idx in pk_indices)
                                        if pk_tuple not in existing_pks:
                                            row_dict = {columns[i]: row[i] for i in range(len(columns))}
                                            
                                            try:
                                                target_cursor.execute(
                                                    f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                                    list(row)
                                                )
                                                inserted += 1
                                                existing_pks.add(pk_tuple)
                                                if inserted % 100 == 0:
                                                    target_conn.commit()
                                            except Exception as ins_err:
                                                err_str = str(ins_err)
                                                
                                                # Si erreur de conflit d'index unique, essayer UPDATE (Option A)
                                                if 'duplicate key' in err_str.lower() or 'unique index' in err_str.lower() or '2601' in err_str:
                                                    updated_via_unique = False
                                                    
                                                    for idx_name, idx_cols in unique_indexes:
                                                        if not idx_cols:
                                                            continue
                                                        
                                                        # Construire les valeurs de l'index (NULL inclus)
                                                        idx_values = []
                                                        has_all = True
                                                        for col in idx_cols:
                                                            if col not in row_dict:
                                                                has_all = False
                                                                break
                                                            idx_values.append(row_dict.get(col))
                                                        if not has_all:
                                                            continue
                                                        
                                                        # Utiliser find_row_by_unique_index pour gérer les NULL correctement
                                                        existing_row = find_row_by_unique_index(
                                                            target_cursor, table_name, idx_cols, idx_values
                                                        )
                                                        if not existing_row:
                                                            # Fallback : essayer de trouver la ligne avec une requête directe
                                                            try:
                                                                where_parts_fallback = []
                                                                where_params_fallback = []
                                                                for col, val in zip(idx_cols, idx_values):
                                                                    if val is None:
                                                                        where_parts_fallback.append(f"[{col}] IS NULL")
                                                                    else:
                                                                        where_parts_fallback.append(f"[{col}] = ?")
                                                                        where_params_fallback.append(val)
                                                                where_clause_fallback = " AND ".join(where_parts_fallback)
                                                                target_cursor.execute(f"SELECT * FROM [{table_name}] WHERE {where_clause_fallback}", tuple(where_params_fallback))
                                                                row_fallback = target_cursor.fetchone()
                                                                if row_fallback:
                                                                    target_cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
                                                                    cols_fallback = [desc[0] for desc in target_cursor.description]
                                                                    existing_row = {cols_fallback[i]: row_fallback[i] for i in range(len(cols_fallback))}
                                                            except Exception:
                                                                pass
                                                        
                                                        if not existing_row:
                                                            continue
                                                        
                                                        # Conflit détecté : UPDATE uniquement si différence
                                                        exclude_for_update = exclude_cols | set(idx_cols)
                                                        has_diff, diff_cols, update_values = compare_rows_for_update(
                                                            row_dict, existing_row, exclude_for_update
                                                        )
                                                        if not has_diff:
                                                            updated_via_unique = True
                                                            break
                                                        
                                                        # WHERE unique avec gestion des NULL
                                                        where_parts = []
                                                        where_params = []
                                                        for col, val in zip(idx_cols, idx_values):
                                                            if val is None:
                                                                where_parts.append(f"[{col}] IS NULL")
                                                            else:
                                                                where_parts.append(f"[{col}] = ?")
                                                                where_params.append(val)
                                                        where_unique = " AND ".join(where_parts)
                                                        
                                                        set_clauses = [f"[{col}] = ?" for col in diff_cols]
                                                        set_values = [update_values[col] for col in diff_cols]
                                                        update_sql = f"UPDATE [{table_name}] SET {', '.join(set_clauses)} WHERE {where_unique}"
                                                        target_cursor.execute(update_sql, tuple(set_values) + tuple(where_params))
                                                        
                                                        updated_ultime += 1
                                                        if updated_ultime % 100 == 0:
                                                            target_conn.commit()
                                                        updated_via_unique = True
                                                        break
                                                    
                                                    if not updated_via_unique:
                                                        if len(ultime_diag["insert_samples"]) < 3:
                                                            ultime_diag["insert_samples"].append(str(ins_err))
                                                else:
                                                    if len(ultime_diag["insert_samples"]) < 3:
                                                        ultime_diag["insert_samples"].append(str(ins_err))
                                                continue
                                    
                                    if inserted > 0 or updated_ultime > 0:
                                        target_conn.commit()
                            else:
                                    # Lire tous les enregistrements de la source
                                    source_cursor.execute(f"SELECT {col_list} FROM [{table_name}]")
                                    source_rows = source_cursor.fetchall()
                                    
                                    pk_indices = [columns.index(pk) for pk in pk_columns]
                                    placeholders = ", ".join(["?" for _ in columns])
                                    
                                    # Index uniques disponibles pour résoudre les conflits (INSERT -> UPDATE)
                                    unique_indexes_ultime = _get_unique_indexes(target_cursor, table_name)
                                    exclude_cols_ultime = set(pk_columns)
                                    
                                    updated_ultime_2 = 0
                                    
                                    for row in source_rows:
                                        row_dict = {columns[i]: row[i] for i in range(len(columns))}
                                        
                                        # Vérifier si déjà présent
                                        if len(pk_columns) == 1:
                                            pk_value = row[pk_indices[0]]
                                            if pk_value in existing_pks:
                                                continue
                                        else:
                                            pk_tuple = tuple(row[idx] for idx in pk_indices)
                                            if pk_tuple in existing_pks:
                                                continue
                                        
                                        # Insérer sans vérifier les FK
                                        try:
                                            target_cursor.execute(
                                                f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                                list(row)
                                            )
                                            inserted += 1
                                            if len(pk_columns) == 1:
                                                existing_pks.add(row[pk_indices[0]])
                                            else:
                                                existing_pks.add(tuple(row[idx] for idx in pk_indices))
                                            if inserted % 100 == 0:
                                                target_conn.commit()
                                        except Exception as ins_err:
                                            err_str = str(ins_err)
                                            
                                            # Si erreur de conflit d'index unique, essayer UPDATE (Option A)
                                            if 'duplicate key' in err_str.lower() or 'unique index' in err_str.lower() or '2601' in err_str:
                                                updated_via_unique = False

                                                for idx_name, idx_cols in unique_indexes_ultime:
                                                    if not idx_cols:
                                                        continue

                                                    # Construire les valeurs de l'index (NULL inclus)
                                                    idx_values = []
                                                    has_all = True
                                                    for col in idx_cols:
                                                        if col not in row_dict:
                                                            has_all = False
                                                            break
                                                        idx_values.append(row_dict.get(col))
                                                    if not has_all:
                                                        continue

                                                    existing_row = find_row_by_unique_index(
                                                        target_cursor, table_name, idx_cols, idx_values
                                                    )
                                                    if not existing_row:
                                                        continue

                                                    # Conflit détecté : UPDATE uniquement si différence
                                                    exclude_for_update = exclude_cols_ultime | set(idx_cols)
                                                    has_diff, diff_cols, update_values = compare_rows_for_update(
                                                        row_dict, existing_row, exclude_for_update
                                                    )
                                                    if not has_diff:
                                                        updated_via_unique = True
                                                        break

                                                    # WHERE unique avec gestion des NULL
                                                    where_parts = []
                                                    where_params = []
                                                    for col, val in zip(idx_cols, idx_values):
                                                        if val is None:
                                                            where_parts.append(f"[{col}] IS NULL")
                                                        else:
                                                            where_parts.append(f"[{col}] = ?")
                                                            where_params.append(val)
                                                    where_unique = " AND ".join(where_parts)

                                                    set_clauses = [f"[{col}] = ?" for col in diff_cols]
                                                    set_values = [update_values[col] for col in diff_cols]
                                                    update_sql = f"UPDATE [{table_name}] SET {', '.join(set_clauses)} WHERE {where_unique}"
                                                    target_cursor.execute(update_sql, tuple(set_values) + tuple(where_params))

                                                    updated_ultime_2 += 1
                                                    if updated_ultime_2 % 100 == 0:
                                                        target_conn.commit()
                                                    updated_via_unique = True
                                                    break
                                                
                                                if not updated_via_unique:
                                                    if len(ultime_diag["insert_samples"]) < 3:
                                                        ultime_diag["insert_samples"].append(str(ins_err))
                                            else:
                                                if len(ultime_diag["insert_samples"]) < 3:
                                                    ultime_diag["insert_samples"].append(str(ins_err))
                                            # Ignorer les erreurs (doublons, types non supportés, etc.)
                                            continue
                                    
                                    if inserted > 0 or updated_ultime_2 > 0:
                                        target_conn.commit()
                                    updated_ultime += updated_ultime_2
                        except Exception as read_err:
                                ultime_diag["read"] = str(read_err)
                                # Si la lecture échoue (types non supportés), essayer avec conversion SQL
                                try:
                                    # Construire les colonnes avec conversion
                                    converted_cols = []
                                    for col_name, data_type, max_len, nullable, default in cols_info:
                                        if data_type in ('timestamp', 'rowversion'):
                                            converted_cols.append(f"CONVERT(VARBINARY(8), [{col_name}]) AS [{col_name}]")
                                        elif data_type in ('sql_variant', 'xml', 'text', 'ntext'):
                                            converted_cols.append(f"CONVERT(NVARCHAR(MAX), CAST([{col_name}] AS NVARCHAR(MAX))) AS [{col_name}]")
                                        elif data_type in ('geography', 'geometry', 'hierarchyid', 'image'):
                                            converted_cols.append(f"CONVERT(VARBINARY(MAX), [{col_name}]) AS [{col_name}]")
                                        else:
                                            converted_cols.append(f"[{col_name}]")
                                    
                                    converted_col_list = ", ".join(converted_cols)
                                    
                                    # Si on a la liste des PKs manquantes, filtrer directement
                                    if missing_pks_list and len(missing_pks_list) > 0 and len(pk_columns) == 1:
                                        pk_values = [pk for pk in missing_pks_list if pk is not None]
                                        # Traiter par lots de 1000
                                        for batch in [pk_values[i:i+1000] for i in range(0, len(pk_values), 1000)]:
                                            pk_strs = []
                                            for pk_val in batch:
                                                if isinstance(pk_val, str):
                                                    pk_strs.append(f"'{pk_val.replace("'", "''")}'")
                                                else:
                                                    pk_strs.append(str(pk_val))
                                            where_clause = f"WHERE [{pk_columns[0]}] IN ({','.join(pk_strs)})"
                                            source_cursor.execute(f"SELECT {converted_col_list} FROM [{table_name}] {where_clause}")
                                            source_rows = source_cursor.fetchall()
                                            
                                            pk_indices = [columns.index(pk) for pk in pk_columns]
                                            placeholders = ", ".join(["?" for _ in columns])
                                            
                                            for row in source_rows:
                                                try:
                                                    target_cursor.execute(
                                                        f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                                        list(row)
                                                    )
                                                    inserted += 1
                                                    if inserted % 100 == 0:
                                                        target_conn.commit()
                                                except Exception as ins_err:
                                                    if len(ultime_diag["insert_samples"]) < 3:
                                                        ultime_diag["insert_samples"].append(str(ins_err))
                                                    continue
                                        
                                        if inserted > 0:
                                            target_conn.commit()
                                    else:
                                        # Lire tous et filtrer
                                        source_cursor.execute(f"SELECT {converted_col_list} FROM [{table_name}]")
                                        source_rows = source_cursor.fetchall()
                                        
                                        pk_indices = [columns.index(pk) for pk in pk_columns]
                                        placeholders = ", ".join(["?" for _ in columns])
                                        
                                        batch_size = 100
                                        for idx, row in enumerate(source_rows):
                                            if len(pk_columns) == 1:
                                                pk_value = row[pk_indices[0]]
                                                if pk_value in existing_pks:
                                                    continue
                                            else:
                                                pk_tuple = tuple(row[idx] for idx in pk_indices)
                                                if pk_tuple in existing_pks:
                                                    continue
                                            
                                            try:
                                                target_cursor.execute(
                                                    f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                                    list(row)
                                                )
                                                inserted += 1
                                                if len(pk_columns) == 1:
                                                    existing_pks.add(row[pk_indices[0]])
                                                else:
                                                    existing_pks.add(tuple(row[idx] for idx in pk_indices))
                                                
                                                # Commit périodique pour éviter les timeouts
                                                if inserted % batch_size == 0:
                                                    target_conn.commit()
                                            except Exception as insert_err:
                                                if len(ultime_diag["insert_samples"]) < 3:
                                                    ultime_diag["insert_samples"].append(str(insert_err))
                                                # Ignorer les erreurs d'insertion (doublons, contraintes, etc.)
                                                continue
                                        
                                        if inserted > 0:
                                            target_conn.commit()
                                except Exception as convert_err:
                                    ultime_diag["convert"] = str(convert_err)
                                    # Si même la conversion échoue, essayer une dernière fois avec l'approche ligne par ligne
                                    # en ignorant les colonnes problématiques si nécessaire
                                    sync_status['details'].append(f"  ⚠ Conversion échouée, tentative ligne par ligne avec gestion d'erreurs...")
                                    try:
                                        # Utiliser missing_pks_list si disponible, sinon recalculer
                                        if missing_pks_list:
                                            missing_pks = set(missing_pks_list)
                                        else:
                                            missing_pks = source_pks - existing_pks
                                        
                                        if missing_pks and len(pk_columns) == 1:
                                            pk_values = [pk for pk in missing_pks if pk is not None]
                                            if pk_values:
                                                # Traiter par lots de 1000
                                                for batch in [pk_values[i:i+1000] for i in range(0, len(pk_values), 1000)]:
                                                    pk_strs = []
                                                    for pk_val in batch:
                                                        if isinstance(pk_val, str):
                                                            pk_strs.append(f"'{pk_val.replace("'", "''")}'")
                                                        else:
                                                            pk_strs.append(str(pk_val))
                                                    where_clause = f"WHERE [{pk_columns[0]}] IN ({','.join(pk_strs)})"
                                                    
                                                    try:
                                                        # Essayer de lire les données même avec types problématiques
                                                        source_cursor.execute(f"SELECT {col_list} FROM [{table_name}] {where_clause}")
                                                        source_rows = source_cursor.fetchall()
                                                        
                                                        pk_indices = [columns.index(pk) for pk in pk_columns]
                                                        placeholders = ", ".join(["?" for _ in columns])
                                                        
                                                        for row in source_rows:
                                                            try:
                                                                target_cursor.execute(
                                                                    f"INSERT INTO [{table_name}] ({col_list}) VALUES ({placeholders})",
                                                                    list(row)
                                                                )
                                                                inserted += 1
                                                                if inserted % 100 == 0:
                                                                    target_conn.commit()
                                                            except Exception as ins_err:
                                                                # Ignorer les erreurs d'insertion (types, contraintes, etc.)
                                                                if len(ultime_diag["insert_samples"]) < 3:
                                                                    ultime_diag["insert_samples"].append(str(ins_err))
                                                                continue
                                                        
                                                        if inserted > 0:
                                                            target_conn.commit()
                                                    except Exception as batch_read_err:
                                                        # Si même la lecture par batch échoue, ignorer ce batch
                                                        if len(ultime_diag["read"]) < 200:
                                                            ultime_diag["read"] = str(batch_read_err)
                                                        continue
                                    except Exception as final_err:
                                        sync_status['details'].append(f"  ⚠ Impossible d'insérer les enregistrements restants: {str(final_err)[:100]}")
                                        pass
                        
                        if inserted > 0 or updated_ultime > 0:
                            target_conn.commit()
                            msg_parts = []
                            if inserted > 0:
                                msg_parts.append(f"{inserted} ajoutés")
                            if updated_ultime > 0:
                                msg_parts.append(f"{updated_ultime} mis à jour")
                            sync_status['details'].append(f"✓ {table_name} (Ultime): {', '.join(msg_parts)} (FK désactivées)")
                        else:
                            # Vérifier s'il y avait vraiment des enregistrements manquants
                            if missing_count > 0:
                                sync_status['details'].append(f"⚠ {table_name} (Ultime): {missing_count} enregistrements manquants non insérés.")
                                # Ajouter un diagnostic concis (ce qui bloque réellement)
                                if ultime_diag.get("openrowset"):
                                    sync_status['details'].append(f"   - Détail OPENROWSET: {ultime_diag['openrowset'][:180]}")
                                if ultime_diag.get("read"):
                                    sync_status['details'].append(f"   - Détail lecture source: {ultime_diag['read'][:180]}")
                                if ultime_diag.get("convert"):
                                    sync_status['details'].append(f"   - Détail conversion: {ultime_diag['convert'][:180]}")
                                if ultime_diag.get("insert_samples"):
                                    for j, msg in enumerate(ultime_diag["insert_samples"][:3], start=1):
                                        sync_status['details'].append(f"   - Exemple erreur INSERT {j}: {msg[:180]}")
                                if ultime_diag.get("openrowset_pk_samples"):
                                    for j, msg in enumerate(ultime_diag["openrowset_pk_samples"][:3], start=1):
                                        sync_status['details'].append(f"   - Exemple erreur OPENROWSET/PK {j}: {msg[:180]}")
                            else:
                                sync_status['details'].append(f"○ {table_name} (Ultime): à jour")
                    finally:
                        # Réactiver les contraintes FK
                        for fk_name in disabled_fks:
                            try:
                                target_cursor.execute(f"ALTER TABLE [{table_name}] CHECK CONSTRAINT [{fk_name}]")
                            except:
                                pass
                        
                        if identity_enabled:
                            target_cursor.execute(f"SET IDENTITY_INSERT [{table_name}] OFF")
                            
                except Exception as e:
                    sync_status['details'].append(f"✗ {table_name} (Ultime): {str(e)}")
        
        # Supprimer en cible les lignes dont la PK n'existe plus dans la source (enfants → parents : ordre inverse du tri topo)
        sync_status['details'].append(
            "\n🧹 Suppression des enregistrements présents en cible mais absents de la source (orphelins)..."
        )
        target_table_set = set(all_target_tables)
        orphans_deleted_total = 0
        for table_name in reversed(sorted_tables):
            if table_name not in target_table_set:
                continue
            pk_orphan = get_primary_keys(source_cursor, table_name)
            if not pk_orphan:
                continue
            try:
                n_del = sync_delete_orphan_rows_from_table(
                    source_cursor, target_cursor, target_conn, table_name, pk_orphan, sync_status
                )
                if n_del > 0:
                    sync_status['details'].append(f"🧹 {table_name}: {n_del} orphelin(s) supprimé(s)")
                    orphans_deleted_total += n_del
            except Exception as ex_or:
                sync_status['details'].append(f"  ⚠ Orphelins {table_name}: {str(ex_or)[:200]}")
        if orphans_deleted_total == 0:
            sync_status['details'].append("○ Aucun orphelin supprimé (ou tables ignorées)")
        
        source_conn.close()
        target_conn.close()
        
        sync_status['message'] = 'Synchronisation terminée'
        sync_status['progress'] = 100
        
    except Exception as e:
        sync_status['message'] = f'Erreur: {str(e)}'
        sync_status['details'].append(f"Erreur globale: {str(e)}")
    
    sync_status['running'] = False

@projet21_bp.route('/')
def index():
    return render_template('projet21/index.html')

@projet21_bp.route('/sync', methods=['POST'])
def start_sync():
    global sync_status
    if sync_status['running']:
        return jsonify({'error': 'Synchronisation déjà en cours'}), 400
    
    thread = threading.Thread(target=sync_databases)
    thread.start()
    return jsonify({'status': 'started'})

@projet21_bp.route('/status')
def get_status():
    # Toujours exposer la version de code réellement chargée,
    # même si sync_status a été réinitialisé ailleurs.
    payload = dict(sync_status) if isinstance(sync_status, dict) else {'running': False, 'message': '', 'progress': 0, 'details': []}
    payload['code_version'] = PROJET21_CODE_VERSION
    return jsonify(payload)

@projet21_bp.route('/xrt/sync', methods=['POST'])
def start_xrt_sync():
    global xrt_sync_status
    if xrt_sync_status.get('running'):
        return jsonify({'error': 'Synchronisation XRT déjà en cours'}), 400
    thread = threading.Thread(target=xrt_sync_databases_incremental)
    thread.start()
    return jsonify({'status': 'started'})


@projet21_bp.route('/xrt/sync-full', methods=['POST'])
def start_xrt_sync_full():
    """
    Synchronisation XRT « plein » : vide le schéma XRT puis recopie tout.
    Réservé aux cas exceptionnels (pas exposé dans l’UI standard).
    """
    global xrt_sync_status
    if xrt_sync_status.get('running'):
        return jsonify({'error': 'Synchronisation XRT déjà en cours'}), 400
    thread = threading.Thread(target=xrt_sync_databases)
    thread.start()
    return jsonify({'status': 'started'})


@projet21_bp.route('/xrt/sync-missing', methods=['POST'])
def start_xrt_sync_missing():
    global xrt_sync_status
    if xrt_sync_status.get('running'):
        return jsonify({'error': 'Synchronisation XRT déjà en cours'}), 400
    thread = threading.Thread(target=xrt_sync_databases_missing_only)
    thread.start()
    return jsonify({'status': 'started'})

@projet21_bp.route('/xrt/sync-mismatch', methods=['POST'])
def start_xrt_sync_mismatch():
    global xrt_sync_status
    if xrt_sync_status.get('running'):
        return jsonify({'error': 'Synchronisation XRT déjà en cours'}), 400
    thread = threading.Thread(target=xrt_sync_databases_count_mismatch_only)
    thread.start()
    return jsonify({'status': 'started'})

@projet21_bp.route('/xrt/status')
def get_xrt_status():
    payload = dict(xrt_sync_status) if isinstance(xrt_sync_status, dict) else {'running': False, 'message': '', 'progress': 0, 'details': []}
    payload['code_version'] = PROJET21_CODE_VERSION
    return jsonify(payload)

@projet21_bp.route('/xrt/verify', methods=['POST'])
def verify_xrt_sync():
    """
    Vérifie la synchronisation XRT : comptes source vs cible, puis si égalité des comptes,
    agrégat CHECKSUM_AGG(BINARY_CHECKSUM(*)) comme indicateur de divergence de contenu.
    La source reste strictement en lecture seule.
    """
    try:
        xrt_cfg = get_xrt_source_runtime_config()
        source_conn = get_connection(xrt_cfg, readonly=True)
        source_cursor = source_conn.cursor()

        target_conn = get_connection(TARGET_CONFIG)
        target_cursor = target_conn.cursor()
        ensure_target_schema(target_cursor, 'XRT')

        tables = xrt_list_source_tables(source_cursor)

        ok = 0
        ok_count_only = 0  # comptages égaux, checksum NULL des deux côtés (non = échec de synchro)
        ko_absent = 0
        ko_mismatch = 0
        ko_checksum = 0
        ko_checksum_na = 0
        ko_other = 0
        lines = []
        lines.append("=== Vérification synchronisation XRT ===")
        lines.append(f"Date/heure: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Source: {xrt_cfg['server']} / {xrt_cfg['database']} (READ ONLY)")
        lines.append(f"Cible: {TARGET_CONFIG['server']} / {TARGET_CONFIG['database']} (schema XRT)")
        lines.append(
            "Méthode: COUNT puis, si comptes égaux et table non vide, CHECKSUM_AGG(CAST(BINARY_CHECKSUM(*) AS BIGINT))."
        )
        lines.append(
            "Si le checksum ne peut pas être calculé (NULL) côté source et cible, seul le comptage est retenu (○) — "
            "ce n’est pas un échec de synchronisation."
        )
        lines.append("")

        for schema_name, table_name in tables:
            tgt_name = xrt_target_table_name(schema_name, table_name)
            try:
                source_cursor.execute(f"SELECT COUNT(*) FROM [{schema_name}].[{table_name}]")
                src_count = int(source_cursor.fetchone()[0] or 0)

                if not xrt_target_table_exists(target_cursor, 'XRT', tgt_name):
                    ko_absent += 1
                    lines.append(
                        f"✗ {schema_name}.{table_name} -> XRT.{tgt_name}: absente en cible "
                        f"(source: {src_count} ligne(s))"
                    )
                    continue

                target_cursor.execute(f"SELECT COUNT(*) FROM [XRT].[{tgt_name}]")
                tgt_count = int(target_cursor.fetchone()[0] or 0)

                if src_count != tgt_count:
                    ko_mismatch += 1
                    lines.append(
                        f"✗ {schema_name}.{table_name} -> XRT.{tgt_name}: "
                        f"comptage source={src_count} cible={tgt_count}"
                    )
                    continue

                # Comptages égaux
                if src_count == 0:
                    ok += 1
                    lines.append(
                        f"✓ {schema_name}.{table_name} -> XRT.{tgt_name}: vide aligné (0 ligne)"
                    )
                    continue

                cs_s = xrt_table_checksum_agg(source_cursor, schema_name, table_name)
                cs_t = xrt_table_checksum_agg(target_cursor, 'XRT', tgt_name)

                if cs_s is None and cs_t is None:
                    ok += 1
                    ok_count_only += 1
                    lines.append(
                        f"○ {schema_name}.{table_name} -> XRT.{tgt_name}: {src_count} ligne(s), "
                        f"comptage OK — checksum agrégé indisponible (BINARY_CHECKSUM limité selon types de colonnes)"
                    )
                    continue

                if cs_s is None or cs_t is None:
                    ko_checksum_na += 1
                    lines.append(
                        f"✗ {schema_name}.{table_name} -> XRT.{tgt_name}: {src_count} ligne(s), "
                        f"checksum partiellement indisponible (src={cs_s!r} cible={cs_t!r}) "
                        f"— situation inhabituelle, comparer ou resynchroniser"
                    )
                    continue

                if cs_s != cs_t:
                    ko_checksum += 1
                    lines.append(
                        f"✗ {schema_name}.{table_name} -> XRT.{tgt_name}: comptage OK ({src_count}) "
                        f"mais checksum différent (source={cs_s} cible={cs_t})"
                    )
                else:
                    ok += 1
                    lines.append(
                        f"✓ {schema_name}.{table_name} -> XRT.{tgt_name}: {src_count} ligne(s), checksum aligné"
                    )
            except Exception as e_tbl:
                ko_other += 1
                lines.append(f"✗ {schema_name}.{table_name}: erreur vérif: {str(e_tbl)[:200]}")

        ko = ko_absent + ko_mismatch + ko_checksum + ko_checksum_na + ko_other
        lines.append("")
        lines.append(
            f"Résumé: OK={ok} | KO={ko} | Tables={len(tables)} "
            f"(absentes en XRT: {ko_absent}, comptes différents: {ko_mismatch}, "
            f"contenu différent (checksum): {ko_checksum}, "
            f"checksum partiellement indisponible (KO): {ko_checksum_na}, "
            f"OK sur comptage seul (checksum NULL des deux côtés): {ok_count_only}, "
            f"autres erreurs: {ko_other})"
        )
        note_chk = (
            "Note: collision BINARY_CHECKSUM rare ; ○ = même nombre de lignes sans contrôle checksum SQL complet."
        )
        lines.append(note_chk)

        if ko_absent and not ko_mismatch and not ko_checksum and not ko_checksum_na and not ko_other:
            lines.append(
                "→ Interprétation: les KO sont des tables non créées en schéma XRT (synchro incomplète). "
                "Utilisez « Synchroniser XRT » puis relancez la vérification."
            )
        if ko_mismatch and not ko_absent and not ko_checksum and not ko_checksum_na and not ko_other:
            lines.append(
                "→ Interprétation: écarts de comptage typiques = activité sur la source entre synchro et vérif, "
                "ou résidus (*_WM). Utilisez « Synchroniser XRT » (miroir par PK / checksum, orphelins ; "
                "recopie complète seulement si table volumineuse ou sans PK), puis revérifiez."
            )
        if ko_checksum and not ko_absent and not ko_mismatch and not ko_checksum_na and not ko_other:
            lines.append(
                "→ Interprétation: même nombre de lignes mais données différentes (mises à jour source, "
                "remplacements de lignes, etc.). Utilisez « Synchroniser XRT » pour réaligner, puis revérifiez."
            )
        if ko_checksum_na and not ko_absent and not ko_mismatch and not ko_checksum and not ko_other:
            lines.append(
                "→ Interprétation: checksum NULL d’un seul côté (rare). Contrôlez la table ou resynchronisez ; "
                "les lignes ○ (checksum NULL des deux côtés mais même comptage) sont des OK, pas des KO."
            )
        if ko_absent and ko_mismatch and not ko_checksum and not ko_checksum_na and not ko_other:
            lines.append(
                "→ Interprétation: tables absentes et écarts de comptage. « Synchroniser XRT » complète les "
                "manquantes et réaligne les divergences (miroir ou recopie selon le cas), puis revérifiez."
            )
        mixed = sum(
            1
            for x in (ko_absent, ko_mismatch, ko_checksum, ko_checksum_na, ko_other)
            if x > 0
        )
        if mixed >= 2 and ko > 0:
            lines.append(
                "→ Interprétation (situation mixte): combinez les actions ci-dessus selon les catégories du résumé ; "
                "« Synchroniser XRT » traite en général comptages, contenu (checksum) et orphelins."
            )

        try:
            source_conn.close()
        except:
            pass
        try:
            target_conn.close()
        except:
            pass

        return jsonify({'success': True, 'output': "\n".join(lines)})
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'traceback': traceback.format_exc()}), 500

@projet21_bp.route('/xrt/ping', methods=['GET'])
def xrt_ping():
    """
    Diagnostic de connectivité XRT (lecture seule).
    Retourne un JSON détaillé, sans toucher aux données.
    """
    try:
        return jsonify({'success': True, 'diagnostic': xrt_connection_diagnostics()})
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'traceback': traceback.format_exc()}), 500

@projet21_bp.route('/verify', methods=['POST'])
def verify_sync():
    """Vérifie la synchronisation en comparant les comptes de toutes les tables"""
    from routes.projet21_verification import verify_sync
    
    try:
        results = verify_sync()
        
        total_manquants = 0
        for item in results.get('ecarts_critiques', []):
            if len(item) >= 4:
                total_manquants += item[3]  # missing_count
        
        total_doublons = 0
        for item in results.get('doublons_pk', []):
            if len(item) >= 3:
                total_doublons += item[2]  # duplicate_rows
        
        return jsonify({
            'success': True,
            'output': results.get('output', ''),
            'summary': {
                'synchronisees': len(results['synchronisees']),
                'ecarts_critiques': len(results.get('ecarts_critiques', [])),
                'ecarts_normaux': len(results.get('ecarts_normaux', [])),
                'doublons_pk': len(results.get('doublons_pk', [])),
                'manquantes_cible': len(results['manquantes_cible']),
                'manquantes_source': len(results['manquantes_source']),
                'total_manquants': total_manquants,
                'total_doublons': total_doublons
            },
            'details': {
                'synchronisees': results['synchronisees'],
                'ecarts_critiques': results.get('ecarts_critiques', []),
                'ecarts_normaux': results.get('ecarts_normaux', []),
                'doublons_pk': results.get('doublons_pk', []),
                'manquantes_cible': results['manquantes_cible'],
                'manquantes_source': results['manquantes_source']
            }
        })
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'traceback': traceback.format_exc()}), 500


def _normalize_pyodbc_value(v):
    # Valeurs JSON-safe
    if isinstance(v, (datetime,)):
        return v.isoformat(sep=" ", timespec="seconds")
    return v


def _fetch_missing_pks_for_table(source_cursor, target_cursor, table_name):
    """
    Retourne (pk_columns, missing_pks_sorted).
    missing_pks_sorted: liste (peut être valeurs simples ou tuples).
    """
    pk_columns = get_primary_keys(source_cursor, table_name)
    if not pk_columns:
        # Pas de PK -> pas supporté pour "lignes manquantes" en mode efficace
        return [], []

    pk_list = ", ".join([f"[{pk}]" for pk in pk_columns])

    source_cursor.execute(f"SELECT {pk_list} FROM [{table_name}]")
    source_pks = set()
    for row in source_cursor.fetchall():
        if len(pk_columns) == 1:
            source_pks.add(row[0])
        else:
            source_pks.add(tuple(row))

    target_cursor.execute(f"SELECT {pk_list} FROM [{table_name}]")
    target_pks = set()
    for row in target_cursor.fetchall():
        if len(pk_columns) == 1:
            target_pks.add(row[0])
        else:
            target_pks.add(tuple(row))

    missing = list(source_pks - target_pks)
    try:
        missing.sort()
    except Exception:
        # types mixtes => tri non garanti
        pass
    return pk_columns, missing


def _fetch_extra_pks_for_table(source_cursor, target_cursor, table_name):
    """
    Retourne (pk_columns, extra_pks_sorted).
    extra_pks_sorted: PK présentes dans la CIBLE mais absentes en SOURCE.
    """
    pk_columns = get_primary_keys(source_cursor, table_name)
    if not pk_columns:
        return [], []

    pk_list = ", ".join([f"[{pk}]" for pk in pk_columns])

    source_cursor.execute(f"SELECT {pk_list} FROM [{table_name}]")
    source_pks = set()
    for row in source_cursor.fetchall():
        if len(pk_columns) == 1:
            source_pks.add(row[0])
        else:
            source_pks.add(tuple(row))

    target_cursor.execute(f"SELECT {pk_list} FROM [{table_name}]")
    target_pks = set()
    for row in target_cursor.fetchall():
        if len(pk_columns) == 1:
            target_pks.add(row[0])
        else:
            target_pks.add(tuple(row))

    extra = list(target_pks - source_pks)
    try:
        extra.sort()
    except Exception:
        pass
    return pk_columns, extra


def _select_rows_by_pks(cursor, table_name, pk_columns, pks):
    """
    Récupère les lignes complètes depuis la base (source) pour une liste de PK.
    """
    if not pks:
        return [], []

    if len(pk_columns) == 1:
        pk = pk_columns[0]
        placeholders = ", ".join(["?"] * len(pks))
        sql = f"SELECT * FROM [{table_name}] WHERE [{pk}] IN ({placeholders}) ORDER BY [{pk}]"
        cursor.execute(sql, tuple(pks))
    else:
        # composite PK: (A=? AND B=?) OR (A=? AND B=?)
        clauses = []
        params = []
        for pk_tuple in pks:
            sub = []
            for i, pk in enumerate(pk_columns):
                sub.append(f"[{pk}] = ?")
                params.append(pk_tuple[i])
            clauses.append("(" + " AND ".join(sub) + ")")
        sql = f"SELECT * FROM [{table_name}] WHERE " + " OR ".join(clauses)
        cursor.execute(sql, tuple(params))

    columns = [desc[0] for desc in cursor.description]
    rows = []
    for r in cursor.fetchall():
        rows.append({columns[i]: _normalize_pyodbc_value(r[i]) for i in range(len(columns))})
    return columns, rows


def _get_unique_indexes(target_cursor, table_name):
    """
    Retourne une liste d'index uniques (hors PK) sous forme de listes de colonnes.
    """
    target_cursor.execute("""
        SELECT i.name, c.name, ic.key_ordinal
        FROM sys.indexes i
        INNER JOIN sys.index_columns ic
            ON i.object_id = ic.object_id AND i.index_id = ic.index_id
        INNER JOIN sys.columns c
            ON ic.object_id = c.object_id AND ic.column_id = c.column_id
        INNER JOIN sys.tables t
            ON i.object_id = t.object_id
        WHERE t.name = ?
          AND i.is_unique = 1
          AND i.is_primary_key = 0
          AND ic.key_ordinal > 0
        ORDER BY i.name, ic.key_ordinal
    """, (table_name,))
    rows = target_cursor.fetchall()
    by_index = {}
    for idx_name, col_name, key_ordinal in rows:
        by_index.setdefault(idx_name, []).append((key_ordinal, col_name))
    indexes = []
    for idx_name, cols in by_index.items():
        cols_sorted = [c for _, c in sorted(cols, key=lambda x: x[0])]
        if cols_sorted:
            indexes.append((idx_name, cols_sorted))
    return indexes


def _get_column_types(source_cursor, table_name):
    cols = get_table_columns(source_cursor, table_name)
    # cols: (COLUMN_NAME, DATA_TYPE, ...)
    return {c[0]: (c[1] or '').lower() for c in cols}


def _diagnose_missing_row(table_name, row_dict, source_types, target_cursor):
    """
    Heuristique "cause de non synchronisation" (non destructive) :
    - conflit de clé unique (existe déjà dans cible sous un autre ID)
    - FK manquante (référence absente en cible)
    - types potentiellement problématiques
    """
    # 1) Conflit de clé unique (index unique hors PK)
    try:
        unique_indexes = _get_unique_indexes(target_cursor, table_name)
        for idx_name, cols in unique_indexes:
            values = []
            has_all = True
            has_null = False
            for c in cols:
                if c not in row_dict:
                    has_all = False
                    break
                v = row_dict.get(c)
                if v is None:
                    has_null = True
                values.append(v)
            if not has_all:
                continue
            # si une des colonnes est NULL, SQL Server autorise souvent plusieurs NULL selon index,
            # donc ce test est moins fiable
            if has_null:
                continue
            where = " AND ".join([f"[{c}] = ?" for c in cols])
            sql = f"SELECT TOP 1 1 FROM [{table_name}] WHERE {where}"
            target_cursor.execute(sql, tuple(values))
            if target_cursor.fetchone():
                return f"Conflit clé unique ({idx_name}) : la ligne existe déjà dans la cible (mêmes {cols})"
    except Exception:
        pass

    # 2) FK manquante
    try:
        fks = get_foreign_keys(target_cursor, table_name)
        missing_fks = []
        for fk_name, parent_table, parent_col, ref_table, ref_col in fks:
            if parent_col not in row_dict:
                continue
            v = row_dict.get(parent_col)
            if v is None:
                continue
            if not check_fk_exists(target_cursor, ref_table, ref_col, v):
                missing_fks.append(f"{parent_col} -> {ref_table}.{ref_col}={v}")
        if missing_fks:
            # limiter l'affichage
            shown = ", ".join(missing_fks[:3])
            more = f" (+{len(missing_fks)-3})" if len(missing_fks) > 3 else ""
            return f"Clé étrangère manquante en cible: {shown}{more}"
    except Exception:
        pass

    # 3) Types potentiellement non supportés / délicats
    try:
        problematic = {"sql_variant", "hierarchyid", "geography", "geometry", "image", "text", "ntext"}
        prob_cols = []
        for col, typ in source_types.items():
            if typ in problematic and col in row_dict:
                prob_cols.append(f"{col}({typ})")
        if prob_cols:
            return "Type(s) délicat(s): " + ", ".join(prob_cols[:6]) + ("..." if len(prob_cols) > 6 else "")
    except Exception:
        pass

    return "Cause inconnue (nécessite diagnostic SQL plus poussé)"


@projet21_bp.route('/missing-rows', methods=['GET'])
def missing_rows():
    """
    Renvoie les lignes manquantes (dans la cible) pour une table donnée,
    avec toutes les colonnes + valeurs depuis la SOURCE.

    Query:
      - table: nom de table (obligatoire)
      - limit: nb lignes (défaut 50, max 200)
      - offset: pagination (défaut 0)
    """
    table_name = (request.args.get('table') or '').strip()
    if not table_name:
        return jsonify({'success': False, 'error': "Paramètre 'table' obligatoire"}), 400

    try:
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
    except Exception:
        return jsonify({'success': False, 'error': "Paramètres 'limit' et 'offset' doivent être des entiers"}), 400

    if limit < 1:
        limit = 1
    if limit > 200:
        limit = 200
    if offset < 0:
        offset = 0

    source_conn = None
    target_conn = None
    try:
        source_conn = get_connection(SOURCE_CONFIG, readonly=True)
        target_conn = get_connection(TARGET_CONFIG)
        source_cursor = source_conn.cursor()
        target_cursor = target_conn.cursor()

        pk_columns, missing_pks = _fetch_missing_pks_for_table(source_cursor, target_cursor, table_name)
        if not pk_columns:
            return jsonify({
                'success': False,
                'error': f"Impossible d'afficher les lignes manquantes: table '{table_name}' sans clé primaire détectée."
            }), 400

        total_missing = len(missing_pks)
        page_pks = missing_pks[offset: offset + limit]

        columns, rows = _select_rows_by_pks(source_cursor, table_name, pk_columns, page_pks)
        source_types = _get_column_types(source_cursor, table_name)

        # Ajouter la cause à chaque ligne
        for r in rows:
            r["__cause"] = _diagnose_missing_row(table_name, r, source_types, target_cursor)

        return jsonify({
            'success': True,
            'table': table_name,
            'pk_columns': pk_columns,
            'total_missing': total_missing,
            'offset': offset,
            'limit': limit,
            'columns': columns,
            'rows': rows,
        })
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'traceback': traceback.format_exc()}), 500
    finally:
        try:
            if source_conn:
                source_conn.close()
        except Exception:
            pass
        try:
            if target_conn:
                target_conn.close()
        except Exception:
            pass


def _diagnose_extra_row(table_name, row_dict, source_cursor, target_cursor):
    """
    Indique si une ligne "supplémentaire en cible" ressemble à un doublon (même index unique que la source),
    ou plutôt une donnée locale (ex: tables web).
    """
    try:
        pk_columns = get_primary_keys(source_cursor, table_name)
        unique_indexes = _get_unique_indexes(target_cursor, table_name)
        for idx_name, idx_cols in unique_indexes:
            values = []
            has_all = True
            has_null = False
            for c in idx_cols:
                if c not in row_dict:
                    has_all = False
                    break
                v = row_dict.get(c)
                if v is None:
                    has_null = True
                values.append(v)
            if not has_all or has_null:
                continue

            # Chercher une ligne en source avec la même clé unique.
            where_parts = [f"[{c}] = ?" for c in idx_cols]
            sql = f"SELECT TOP 1 {', '.join([f'[{pk}]' for pk in pk_columns])} FROM [{table_name}] WHERE " + " AND ".join(where_parts)
            source_cursor.execute(sql, tuple(values))
            src = source_cursor.fetchone()
            if src:
                return f"Doublon probable vs SOURCE (index unique {idx_name})"
        return "Présente en CIBLE mais absente en SOURCE (donnée locale / spécifique cible)"
    except Exception:
        return "Présente en CIBLE mais absente en SOURCE (diagnostic non disponible)"


@projet21_bp.route('/extra-rows', methods=['GET'])
def extra_rows():
    """
    Renvoie les lignes présentes en CIBLE mais absentes en SOURCE (PK-based),
    avec toutes les colonnes + valeurs depuis la CIBLE.

    Query:
      - table: nom de table (obligatoire)
      - limit: nb lignes (défaut 50, max 200)
      - offset: pagination (défaut 0)
    """
    table_name = (request.args.get('table') or '').strip()
    if not table_name:
        return jsonify({'success': False, 'error': "Paramètre 'table' obligatoire"}), 400

    try:
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
    except Exception:
        return jsonify({'success': False, 'error': "Paramètres 'limit' et 'offset' doivent être des entiers"}), 400

    if limit < 1:
        limit = 1
    if limit > 200:
        limit = 200
    if offset < 0:
        offset = 0

    source_conn = None
    target_conn = None
    try:
        source_conn = get_connection(SOURCE_CONFIG, readonly=True)
        target_conn = get_connection(TARGET_CONFIG)
        source_cursor = source_conn.cursor()
        target_cursor = target_conn.cursor()

        pk_columns, extra_pks = _fetch_extra_pks_for_table(source_cursor, target_cursor, table_name)
        if not pk_columns:
            return jsonify({
                'success': False,
                'error': f"Impossible d'afficher les lignes supplémentaires: table '{table_name}' sans clé primaire détectée."
            }), 400

        total_extra = len(extra_pks)
        page_pks = extra_pks[offset: offset + limit]

        columns, rows = _select_rows_by_pks(target_cursor, table_name, pk_columns, page_pks)

        # Ajouter une note à chaque ligne
        for r in rows:
            r["__note"] = _diagnose_extra_row(table_name, r, source_cursor, target_cursor)

        return jsonify({
            'success': True,
            'table': table_name,
            'pk_columns': pk_columns,
            'total_extra': total_extra,
            'offset': offset,
            'limit': limit,
            'columns': columns,
            'rows': rows,
        })
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'traceback': traceback.format_exc()}), 500
    finally:
        try:
            if source_conn:
                source_conn.close()
        except Exception:
            pass
        try:
            if target_conn:
                target_conn.close()
        except Exception:
            pass


def _select_row_by_pk(cursor, table_name, pk_columns, pk_value):
    """Retourne une ligne (dict) depuis la base cible pour une PK donnée, ou None."""
    if not pk_columns:
        return None
    if len(pk_columns) == 1:
        sql = f"SELECT * FROM [{table_name}] WHERE [{pk_columns[0]}] = ?"
        cursor.execute(sql, (pk_value,))
    else:
        # composite PK
        where = " AND ".join([f"[{c}] = ?" for c in pk_columns])
        sql = f"SELECT * FROM [{table_name}] WHERE {where}"
        cursor.execute(sql, tuple(pk_value))
    row = cursor.fetchone()
    if not row:
        return None
    columns = [desc[0] for desc in cursor.description]
    return {columns[i]: _normalize_pyodbc_value(row[i]) for i in range(len(columns))}

def _select_rows_by_pk_duplicates(target_cursor, table_name, pk_columns, duplicate_pks):
    """
    Récupère TOUTES les lignes dupliquées pour une liste de PK dupliquées.
    Retourne (columns, rows) où chaque row a aussi '__duplicate_count' indiquant combien de fois cette PK apparaît.
    """
    if not duplicate_pks or not pk_columns:
        return [], []
    
    all_rows = []
    columns = None
    
    for pk_info in duplicate_pks:
        pk_value, count = pk_info
        
        # Récupérer toutes les lignes avec cette PK
        if len(pk_columns) == 1:
            sql = f"SELECT * FROM [{table_name}] WHERE [{pk_columns[0]}] = ?"
            target_cursor.execute(sql, (pk_value,))
        else:
            where = " AND ".join([f"[{c}] = ?" for c in pk_columns])
            sql = f"SELECT * FROM [{table_name}] WHERE {where}"
            target_cursor.execute(sql, tuple(pk_value))
        
        rows_batch = target_cursor.fetchall()
        if not columns:
            columns = [desc[0] for desc in target_cursor.description]
        
        for row in rows_batch:
            row_dict = {columns[i]: _normalize_pyodbc_value(row[i]) for i in range(len(columns))}
            row_dict['__duplicate_count'] = count
            all_rows.append(row_dict)
    
    return columns, all_rows

@projet21_bp.route('/duplicate-pks', methods=['GET'])
def duplicate_pks():
    """
    Renvoie les lignes avec des PK dupliquées dans la CIBLE pour une table donnée.
    
    Query:
      - table: nom de table (obligatoire)
      - limit: nb PK dupliquées à traiter (défaut 20, max 100)
      - offset: pagination (défaut 0)
    """
    table_name = (request.args.get('table') or '').strip()
    if not table_name:
        return jsonify({'success': False, 'error': "Paramètre 'table' obligatoire"}), 400
    
    try:
        limit = int(request.args.get('limit', 20))
        offset = int(request.args.get('offset', 0))
    except Exception:
        return jsonify({'success': False, 'error': "Paramètres 'limit' et 'offset' doivent être des entiers"}), 400
    
    if limit < 1:
        limit = 1
    if limit > 100:
        limit = 100
    if offset < 0:
        offset = 0
    
    target_conn = None
    try:
        from routes.projet21_verification import check_duplicate_pks
        target_conn = get_connection(TARGET_CONFIG)
        target_cursor = target_conn.cursor()
        
        pk_columns = get_primary_keys(target_cursor, table_name)
        if not pk_columns:
            return jsonify({
                'success': False,
                'error': f"Impossible d'afficher les doublons: table '{table_name}' sans clé primaire détectée."
            }), 400
        
        duplicate_check = check_duplicate_pks(target_cursor, table_name)
        if 'error' in duplicate_check:
            return jsonify({
                'success': False,
                'error': f"Erreur lors de la vérification des doublons: {duplicate_check['error']}"
            }), 500
        
        if not duplicate_check.get('has_duplicates', False):
            return jsonify({
                'success': True,
                'table': table_name,
                'pk_columns': pk_columns,
                'total_duplicates': 0,
                'offset': 0,
                'limit': limit,
                'columns': [],
                'rows': []
            })
        
        duplicate_pks_list = duplicate_check.get('duplicate_pks', [])
        total_duplicates = len(duplicate_pks_list)
        page_duplicates = duplicate_pks_list[offset: offset + limit]
        
        columns, rows = _select_rows_by_pk_duplicates(target_cursor, table_name, pk_columns, page_duplicates)
        
        return jsonify({
            'success': True,
            'table': table_name,
            'pk_columns': pk_columns,
            'total_duplicates': total_duplicates,
            'total_duplicate_rows': duplicate_check.get('duplicate_rows', 0),
            'offset': offset,
            'limit': limit,
            'columns': columns,
            'rows': rows,
        })
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'traceback': traceback.format_exc()}), 500
    finally:
        try:
            if target_conn:
                target_conn.close()
        except Exception:
            pass


@projet21_bp.route('/compare-rows', methods=['GET'])
def compare_rows():
    """
    Génère un tableau de comparaison SOURCE vs CIBLE pour des PK "manquantes" (PK-based),
    en affichant sur une seule ligne les valeurs des 2 bases.

    - Si la PK n'existe pas en cible, on tente de trouver la ligne cible via un index unique (Option A).

    Query:
      - table: nom de table (obligatoire)
      - limit: nb lignes (défaut 10, max 50)
      - offset: pagination (défaut 0)
    """
    table_name = (request.args.get('table') or '').strip()
    if not table_name:
        return jsonify({'success': False, 'error': "Paramètre 'table' obligatoire"}), 400

    try:
        limit = int(request.args.get('limit', 10))
        offset = int(request.args.get('offset', 0))
    except Exception:
        return jsonify({'success': False, 'error': "Paramètres 'limit' et 'offset' doivent être des entiers"}), 400

    if limit < 1:
        limit = 1
    if limit > 50:
        limit = 50
    if offset < 0:
        offset = 0

    source_conn = None
    target_conn = None
    try:
        source_conn = get_connection(SOURCE_CONFIG, readonly=True)
        target_conn = get_connection(TARGET_CONFIG)
        source_cursor = source_conn.cursor()
        target_cursor = target_conn.cursor()

        pk_columns, missing_pks = _fetch_missing_pks_for_table(source_cursor, target_cursor, table_name)
        if not pk_columns:
            return jsonify({
                'success': False,
                'error': f"Impossible de comparer: table '{table_name}' sans clé primaire détectée."
            }), 400

        total_missing = len(missing_pks)
        page_pks = missing_pks[offset: offset + limit]

        # Charger les lignes source (celles dont la PK est manquante côté cible)
        source_cols, source_rows = _select_rows_by_pks(source_cursor, table_name, pk_columns, page_pks)

        # Colonnes cible (pour affichage stable)
        target_cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
        target_cols = [desc[0] for desc in target_cursor.description]

        unique_indexes = _get_unique_indexes(target_cursor, table_name)

        out_rows = []
        for src in source_rows:
            pk_val = src.get(pk_columns[0]) if len(pk_columns) == 1 else tuple(src.get(c) for c in pk_columns)

            # 1) Essayer match exact par PK
            tgt = _select_row_by_pk(target_cursor, table_name, pk_columns, pk_val)
            matched_by = {'type': 'pk', 'columns': pk_columns} if tgt else None

            # 2) Sinon, essayer match par index unique (Option A)
            if not tgt:
                for idx_name, idx_cols in unique_indexes:
                    # ne tester que si toutes les colonnes existent dans la source
                    vals = []
                    ok = True
                    for c in idx_cols:
                        if c not in src:
                            ok = False
                            break
                        vals.append(src.get(c))
                    if not ok:
                        continue

                    tgt_candidate = find_row_by_unique_index(target_cursor, table_name, idx_cols, vals)
                    if tgt_candidate:
                        tgt = tgt_candidate
                        matched_by = {'type': 'unique', 'index': idx_name, 'columns': idx_cols}
                        break

            # Construire une ligne "plate" : SRC_* et CIBLE_*
            flat = {
                '__table': table_name,
                '__pk': pk_val,
                '__matched_by': matched_by,
            }
            for c in source_cols:
                flat[f"SRC.{c}"] = src.get(c)
            for c in target_cols:
                flat[f"CIBLE.{c}"] = (tgt.get(c) if tgt else None)

            out_rows.append(flat)

        # Colonnes du tableau (une seule ligne = toutes les colonnes)
        columns = ['__table', '__pk', '__matched_by'] + [f"SRC.{c}" for c in source_cols] + [f"CIBLE.{c}" for c in target_cols]

        return jsonify({
            'success': True,
            'table': table_name,
            'pk_columns': pk_columns,
            'total_missing': total_missing,
            'offset': offset,
            'limit': limit,
            'columns': columns,
            'rows': out_rows,
        })
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'traceback': traceback.format_exc()}), 500
    finally:
        try:
            if source_conn:
                source_conn.close()
        except Exception:
            pass
        try:
            if target_conn:
                target_conn.close()
        except Exception:
            pass


@projet21_bp.route('/compare-rows.xlsx', methods=['GET'])
def compare_rows_xlsx():
    """
    Export Excel (.xlsx) du tableau de comparaison SOURCE vs CIBLE.
    Exporte TOUTES les lignes "manquantes PK" (PK-based) pour la table sélectionnée.

    Query:
      - table (obligatoire)
    """
    table_name = (request.args.get('table') or '').strip()
    if not table_name:
        return jsonify({'success': False, 'error': "Paramètre 'table' obligatoire"}), 400

    source_conn = None
    target_conn = None
    try:
        source_conn = get_connection(SOURCE_CONFIG, readonly=True)
        target_conn = get_connection(TARGET_CONFIG)
        source_cursor = source_conn.cursor()
        target_cursor = target_conn.cursor()

        pk_columns, missing_pks = _fetch_missing_pks_for_table(source_cursor, target_cursor, table_name)
        if not pk_columns:
            return jsonify({
                'success': False,
                'error': f"Impossible d'exporter: table '{table_name}' sans clé primaire détectée."
            }), 400

        # Export complet
        source_cols, source_rows = _select_rows_by_pks(source_cursor, table_name, pk_columns, list(missing_pks))

        target_cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
        target_cols = [desc[0] for desc in target_cursor.description]

        unique_indexes = _get_unique_indexes(target_cursor, table_name)

        out_rows = []
        for src in source_rows:
            pk_val = src.get(pk_columns[0]) if len(pk_columns) == 1 else tuple(src.get(c) for c in pk_columns)
            tgt = _select_row_by_pk(target_cursor, table_name, pk_columns, pk_val)
            matched_by = {'type': 'pk', 'columns': pk_columns} if tgt else None

            if not tgt:
                for idx_name, idx_cols in unique_indexes:
                    vals = []
                    ok = True
                    for c in idx_cols:
                        if c not in src:
                            ok = False
                            break
                        vals.append(src.get(c))
                    if not ok:
                        continue
                    tgt_candidate = find_row_by_unique_index(target_cursor, table_name, idx_cols, vals)
                    if tgt_candidate:
                        tgt = tgt_candidate
                        matched_by = {'type': 'unique', 'index': idx_name, 'columns': idx_cols}
                        break

            flat = {'__table': table_name, '__pk': pk_val, '__matched_by': matched_by}
            for c in source_cols:
                flat[f"SRC.{c}"] = src.get(c)
            for c in target_cols:
                flat[f"CIBLE.{c}"] = (tgt.get(c) if tgt else None)
            out_rows.append(flat)

        columns = ['__table', '__pk', '__matched_by'] + [f"SRC.{c}" for c in source_cols] + [f"CIBLE.{c}" for c in target_cols]

        df = pd.DataFrame(out_rows, columns=columns)
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Comparaison')

            # Mise en forme: colorer en rouge les valeurs différentes SRC vs CIBLE
            try:
                ws = writer.sheets.get('Comparaison')
                if ws is not None:
                    header_to_col = {}
                    for col_idx in range(1, ws.max_column + 1):
                        v = ws.cell(row=1, column=col_idx).value
                        if isinstance(v, str) and v:
                            header_to_col[v] = col_idx

                    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

                    # Comparer seulement les paires SRC.<col> et CIBLE.<col> basées sur les colonnes source
                    comparable_cols = [c for c in source_cols if f"SRC.{c}" in header_to_col and f"CIBLE.{c}" in header_to_col]

                    def _eq(a, b):
                        # Considérer None, "" et espaces comme équivalents
                        def _norm(x):
                            if x is None:
                                return ""
                            if isinstance(x, str):
                                return x.strip()
                            return x
                        na = _norm(a)
                        nb = _norm(b)
                        # Comparaison robuste par string si types différents
                        if type(na) != type(nb):
                            return str(na) == str(nb)
                        return na == nb

                    for row_idx in range(2, ws.max_row + 1):
                        for c in comparable_cols:
                            src_col_idx = header_to_col[f"SRC.{c}"]
                            tgt_col_idx = header_to_col[f"CIBLE.{c}"]
                            src_val = ws.cell(row=row_idx, column=src_col_idx).value
                            tgt_val = ws.cell(row=row_idx, column=tgt_col_idx).value
                            if not _eq(src_val, tgt_val):
                                ws.cell(row=row_idx, column=src_col_idx).fill = red_fill
                                ws.cell(row=row_idx, column=tgt_col_idx).fill = red_fill
            except Exception:
                # Ne pas bloquer l'export si la mise en forme échoue
                pass
        bio.seek(0)

        filename = f"comparaison_{table_name}_TOTAL_{len(missing_pks)}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'traceback': traceback.format_exc()}), 500
    finally:
        try:
            if source_conn:
                source_conn.close()
        except Exception:
            pass
        try:
            if target_conn:
                target_conn.close()
        except Exception:
            pass


@projet21_bp.route('/analyse-schema-papiers')
def analyse_schema_papiers():
    """
    Analyse le schéma de la base de données et génère un rapport détaillé
    listant toutes les tables ayant une relation directe ou indirecte
    avec PAPIERS_ARTICLES et PAPIERS_IMPRIMEURS
    """
    try:
        from collections import defaultdict, deque
        
        conn = get_connection(TARGET_CONFIG, readonly=True)
        cursor = conn.cursor()
        
        # Récupérer toutes les clés étrangères
        cursor.execute("""
            SELECT 
                fk.name AS FK_Name,
                tp.name AS Parent_Table,
                cp.name AS Parent_Column,
                tr.name AS Referenced_Table,
                cr.name AS Referenced_Column,
                fk.is_disabled,
                fk.is_not_trusted
            FROM sys.foreign_keys AS fk
            INNER JOIN sys.foreign_key_columns AS fkc ON fk.object_id = fkc.constraint_object_id
            INNER JOIN sys.tables AS tp ON fkc.parent_object_id = tp.object_id
            INNER JOIN sys.columns AS cp ON fkc.parent_object_id = cp.object_id 
                AND fkc.parent_column_id = cp.column_id
            INNER JOIN sys.tables AS tr ON fkc.referenced_object_id = tr.object_id
            INNER JOIN sys.columns AS cr ON fkc.referenced_object_id = cr.object_id 
                AND fkc.referenced_column_id = cr.column_id
            ORDER BY tp.name, fk.name, fkc.constraint_column_id
        """)
        
        fk_dict = defaultdict(list)
        reverse_graph = defaultdict(set)
        
        for row in cursor.fetchall():
            parent_table = row.Parent_Table
            referenced_table = row.Referenced_Table
            
            fk_info = {
                'fk_name': row.FK_Name,
                'parent_column': row.Parent_Column,
                'referenced_column': row.Referenced_Column,
                'is_disabled': row.is_disabled,
                'is_not_trusted': row.is_not_trusted
            }
            
            fk_dict[parent_table].append({
                'referenced_table': referenced_table,
                'details': fk_info
            })
            
            # Construire le graphe inversé
            reverse_graph[referenced_table].add(parent_table)
        
        # Tables de départ
        start_tables = ['PAPIERS_ARTICLES', 'PAPIERS_IMPRIMEURS']
        
        # Vérifier que les tables existent
        cursor.execute("""
            SELECT TABLE_NAME 
            FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_NAME IN (?, ?) AND TABLE_TYPE = 'BASE TABLE'
        """, start_tables[0], start_tables[1])
        
        existing_tables = [row.TABLE_NAME for row in cursor.fetchall()]
        missing_tables = [t for t in start_tables if t not in existing_tables]
        
        if missing_tables:
            return jsonify({
                'success': False,
                'error': f"Tables non trouvées: {', '.join(missing_tables)}"
            }), 404
        
        # Trouver toutes les tables liées (BFS)
        related_tables = set()
        visited = set()
        queue = deque()
        
        for table in start_tables:
            if table not in visited:
                visited.add(table)
                queue.append((table, 0))
                related_tables.add(table)
        
        # Parcourir les enfants et parents directs
        for table in start_tables:
            if table in reverse_graph:
                for child_table in reverse_graph[table]:
                    if child_table not in visited:
                        visited.add(child_table)
                        queue.append((child_table, 1))
                        related_tables.add(child_table)
            
            if table in fk_dict:
                for fk in fk_dict[table]:
                    parent_table = fk['referenced_table']
                    if parent_table not in visited:
                        visited.add(parent_table)
                        queue.append((parent_table, 1))
                        related_tables.add(parent_table)
        
        # BFS pour relations indirectes
        while queue:
            current_table, level = queue.popleft()
            
            if current_table in reverse_graph:
                for child_table in reverse_graph[current_table]:
                    if child_table not in visited:
                        visited.add(child_table)
                        queue.append((child_table, level + 1))
                        related_tables.add(child_table)
            
            if current_table in fk_dict:
                for fk in fk_dict[current_table]:
                    parent_table = fk['referenced_table']
                    if parent_table not in visited:
                        visited.add(parent_table)
                        queue.append((parent_table, level + 1))
                        related_tables.add(parent_table)
        
        # Construire le rapport détaillé
        report = {
            'generation_date': datetime.now().isoformat(),
            'start_tables': start_tables,
            'summary': {
                'total_related_tables': len(related_tables),
                'start_tables_count': len(start_tables),
                'related_tables_count': len(related_tables) - len(start_tables)
            },
            'tables': {}
        }
        
        # Séparer par type de relation
        direct_children = set()
        direct_parents = set()
        
        for table in related_tables:
            if table in start_tables:
                continue
            
            for start_table in start_tables:
                if start_table in reverse_graph and table in reverse_graph[start_table]:
                    direct_children.add(table)
                if start_table in fk_dict:
                    for fk in fk_dict[start_table]:
                        if fk['referenced_table'] == table:
                            direct_parents.add(table)
        
        report['summary']['direct_children_count'] = len(direct_children)
        report['summary']['direct_parents_count'] = len(direct_parents)
        report['summary']['indirect_related_count'] = len(related_tables) - len(start_tables) - len(direct_children) - len(direct_parents)
        
        # Détails par table
        for table in sorted(related_tables):
            relation_types = []
            if table in start_tables:
                relation_types.append('start')
            if table in direct_children:
                relation_types.append('direct_child')
            if table in direct_parents:
                relation_types.append('direct_parent')
            if table not in start_tables and table not in direct_children and table not in direct_parents:
                relation_types.append('indirect')
            
            # Relations FK
            fk_relations = []
            if table in fk_dict:
                for fk in fk_dict[table]:
                    fk_relations.append({
                        'type': 'references',
                        'referenced_table': fk['referenced_table'],
                        'fk_name': fk['details']['fk_name'],
                        'parent_column': fk['details']['parent_column'],
                        'referenced_column': fk['details']['referenced_column'],
                        'is_disabled': bool(fk['details']['is_disabled']),
                        'is_not_trusted': bool(fk['details']['is_not_trusted'])
                    })
            
            if table in reverse_graph:
                for child_table in reverse_graph[table]:
                    fk_relations.append({
                        'type': 'referenced_by',
                        'referencing_table': child_table
                    })
            
            # Compter les lignes (peut être lent)
            try:
                cursor.execute(f"SELECT COUNT(*) FROM [{table}]")
                row_count = cursor.fetchone()[0]
            except:
                row_count = None
            
            # Colonnes
            cursor.execute("""
                SELECT COLUMN_NAME, DATA_TYPE, IS_NULLABLE
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_NAME = ?
                ORDER BY ORDINAL_POSITION
            """, (table,))
            
            columns = []
            for col_row in cursor.fetchall():
                columns.append({
                    'name': col_row.COLUMN_NAME,
                    'type': col_row.DATA_TYPE,
                    'nullable': col_row.IS_NULLABLE == 'YES'
                })
            
            report['tables'][table] = {
                'relation_types': relation_types,
                'row_count': row_count,
                'column_count': len(columns),
                'columns': columns,
                'foreign_key_relations': fk_relations
            }
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'report': report
        })
        
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@projet21_bp.route('/auto-sync/config', methods=['GET'])
def get_auto_sync_config():
    """Récupère la configuration de synchronisation automatique"""
    try:
        from routes.projet21_auto_sync import load_auto_sync_config
        config = load_auto_sync_config()
        return jsonify({
            'success': True,
            'config': config
        })
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@projet21_bp.route('/auto-sync/config', methods=['POST'])
def set_auto_sync_config():
    """Active ou désactive la synchronisation automatique"""
    try:
        from routes.projet21_auto_sync import set_auto_sync_enabled
        from flask import current_app
        
        data = request.get_json()
        enabled = data.get('enabled', True)
        config = set_auto_sync_enabled(enabled)
        
        # Mettre à jour le scheduler APScheduler si disponible
        # Note: Si vous utilisez Task Scheduler Windows, cette partie ne s'applique pas
        scheduler = current_app.config.get('PROJET21_SCHEDULER')
        if scheduler:
            try:
                from apscheduler.triggers.cron import CronTrigger
                from routes.projet21_auto_sync import run_auto_sync_and_verify
                
                if enabled:
                    # Ajouter ou mettre à jour le job
                    scheduler.add_job(
                        func=run_auto_sync_and_verify,
                        trigger=CronTrigger(hour=5, minute=0),
                        id='projet21_auto_sync',
                        name='Synchronisation automatique Projet 21',
                        replace_existing=True
                    )
                else:
                    # Supprimer le job
                    try:
                        scheduler.remove_job('projet21_auto_sync')
                    except:
                        pass  # Le job n'existe peut-être pas
            except ImportError:
                pass  # APScheduler non disponible, utiliser Task Scheduler Windows
        
        # Message adapté selon le type de scheduler
        if scheduler:
            message = f'Synchronisation automatique ' + ('activée' if enabled else 'désactivée') + ' (APScheduler)'
        else:
            message = f'Synchronisation automatique ' + ('activée' if enabled else 'désactivée') + ' (Task Scheduler Windows)\n\n'
            if enabled:
                message += '⚠️ Important: Pour que la synchronisation fonctionne, assurez-vous que la tâche "Projet21 - Synchronisation Automatique" est activée dans le Planificateur de tâches Windows.'
            else:
                message += 'Pour désactiver complètement, désactivez également la tâche dans le Planificateur de tâches Windows.'
        
        return jsonify({
            'success': True,
            'config': config,
            'message': message
        })
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@projet21_bp.route('/auto-sync/status', methods=['GET'])
def get_auto_sync_status():
    """Vérifie l'état de la synchronisation automatique avec Task Scheduler Windows"""
    try:
        from routes.projet21_auto_sync import load_auto_sync_config, is_auto_sync_enabled, RESULTS_FILE
        from datetime import datetime, timedelta
        from pathlib import Path
        
        config = load_auto_sync_config()
        enabled = is_auto_sync_enabled()
        
        # Vérifier Task Scheduler (Windows)
        # On vérifie si le script de synchronisation existe
        task_script_path = Path(__file__).parent / 'projet21_auto_sync_task.py'
        task_scheduler_status = {
            'type': 'windows_task_scheduler',
            'script_exists': task_script_path.exists(),
            'script_path': str(task_script_path)
        }
        
        # Vérifier le dernier résultat
        from routes.projet21_auto_sync import load_last_verification_result
        last_result = load_last_verification_result()
        
        last_result_timestamp = None
        last_result_file_exists = False
        if RESULTS_FILE and RESULTS_FILE.exists():
            last_result_file_exists = True
            if last_result:
                last_result_timestamp = last_result.get('timestamp')
            else:
                # Essayer de lire directement le timestamp du fichier
                try:
                    import json
                    with open(RESULTS_FILE, 'r', encoding='utf-8') as f:
                        file_data = json.load(f)
                        last_result_timestamp = file_data.get('timestamp')
                except:
                    pass
        
        # Calculer le prochain exécution attendue (05:00 AM chaque jour)
        now = datetime.now()
        next_run = now.replace(hour=5, minute=0, second=0, microsecond=0)
        if now.hour >= 5:
            next_run += timedelta(days=1)
        
        # Calculer le temps depuis la dernière exécution
        hours_since_last_run = None
        if last_result_timestamp:
            try:
                last_run_dt = datetime.fromisoformat(last_result_timestamp.replace('Z', '+00:00'))
                if last_run_dt.tzinfo:
                    # Convertir en datetime naïf pour comparaison
                    last_run_dt = last_run_dt.replace(tzinfo=None)
                hours_since_last_run = (now - last_run_dt).total_seconds() / 3600
            except:
                pass
        
        status_info = {
            'enabled': enabled,
            'config': config,
            'scheduler': task_scheduler_status,
            'last_result_file_exists': last_result_file_exists,
            'last_result_timestamp': last_result_timestamp,
            'current_time': datetime.now().isoformat(),
            'next_scheduled_run': next_run.isoformat(),
            'hours_since_last_run': hours_since_last_run
        }
        
        return jsonify({
            'success': True,
            'status': status_info
        })
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@projet21_bp.route('/auto-sync/test-run', methods=['POST'])
def test_auto_sync_run():
    """Déclenche manuellement une synchronisation automatique pour test"""
    try:
        from routes.projet21_auto_sync import run_auto_sync_and_verify
        
        print("🧪 Test manuel de synchronisation automatique déclenché")
        run_auto_sync_and_verify()
        
        return jsonify({
            'success': True,
            'message': 'Synchronisation automatique de test exécutée. Vérifiez les résultats dans la section "Vérifier synchronisation automatique".'
        })
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@projet21_bp.route('/auto-sync/debug-file', methods=['GET'])
def debug_result_file():
    """Route de diagnostic pour vérifier le fichier de résultats"""
    try:
        from routes.projet21_auto_sync import RESULTS_FILE, RESULTS_DIR
        from pathlib import Path
        import json
        import os
        
        debug_info = {
            'results_dir_exists': RESULTS_DIR.exists() if RESULTS_DIR else False,
            'results_dir_path': str(RESULTS_DIR) if RESULTS_DIR else None,
            'results_file_exists': RESULTS_FILE.exists() if RESULTS_FILE else False,
            'results_file_path': str(RESULTS_FILE) if RESULTS_FILE else None,
            'current_working_dir': os.getcwd(),
            'script_file': str(Path(__file__).resolve()),
        }
        
        if RESULTS_FILE and RESULTS_FILE.exists():
            try:
                with open(RESULTS_FILE, 'r', encoding='utf-8') as f:
                    file_data = json.load(f)
                    debug_info['file_timestamp'] = file_data.get('timestamp')
                    debug_info['file_sync_success'] = file_data.get('sync_success')
                    debug_info['file_has_problems'] = file_data.get('has_problems')
                    debug_info['file_size'] = RESULTS_FILE.stat().st_size
                    debug_info['file_modified'] = os.path.getmtime(RESULTS_FILE)
            except Exception as e:
                debug_info['file_read_error'] = str(e)
        
        return jsonify({
            'success': True,
            'debug': debug_info
        })
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@projet21_bp.route('/auto-sync/last-result', methods=['GET'])
def get_last_auto_sync_result():
    """Récupère le dernier résultat de vérification automatique"""
    try:
        from routes.projet21_auto_sync import load_last_verification_result
        
        result = load_last_verification_result()
        if not result:
            return jsonify({
                'success': True,
                'has_result': False,
                'message': 'Aucun résultat de synchronisation automatique disponible'
            })
        
        # Formater les résultats comme la vérification manuelle
        results_data = result.get('results', {})
        
        # Calculer les totaux pour le résumé
        total_manquants = 0
        for item in results_data.get('ecarts_critiques', []):
            if len(item) >= 4:
                total_manquants += item[3]
        
        total_doublons = 0
        for item in results_data.get('doublons_pk', []):
            if len(item) >= 3:
                total_doublons += item[2]
        
        return jsonify({
            'success': True,
            'has_result': True,
            'timestamp': result.get('timestamp'),
            'sync_success': result.get('sync_success', True),
            'has_problems': result.get('has_problems', False),
            'output': results_data.get('output', ''),
            'summary': {
                'synchronisees': len(results_data.get('synchronisees', [])),
                'ecarts_critiques': len(results_data.get('ecarts_critiques', [])),
                'ecarts_normaux': len(results_data.get('ecarts_normaux', [])),
                'doublons_pk': len(results_data.get('doublons_pk', [])),
                'manquantes_cible': len(results_data.get('manquantes_cible', [])),
                'manquantes_source': len(results_data.get('manquantes_source', [])),
                'total_manquants': total_manquants,
                'total_doublons': total_doublons
            },
            'details': {
                'synchronisees': results_data.get('synchronisees', []),
                'ecarts_critiques': results_data.get('ecarts_critiques', []),
                'ecarts_normaux': results_data.get('ecarts_normaux', []),
                'doublons_pk': results_data.get('doublons_pk', []),
                'manquantes_cible': results_data.get('manquantes_cible', []),
                'manquantes_source': results_data.get('manquantes_source', [])
            }
        })
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500
